import os
import json
import sqlite3
import shutil
import stat
import sys
import threading
import uuid
import webbrowser
import zipfile
import ctypes
from datetime import datetime
from pathlib import Path, PurePosixPath
from urllib import error as urlerror
from urllib.parse import urlparse
from urllib import request as urlrequest
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo

try:
    from PIL import Image, ImageTk, ImageOps, ImageDraw
except Exception:
    Image = None
    ImageTk = None
    ImageOps = None
    ImageDraw = None

try:
    from tkinterdnd2 import TkinterDnD, DND_FILES
except Exception:
    TkinterDnD = None
    DND_FILES = None

APP_TITLE = "مدير مصاريف الزواج"
APP_SLUG = "MarriageExpensesManager"
APP_VERSION = "1.0.0"
APP_USER_MODEL_ID = "Abouels.MarriageExpensesManager"
SOURCE_DIR = Path(__file__).resolve().parent
IS_FROZEN = getattr(sys, "frozen", False)
BASE_DIR = Path(sys.executable).resolve().parent if IS_FROZEN else SOURCE_DIR
RESOURCE_DIR = Path(getattr(sys, "_MEIPASS", BASE_DIR)).resolve()
ASSETS_DIR = RESOURCE_DIR / "assets"
ICONS_DIR = ASSETS_DIR / "icons"
APP_ICON_PNG = ASSETS_DIR / "app_icon.png"
APP_ICON_ICO = ASSETS_DIR / "app_icon.ico"


def get_data_dir():
    override = os.environ.get("MARRIAGE_MANAGER_DATA_DIR")
    if override:
        return Path(override).expanduser().resolve()
    portable_markers = (BASE_DIR / "portable.flag", BASE_DIR / "app_data")
    if not IS_FROZEN or any(marker.exists() for marker in portable_markers):
        return BASE_DIR / "app_data"
    if sys.platform == "darwin":
        return Path.home() / "Library" / "Application Support" / APP_SLUG / "app_data"
    local_root = os.environ.get("LOCALAPPDATA") or os.environ.get("APPDATA")
    if local_root:
        return Path(local_root) / APP_SLUG / "app_data"
    return Path.home() / "AppData" / "Local" / APP_SLUG / "app_data"


DATA_DIR = get_data_dir()
INVOICES_DIR = DATA_DIR / "invoices"
LOAN_RECEIPTS_DIR = DATA_DIR / "loan_receipts"
DB_PATH = DATA_DIR / "apartment_costs.db"
EXPORTS_DIR = DATA_DIR / "exports"
CONFIG_PATH = DATA_DIR / "config.json"
TEST_FIRST_RUN_EVERY_LAUNCH = False
OWNER_NAME = "Eng. I. Abouelsaad"
OWNER_EMAIL = "i.abouelsaad9@gmail.com"
COPYRIGHT_NOTICE = f"Designed & Developed by {OWNER_NAME}"

OTHER_EXPENSES_TYPE = "مصاريف أخرى"
WEDDING_TYPE = "حفل الزفاف"
LEGACY_WEDDING_TYPE = "حفل الزواج"
TYPE_OPTIONS = ["فرش", "تشطيب", WEDDING_TYPE, OTHER_EXPENSES_TYPE]
FINISHING_OPTIONS = ["مشتريات", "مصنعيات", "عمولة مهندس", "غيره"]
DEFAULT_USER_NAME = "أنا"
DEFAULT_OTHER_PARTY_NAME = "الطرف الآخر"
LEGACY_OTHER_PARTY_NAME = "د. ايمن"
PAYMENT_METHODS = ["كاش", "إنستا باي"]
DEFAULT_PAYMENT_TARGETS = ["والد العروسة", "مهندس التشطيبات"]
LOAN_CURRENCIES = ["جنيه مصري", "يورو", "دولار", "ريال"]
LOAN_PAYMENT_METHODS = ["كاش", "تحويل بنكي", "إنستا باي"]
DEFAULT_ROOM_NAME = "مجمع"
DEFAULT_ROOM_OPTIONS = [DEFAULT_ROOM_NAME]
PARTY_ROLE_OPTIONS = ["العريس", "العروسة", "والد العروسة", "شريك آخر"]
SPLIT_MODE_EQUAL = "50 / 50"
SPLIT_MODE_PARTY1 = "الطرف الأول بالكامل"
SPLIT_MODE_PARTY2 = "الطرف الثاني بالكامل"
SPLIT_MODE_CUSTOM = "نسبة مخصصة"
SPLIT_MODE_OPTIONS = [SPLIT_MODE_EQUAL, SPLIT_MODE_PARTY1, SPLIT_MODE_PARTY2, SPLIT_MODE_CUSTOM]
ACCOUNTING_SINGLE = "نظام الطرف الواحد"
ACCOUNTING_EQUAL = "نظام 50/50"
ACCOUNTING_ITEMIZED = "نظام بنود على الطرفين"
ACCOUNTING_MODE_OPTIONS = [ACCOUNTING_SINGLE, ACCOUNTING_EQUAL, ACCOUNTING_ITEMIZED]
WEDDING_EQUAL = "حفل الزفاف بالمناصفة"
WEDDING_PARTY1 = "حفل الزفاف على الطرف الأول"
WEDDING_PARTY2 = "حفل الزفاف على الطرف الثاني"
WEDDING_POLICY_OPTIONS = [WEDDING_EQUAL, WEDDING_PARTY1, WEDDING_PARTY2]
SUPPORT_MESSAGE_TYPES = ["مشكلة تقنية", "طلب ميزة", "اقتراح تحسين", "ملاحظة عامة"]
SUPPORT_ENDPOINT_FILE = RESOURCE_DIR / "support_endpoint.txt"
THUMB_SIZE = (280, 220)
ALLOWED_ATTACHMENT_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp", ".pdf"}
MAX_ATTACHMENT_BYTES = 100 * 1024 * 1024
MAX_BACKUP_EXTRACTED_BYTES = 2 * 1024 * 1024 * 1024
MAX_SUPPORT_NAME_CHARS = 120
MAX_SUPPORT_EMAIL_CHARS = 254
MAX_SUPPORT_SUBJECT_CHARS = 160
MAX_SUPPORT_DETAILS_CHARS = 4000
# Theme constants
BG_APP = "#F3EEE3"
SURFACE = "#FFFDF7"
SIDEBAR_BG = "#E9E1D3"
PRIMARY = "#064E3B"
PRIMARY_HOVER = "#043D2F"
ACTIVE_BG = "#F4E8C8"
NAVY = "#123D2F"
NAVY_2 = "#174B3B"
TEXT = "#0F2F26"
TEXT_2 = "#7A705F"
FIELD_BG = "#F8F1E4"
INPUT_BG = "#FBF5E8"
INPUT_BORDER = "#BFA66A"
INPUT_FOCUS = "#064E3B"
BORDER = "#D8C7A4"
GOLD = "#B9954A"
GOLD_HOVER = "#D6B76F"
DANGER = "#A63A2B"
SUCCESS = "#0F7A55"
WARNING = "#B9954A"
FONT_FAMILY = "Segoe UI"
RTL_START = "\u202B"
RTL_END = "\u202C"


class SidebarButton(tk.Frame):
    def __init__(self, parent, text, command, app, **kwargs):
        super().__init__(parent, bg=SIDEBAR_BG, cursor="hand2", **kwargs)
        self.command = command
        self.app = app
        self.active = False
        self.hovered = False
        self.indicator = tk.Frame(self, bg=SIDEBAR_BG, width=3)
        self.indicator.pack(side="right", fill="y")
        self.label = tk.Label(
            self,
            text=text,
            bg=SIDEBAR_BG,
            fg=TEXT_2,
            font=(FONT_FAMILY, 10, "bold"),
            anchor="e",
            justify="right",
            padx=18,
            pady=7,
            cursor="hand2",
        )
        self.label.pack(side="right", fill="both", expand=True)
        for widget in (self, self.label, self.indicator):
            widget.bind("<Enter>", self._on_enter)
            widget.bind("<Leave>", self._on_leave)
            widget.bind("<Button-1>", self._on_click)

    def _paint(self, bg):
        self.configure(bg=bg)
        self.label.configure(bg=bg)
        if self.active:
            indicator_bg = PRIMARY
        elif self.hovered:
            indicator_bg = bg
        else:
            indicator_bg = bg
        self.indicator.configure(bg=indicator_bg)

    def _on_enter(self, _event=None):
        self.hovered = True
        self._paint(ACTIVE_BG if self.active else "#F5EFE3")

    def _on_leave(self, _event=None):
        self.hovered = False
        self._paint(ACTIVE_BG if self.active else SIDEBAR_BG)

    def _on_click(self, _event=None):
        self.app._set_sidebar_active(self)
        self.command()

    def set_active(self, active):
        self.active = active
        self.label.configure(
            fg=(PRIMARY if active else TEXT_2),
            font=(FONT_FAMILY, 10, "bold"),
        )
        self._paint(ACTIVE_BG if active else SIDEBAR_BG)


class NativeArabicEdit(tk.Frame):
    """Windows EDIT control with stable Arabic selection in RTL fields."""

    if os.name == "nt":
        _user32 = ctypes.windll.user32
        _gdi32 = ctypes.windll.gdi32
        _WS_CHILD = 0x40000000
        _WS_VISIBLE = 0x10000000
        _WS_TABSTOP = 0x00010000
        _ES_RIGHT = 0x0002
        _ES_MULTILINE = 0x0004
        _ES_AUTOHSCROLL = 0x0080
        _ES_AUTOVSCROLL = 0x0040
        _ES_WANTRETURN = 0x1000
        _WS_EX_CLIENTEDGE = 0x00000200
        _WM_SETFONT = 0x0030
        _WM_GETTEXT = 0x000D
        _WM_GETTEXTLENGTH = 0x000E
        _WM_SETTEXT = 0x000C
        _EM_SETSEL = 0x00B1
        _EM_GETSEL = 0x00B0
        _EM_SETREADONLY = 0x00CF
        _SWP_NOMOVE = 0x0002
        _SWP_NOZORDER = 0x0004

        _CreateWindowExW = _user32.CreateWindowExW
        _CreateWindowExW.argtypes = [
            ctypes.c_ulong,
            ctypes.c_wchar_p,
            ctypes.c_wchar_p,
            ctypes.c_ulong,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_int,
            ctypes.c_void_p,
            ctypes.c_void_p,
            ctypes.c_void_p,
            ctypes.c_void_p,
        ]
        _CreateWindowExW.restype = ctypes.c_void_p
        _SendMessageW = _user32.SendMessageW
        _SendMessageW.restype = ctypes.c_void_p

    def __init__(self, parent, textvariable=None, state="normal", font_size=10, multiline=False, rows=1, **kwargs):
        self.textvariable = textvariable
        self._state = state
        self._multiline = multiline
        self._syncing = False
        self._last_value = ""
        self._hwnd = None
        self._font = None
        pixel_height = max(42, int(font_size * 3.8 * max(1, rows)))
        super().__init__(parent, bg=INPUT_BG, height=pixel_height, highlightthickness=0, bd=0, **kwargs)
        self.pack_propagate(False)
        self.grid_propagate(False)
        self.bind("<Configure>", self._resize_child, add="+")
        self.bind("<Destroy>", self._stop_sync, add="+")
        self.after_idle(lambda: self._create_child(font_size))
        if self.textvariable is not None:
            self._trace_id = self.textvariable.trace_add("write", lambda *_: self._set_from_variable())
        else:
            self._trace_id = None
        self.after(120, self._poll_native_text)

    def _create_child(self, font_size):
        if self._hwnd or not self.winfo_exists():
            return
        style = self._WS_CHILD | self._WS_VISIBLE | self._WS_TABSTOP | self._ES_RIGHT
        if self._multiline:
            style |= self._ES_MULTILINE | self._ES_AUTOVSCROLL | self._ES_WANTRETURN
        else:
            style |= self._ES_AUTOHSCROLL
        self._hwnd = self._CreateWindowExW(
            0,
            "EDIT",
            "",
            style,
            0,
            0,
            max(1, self.winfo_width()),
            max(1, self.winfo_height()),
            self.winfo_id(),
            None,
            None,
            None,
        )
        self._font = self._gdi32.CreateFontW(-max(12, int(font_size * 1.55)), 0, 0, 0, 400, 0, 0, 0, 178, 0, 0, 5, 0, FONT_FAMILY)
        self._SendMessageW(self._hwnd, self._WM_SETFONT, self._font, 1)
        self._set_native_text(self.textvariable.get() if self.textvariable is not None else self._last_value)
        self.configure(state=self._state)
        self._resize_child()

    def _resize_child(self, event=None):
        if self._hwnd:
            inset_x = 10
            inset_y = 6
            self._user32.SetWindowPos(
                self._hwnd,
                0,
                inset_x,
                inset_y,
                max(1, self.winfo_width() - (inset_x * 2)),
                max(1, self.winfo_height() - (inset_y * 2)),
                self._SWP_NOZORDER,
            )

    def _stop_sync(self, event=None):
        self._hwnd = None

    def _native_text(self):
        if not self._hwnd:
            return self._last_value
        length = int(self._user32.GetWindowTextLengthW(self._hwnd))
        buf = ctypes.create_unicode_buffer(length + 1)
        self._user32.GetWindowTextW(self._hwnd, buf, length + 1)
        return buf.value

    def _set_native_text(self, value):
        value = value or ""
        self._last_value = value
        if self._hwnd:
            self._user32.SetWindowTextW(self._hwnd, value)

    def _set_from_variable(self):
        if self._syncing or self.textvariable is None:
            return
        value = self.textvariable.get()
        if value != self._last_value:
            self._set_native_text(value)

    def _poll_native_text(self):
        if not self.winfo_exists():
            return
        if self._hwnd:
            value = self._native_text()
            if value != self._last_value:
                self._last_value = value
                if self.textvariable is not None:
                    self._syncing = True
                    self.textvariable.set(value)
                    self._syncing = False
        self.after(120, self._poll_native_text)

    def get(self, *args):
        return self._native_text()

    def delete(self, start, end=None):
        self._set_native_text("")
        if self.textvariable is not None:
            self.textvariable.set("")

    def insert(self, index, value):
        current = self._native_text()
        if index in ("end", tk.END):
            pos = len(current)
        else:
            try:
                pos = int(str(index).split(".")[-1])
            except Exception:
                pos = 0
        self._set_native_text(current[:pos] + (value or "") + current[pos:])
        if self.textvariable is not None:
            self.textvariable.set(self._last_value)

    def selection_range(self, start, end):
        if self._hwnd:
            finish = len(self._native_text()) if end in ("end", tk.END) else int(end)
            self._SendMessageW(self._hwnd, self._EM_SETSEL, int(start), finish)

    def selection_present(self):
        return self.index("sel.first") != self.index("sel.last")

    def selection_clear(self):
        pos = self.index("insert")
        self.selection_range(pos, pos)

    def selection_own(self):
        return None

    def icursor(self, index):
        pos = len(self._native_text()) if index in ("end", tk.END) else int(index)
        if self._hwnd:
            start = ctypes.c_ulong()
            end = ctypes.c_ulong()
            self._SendMessageW(self._hwnd, self._EM_GETSEL, ctypes.byref(start), ctypes.byref(end))
            if int(start.value) != int(end.value) and pos == int(end.value):
                return
        self.selection_range(pos, pos)

    def index(self, index):
        if not self._hwnd:
            return 0
        start = ctypes.c_ulong()
        end = ctypes.c_ulong()
        self._SendMessageW(self._hwnd, self._EM_GETSEL, ctypes.byref(start), ctypes.byref(end))
        if index == "sel.first":
            return int(start.value)
        if index == "sel.last":
            return int(end.value)
        if index == "insert":
            return int(end.value)
        if index in ("end", tk.END):
            return len(self._native_text())
        return int(str(index).split(".")[-1])

    def focus_set(self):
        if self._hwnd:
            self._user32.SetFocus(self._hwnd)
        else:
            super().focus_set()

    focus_force = focus_set

    def configure(self, cnf=None, **kwargs):
        state = kwargs.pop("state", None)
        if state is not None:
            self._state = state
            if self._hwnd:
                self._SendMessageW(self._hwnd, self._EM_SETREADONLY, 1 if state == "readonly" else 0, 0)
        if cnf or kwargs:
            return super().configure(cnf, **kwargs)
        return super().configure()

    config = configure

    def cget(self, key):
        if key == "state":
            return self._state
        return super().cget(key)


class RoundedTextInput(tk.Frame):
    def __init__(
        self,
        parent,
        variable=None,
        state="normal",
        font_size=11,
        multiline=False,
        rows=1,
        surface_bg=None,
        input_bg=None,
        border_color=None,
        focus_color=None,
        radius=15,
        native=True,
        height=None,
    ):
        self._height = height or max(54, int(font_size * 4.4 * max(1, rows)))
        self._surface_bg = surface_bg or SURFACE
        self._input_bg = input_bg or INPUT_BG
        self._border_color = border_color or INPUT_BORDER
        self._focus_color = focus_color or INPUT_FOCUS
        self._radius = radius
        super().__init__(parent, bg=self._surface_bg, height=self._height)
        self.pack_propagate(False)
        self.grid_propagate(False)
        self._focused = False
        self.canvas = tk.Canvas(self, bg=self._surface_bg, highlightthickness=0, bd=0, height=self._height)
        self.canvas.pack(fill="both", expand=True)
        self.inner = tk.Frame(self.canvas, bg=self._input_bg)
        self._multiline = multiline
        if native and os.name == "nt":
            self.child = NativeArabicEdit(self.inner, textvariable=variable, state=state, font_size=font_size, multiline=multiline, rows=rows)
            self.child.configure(bg=self._input_bg)
        elif multiline:
            self.child = tk.Text(
                self.inner,
                height=rows,
                wrap="word",
                font=(FONT_FAMILY, font_size),
                bg=self._input_bg,
                fg="#111827",
                insertbackground="#111827",
                relief="flat",
                bd=0,
                highlightthickness=0,
                padx=8,
                pady=8,
            )
            self.child.tag_configure("rtl", justify="right", rmargin=6, lmargin1=6, lmargin2=6)
            self.child.tag_add("rtl", "1.0", "end")
        else:
            self.child = tk.Entry(
                self.inner,
                textvariable=variable,
                justify="right",
                state=state,
                font=(FONT_FAMILY, font_size),
                bg=self._input_bg,
                fg="#111827",
                insertbackground="#111827",
                readonlybackground=self._input_bg,
                disabledbackground=self._input_bg,
                disabledforeground="#111827",
                relief="flat",
                bd=0,
                highlightthickness=0,
                exportselection=False,
            )
        self.child.pack(fill="both", expand=True)
        self._inner_window = self.canvas.create_window(14, 6, anchor="nw", window=self.inner)
        self.bind("<Configure>", self._draw, add="+")
        for widget in (self, self.canvas, self.inner, self.child):
            widget.bind("<FocusIn>", self._focus_in, add="+")
            widget.bind("<FocusOut>", self._focus_out, add="+")

    def _rounded_rect(self, x1, y1, x2, y2, r, **kwargs):
        points = [
            x1 + r, y1, x2 - r, y1, x2, y1, x2, y1 + r,
            x2, y2 - r, x2, y2, x2 - r, y2, x1 + r, y2,
            x1, y2, x1, y2 - r, x1, y1 + r, x1, y1,
        ]
        return self.canvas.create_polygon(points, smooth=True, splinesteps=16, **kwargs)

    def _draw(self, _event=None):
        width = max(1, self.winfo_width())
        height = max(1, self.winfo_height())
        border = self._focus_color if self._focused else self._border_color
        self.canvas.delete("shape")
        self._rounded_rect(1, 1, width - 2, height - 2, self._radius, fill=self._input_bg, outline=border, width=1, tags="shape")
        self.canvas.coords(self._inner_window, 14, 6)
        self.canvas.itemconfigure(self._inner_window, width=max(1, width - 28), height=max(1, height - 12))

    def _focus_in(self, _event=None):
        self._focused = True
        self._draw()

    def _focus_out(self, _event=None):
        self._focused = False
        self._draw()

    def get(self, *args):
        if isinstance(self.child, tk.Text):
            return self.child.get("1.0", "end-1c")
        return self.child.get(*args)

    def delete(self, *args):
        if isinstance(self.child, tk.Text):
            return self.child.delete("1.0", "end")
        return self.child.delete(*args)

    def insert(self, *args):
        if isinstance(self.child, tk.Text):
            value = args[-1] if args else ""
            self.child.insert("1.0", value)
            self.child.tag_add("rtl", "1.0", "end")
            return None
        return self.child.insert(*args)

    def selection_range(self, *args):
        return self.child.selection_range(*args)

    def selection_present(self, *args):
        return self.child.selection_present(*args)

    def selection_clear(self, *args):
        return self.child.selection_clear(*args)

    def selection_own(self, *args):
        return self.child.selection_own(*args)

    def icursor(self, *args):
        return self.child.icursor(*args)

    def index(self, *args):
        return self.child.index(*args)

    def focus_set(self):
        return self.child.focus_set()

    def configure(self, cnf=None, **kwargs):
        child_keys = {"state"}
        child_kwargs = {key: kwargs.pop(key) for key in list(kwargs) if key in child_keys}
        result = super().configure(cnf, **kwargs)
        if child_kwargs:
            self.child.configure(**child_kwargs)
        return result

    config = configure

    def cget(self, key):
        if key == "state":
            return self.child.cget(key)
        return super().cget(key)


class RoundedComboInput(tk.Frame):
    def __init__(
        self,
        parent,
        variable=None,
        values=(),
        font_size=11,
        surface_bg=None,
        input_bg=None,
        border_color=None,
        focus_color=None,
        radius=15,
        height=54,
    ):
        self._surface_bg = surface_bg or SURFACE
        self._input_bg = input_bg or INPUT_BG
        self._border_color = border_color or INPUT_BORDER
        self._focus_color = focus_color or INPUT_FOCUS
        self._radius = radius
        self._height = height
        super().__init__(parent, bg=self._surface_bg, height=self._height)
        self.pack_propagate(False)
        self.grid_propagate(False)
        self._focused = False
        self.canvas = tk.Canvas(self, bg=self._surface_bg, highlightthickness=0, bd=0, height=self._height)
        self.canvas.pack(fill="both", expand=True)
        self.inner = tk.Frame(self.canvas, bg=self._input_bg)
        self.child = ttk.Combobox(self.inner, textvariable=variable, values=values, justify="right", state="readonly", font=(FONT_FAMILY, font_size), style="TCombobox")
        self.child.pack(fill="both", expand=True)
        self._inner_window = self.canvas.create_window(14, 9, anchor="nw", window=self.inner)
        self.bind("<Configure>", self._draw, add="+")
        for widget in (self, self.canvas, self.inner, self.child):
            widget.bind("<FocusIn>", self._focus_in, add="+")
            widget.bind("<FocusOut>", self._focus_out, add="+")

    def _rounded_rect(self, x1, y1, x2, y2, r, **kwargs):
        points = [
            x1 + r, y1, x2 - r, y1, x2, y1, x2, y1 + r,
            x2, y2 - r, x2, y2, x2 - r, y2, x1 + r, y2,
            x1, y2, x1, y2 - r, x1, y1 + r, x1, y1,
        ]
        return self.canvas.create_polygon(points, smooth=True, splinesteps=16, **kwargs)

    def _draw(self, _event=None):
        width = max(1, self.winfo_width())
        height = max(1, self.winfo_height())
        border = self._focus_color if self._focused else self._border_color
        self.canvas.delete("shape")
        self._rounded_rect(1, 1, width - 2, height - 2, self._radius, fill=self._input_bg, outline=border, width=1, tags="shape")
        self.canvas.coords(self._inner_window, 14, 9)
        self.canvas.itemconfigure(self._inner_window, width=max(1, width - 28), height=max(1, height - 18))

    def _focus_in(self, _event=None):
        self._focused = True
        self._draw()

    def _focus_out(self, _event=None):
        self._focused = False
        self._draw()

    def get(self):
        return self.child.get()

    def set(self, value):
        return self.child.set(value)

    def bind(self, sequence=None, func=None, add=None):
        self.child.bind(sequence, func, add)
        return super().bind(sequence, func, add)

    def configure(self, cnf=None, **kwargs):
        child_keys = {"values", "state", "textvariable"}
        child_kwargs = {key: kwargs.pop(key) for key in list(kwargs) if key in child_keys}
        result = super().configure(cnf, **kwargs)
        if child_kwargs:
            self.child.configure(**child_kwargs)
        return result

    config = configure

    def cget(self, key):
        if key in {"values", "state", "textvariable"}:
            return self.child.cget(key)
        return super().cget(key)

    def __setitem__(self, key, value):
        self.child[key] = value

    def __getitem__(self, key):
        return self.child[key]


class RoundedOptionInput(tk.Frame):
    def __init__(
        self,
        parent,
        variable=None,
        values=(),
        font_size=11,
        surface_bg=None,
        input_bg=None,
        border_color=None,
        focus_color=None,
        radius=10,
        height=46,
    ):
        self.variable = variable or tk.StringVar(value="")
        self.values = list(values)
        self._surface_bg = surface_bg or SURFACE
        self._input_bg = input_bg or INPUT_BG
        self._border_color = border_color or INPUT_BORDER
        self._focus_color = focus_color or INPUT_FOCUS
        self._radius = radius
        self._height = height
        self._focused = False
        super().__init__(parent, bg=self._surface_bg, height=height)
        self.pack_propagate(False)
        self.grid_propagate(False)
        self.canvas = tk.Canvas(self, bg=self._surface_bg, height=height, highlightthickness=0, bd=0, cursor="hand2")
        self.canvas.pack(fill="both", expand=True)
        self.label = tk.Label(
            self.canvas,
            textvariable=self.variable,
            bg=self._input_bg,
            fg=TEXT,
            font=(FONT_FAMILY, font_size),
            anchor="e",
            justify="right",
            padx=12,
            cursor="hand2",
        )
        self._label_window = self.canvas.create_window(0, 0, anchor="nw", window=self.label)
        self.menu = tk.Menu(self, tearoff=0, font=(FONT_FAMILY, font_size))
        self._build_menu()
        self.bind("<Configure>", self._draw, add="+")
        for widget in (self, self.canvas, self.label):
            widget.bind("<Button-1>", self._show_menu, add="+")
            widget.bind("<FocusIn>", self._focus_in, add="+")
            widget.bind("<FocusOut>", self._focus_out, add="+")

    def _rounded_rect(self, x1, y1, x2, y2, r, **kwargs):
        points = [
            x1 + r, y1, x2 - r, y1, x2, y1, x2, y1 + r,
            x2, y2 - r, x2, y2, x2 - r, y2, x1 + r, y2,
            x1, y2, x1, y2 - r, x1, y1 + r, x1, y1,
        ]
        return self.canvas.create_polygon(points, smooth=True, splinesteps=16, **kwargs)

    def _draw(self, _event=None):
        width = max(1, self.winfo_width())
        height = max(1, self.winfo_height())
        border = self._focus_color if self._focused else self._border_color
        self.canvas.delete("shape")
        self._rounded_rect(1, 1, width - 2, height - 2, self._radius, fill=self._input_bg, outline=border, width=1, tags="shape")
        arrow_x = 22
        arrow_y = height // 2
        self.canvas.create_polygon(
            arrow_x - 6, arrow_y - 3,
            arrow_x + 6, arrow_y - 3,
            arrow_x, arrow_y + 5,
            fill=TEXT,
            outline="",
            tags="shape",
        )
        self.canvas.coords(self._label_window, 42, 7)
        self.canvas.itemconfigure(self._label_window, width=max(1, width - 58), height=max(1, height - 14))
        self.canvas.tag_lower("shape")

    def _build_menu(self):
        self.menu.delete(0, "end")
        for value in self.values:
            self.menu.add_command(label=value, command=lambda item=value: self.set(item))

    def _show_menu(self, event=None):
        self.focus_set()
        self._focused = True
        self._draw()
        try:
            self.menu.tk_popup(self.winfo_rootx(), self.winfo_rooty() + self.winfo_height())
        finally:
            self.menu.grab_release()

    def _focus_in(self, _event=None):
        self._focused = True
        self._draw()

    def _focus_out(self, _event=None):
        self._focused = False
        self._draw()

    def get(self):
        return self.variable.get()

    def set(self, value):
        self.variable.set(value)

    def configure(self, cnf=None, **kwargs):
        if cnf:
            kwargs.update(cnf)
        if "values" in kwargs:
            self.values = list(kwargs.pop("values"))
            self._build_menu()
        if "textvariable" in kwargs:
            self.variable = kwargs.pop("textvariable")
            self.label.configure(textvariable=self.variable)
        if "state" in kwargs:
            kwargs.pop("state")
        if kwargs:
            return super().configure(**kwargs)
        return super().configure()

    config = configure

    def cget(self, key):
        if key == "values":
            return tuple(self.values)
        if key == "textvariable":
            return self.variable
        return super().cget(key)

    def __setitem__(self, key, value):
        self.configure(**{key: value})

    def __getitem__(self, key):
        return self.cget(key)


class RoundedCanvasButton(tk.Canvas):
    def __init__(
        self,
        parent,
        text,
        command,
        bg="#064E3B",
        hover_bg=None,
        fg="#FFFFFF",
        width=112,
        height=36,
        radius=8,
        font_size=10,
        shadow=True,
    ):
        self.command = command
        self._label_text = text
        self.btn_bg = bg
        self.hover_bg = hover_bg or bg
        self.fg = fg
        self.radius = radius
        self.shadow = shadow
        super().__init__(
            parent,
            width=width,
            height=height + (3 if shadow else 0),
            bg=parent.cget("bg"),
            highlightthickness=0,
            bd=0,
            cursor="hand2",
        )
        self.font = (FONT_FAMILY, font_size, "bold")
        self.bind("<Configure>", lambda _e: self._draw(self.btn_bg))
        self.bind("<Enter>", lambda _e: self._draw(self.hover_bg))
        self.bind("<Leave>", lambda _e: self._draw(self.btn_bg))
        self.bind("<Button-1>", lambda _e: self.command())

    def _rounded_rect(self, x1, y1, x2, y2, r, **kwargs):
        points = [
            x1 + r, y1, x2 - r, y1, x2, y1, x2, y1 + r,
            x2, y2 - r, x2, y2, x2 - r, y2, x1 + r, y2,
            x1, y2, x1, y2 - r, x1, y1 + r, x1, y1,
        ]
        return self.create_polygon(points, smooth=True, splinesteps=16, **kwargs)

    def _draw(self, fill):
        self.delete("all")
        width = max(1, self.winfo_width())
        height = max(1, self.winfo_height())
        button_h = height - (3 if self.shadow else 0)
        if self.shadow:
            self._rounded_rect(3, 5, width - 1, button_h + 2, self.radius, fill="#D7D7D7", outline="", tags="shadow")
        self._rounded_rect(1, 1, width - 3, button_h, self.radius, fill=fill, outline="", tags="button")
        self.create_text((width - 3) / 2, button_h / 2, text=self._text(), fill=self.fg, font=self.font, anchor="center")

    def _text(self):
        return self._label_text

    def configure(self, cnf=None, **kwargs):
        if "text" in kwargs:
            self._label_text = kwargs.pop("text")
            self._draw(self.btn_bg)
        if kwargs or cnf:
            return super().configure(cnf, **kwargs)
        return super().configure()

    config = configure


class RoundedCardFrame(tk.Frame):
    def __init__(self, parent, bg="#FFFFFF", border="#E5E7EB", shadow="#E6E6E6", radius=12, padx=20, pady=20, auto_size=True):
        super().__init__(parent, bg=parent.cget("bg"))
        self.card_bg = bg
        self.border = border
        self.shadow = shadow
        self.radius = radius
        self.padx = padx
        self.pady = pady
        self.auto_size = auto_size
        self.canvas = tk.Canvas(self, bg=parent.cget("bg"), highlightthickness=0, bd=0)
        self.canvas.pack(fill="both", expand=True)
        self.body = tk.Frame(self.canvas, bg=bg, padx=padx, pady=pady)
        self._window = self.canvas.create_window(0, 0, anchor="nw", window=self.body)
        self.bind("<Configure>", self._draw, add="+")
        if auto_size:
            self.body.bind("<Configure>", self._sync_requested_size, add="+")

    def _rounded_rect(self, x1, y1, x2, y2, r, **kwargs):
        points = [
            x1 + r, y1, x2 - r, y1, x2, y1, x2, y1 + r,
            x2, y2 - r, x2, y2, x2 - r, y2, x1 + r, y2,
            x1, y2, x1, y2 - r, x1, y1 + r, x1, y1,
        ]
        return self.canvas.create_polygon(points, smooth=True, splinesteps=20, **kwargs)

    def _draw(self, _event=None):
        width = max(1, self.winfo_width())
        height = max(1, self.winfo_height())
        self.canvas.delete("shape")
        self._rounded_rect(4, 5, width - 3, height - 3, self.radius, fill=self.shadow, outline="", tags="shape")
        self._rounded_rect(1, 1, width - 6, height - 7, self.radius, fill=self.card_bg, outline=self.border, width=1, tags="shape")
        self.canvas.coords(self._window, 3, 3)
        self.canvas.itemconfigure(self._window, width=max(1, width - 10), height=max(1, height - 12))

    def _sync_requested_size(self, _event=None):
        if not self.auto_size:
            return
        try:
            req_width = self.body.winfo_reqwidth() + 14
            req_height = self.body.winfo_reqheight() + 16
            self.canvas.configure(width=req_width, height=req_height)
        except Exception:
            pass

    def refresh_size(self):
        self.update_idletasks()
        self._sync_requested_size()
        self._draw()


def load_config():
    ensure_dirs()
    if CONFIG_PATH.exists():
        try:
            return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_config(data):
    ensure_dirs()
    temp_path = CONFIG_PATH.with_name(f"{CONFIG_PATH.name}.{uuid.uuid4().hex}.tmp")
    try:
        temp_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        os.replace(temp_path, CONFIG_PATH)
    finally:
        if temp_path.exists():
            try:
                temp_path.unlink()
            except OSError:
                pass


def load_room_options():
    data = load_config()
    vals = data.get("room_options") or []
    vals = normalize_room_options(vals)
    if vals:
        return vals
    return list(DEFAULT_ROOM_OPTIONS)


def normalize_room_options(options):
    vals = []
    seen = set()
    for option in [DEFAULT_ROOM_NAME, *(options or [])]:
        val = str(option).strip()
        if val and val not in seen:
            vals.append(val)
            seen.add(val)
    return vals


def save_room_options(options):
    data = load_config()
    data["room_options"] = normalize_room_options(options)
    save_config(data)


def load_party_names():
    data = load_config()
    user_name = str(data.get("user_name") or "").strip()
    other_party_name = str(data.get("other_party_name") or "").strip()
    if user_name and other_party_name:
        return user_name, other_party_name
    return None, None


def save_party_names(user_name, other_party_name):
    data = load_config()
    data["user_name"] = user_name.strip()
    data["other_party_name"] = other_party_name.strip()
    save_config(data)


def load_project_name():
    data = load_config()
    return str(data.get("project_name") or APP_TITLE).strip() or APP_TITLE


def save_project_name(project_name):
    data = load_config()
    data["project_name"] = (project_name or APP_TITLE).strip() or APP_TITLE
    save_config(data)


def load_support_web_app_url():
    env_url = os.environ.get("MARRIAGE_MANAGER_SUPPORT_URL", "").strip()
    if env_url:
        return env_url
    if SUPPORT_ENDPOINT_FILE.exists():
        try:
            return SUPPORT_ENDPOINT_FILE.read_text(encoding="utf-8").strip()
        except Exception:
            return ""
    data = load_config()
    return str(data.get("support_web_app_url") or "").strip()


def is_valid_support_web_app_url(url):
    if not url:
        return False
    try:
        parsed = urlparse(url.strip())
    except Exception:
        return False
    host = (parsed.hostname or "").lower()
    return (
        parsed.scheme == "https"
        and host == "script.google.com"
        and parsed.path.startswith("/macros/s/")
        and parsed.path.endswith("/exec")
    )


def validate_support_text_lengths(payload):
    limits = {
        "name": MAX_SUPPORT_NAME_CHARS,
        "email": MAX_SUPPORT_EMAIL_CHARS,
        "subject": MAX_SUPPORT_SUBJECT_CHARS,
        "details": MAX_SUPPORT_DETAILS_CHARS,
    }
    for key, limit in limits.items():
        if len(str(payload.get(key) or "")) > limit:
            return False, key, limit
    return True, "", 0


def is_supported_attachment(path):
    return Path(path).suffix.lower() in ALLOWED_ATTACHMENT_EXTENSIONS


def is_attachment_size_allowed(path):
    try:
        return Path(path).stat().st_size <= MAX_ATTACHMENT_BYTES
    except OSError:
        return False


def iter_backup_source_files(output_path):
    output_path = Path(output_path).resolve()
    exports_dir = EXPORTS_DIR.resolve()
    for path in DATA_DIR.rglob("*"):
        if not path.is_file():
            continue
        resolved = path.resolve()
        if resolved == output_path:
            continue
        try:
            resolved.relative_to(exports_dir)
            continue
        except ValueError:
            pass
        yield path


def _safe_backup_member_target(base_dir, member_name):
    normalized = str(member_name or "").replace("\\", "/")
    parsed = PurePosixPath(normalized)
    parts = parsed.parts
    if parsed.is_absolute() or not parts or parts[0] != "app_data":
        raise ValueError("ملف النسخة الاحتياطية يحتوي على مسار غير صالح.")
    if any(part in {"", ".", ".."} or ":" in part for part in parts):
        raise ValueError("ملف النسخة الاحتياطية يحتوي على مسار غير صالح.")
    target = (base_dir / Path(*parts)).resolve()
    try:
        target.relative_to(base_dir.resolve())
    except ValueError as exc:
        raise ValueError("ملف النسخة الاحتياطية يحتوي على مسار غير صالح.") from exc
    return target


def extract_backup_safely(zip_path, destination_dir):
    destination_dir = Path(destination_dir)
    destination_dir.mkdir(parents=True, exist_ok=True)
    total_size = 0
    with zipfile.ZipFile(zip_path, "r") as zf:
        infos = zf.infolist()
        if not infos:
            raise ValueError("ملف النسخة الاحتياطية فارغ.")
        for info in infos:
            mode = info.external_attr >> 16
            if stat.S_ISLNK(mode):
                raise ValueError("ملف النسخة الاحتياطية يحتوي على رابط غير مسموح.")
            total_size += int(info.file_size or 0)
            if total_size > MAX_BACKUP_EXTRACTED_BYTES:
                raise ValueError("حجم النسخة الاحتياطية بعد فك الضغط كبير جدًا.")
            target = _safe_backup_member_target(destination_dir, info.filename)
            if info.is_dir():
                target.mkdir(parents=True, exist_ok=True)
                continue
            target.parent.mkdir(parents=True, exist_ok=True)
            with zf.open(info, "r") as source, target.open("wb") as dest:
                shutil.copyfileobj(source, dest)


def normalize_text_options(options, defaults):
    vals = []
    seen = set()
    for option in [*(defaults or []), *(options or [])]:
        val = str(option).strip()
        if val and val not in seen:
            vals.append(val)
            seen.add(val)
    return vals


def load_payment_targets():
    data = load_config()
    saved = data.get("payment_targets")
    if saved:
        return normalize_text_options(saved, [])
    return list(DEFAULT_PAYMENT_TARGETS)


def save_payment_targets(targets):
    data = load_config()
    data["payment_targets"] = normalize_text_options(targets, [])
    if not data["payment_targets"]:
        data["payment_targets"] = list(DEFAULT_PAYMENT_TARGETS)
    save_config(data)


def load_base_currency():
    data = load_config()
    currency = str(data.get("base_currency") or LOAN_CURRENCIES[0]).strip()
    return currency if currency in LOAN_CURRENCIES else LOAN_CURRENCIES[0]


def save_base_currency(currency):
    data = load_config()
    data["base_currency"] = currency if currency in LOAN_CURRENCIES else LOAN_CURRENCIES[0]
    save_config(data)


def load_party_roles():
    data = load_config()
    return (
        str(data.get("party1_role") or PARTY_ROLE_OPTIONS[0]).strip(),
        str(data.get("party2_role") or PARTY_ROLE_OPTIONS[2]).strip(),
    )


def save_party_roles(party1_role, party2_role):
    data = load_config()
    data["party1_role"] = (party1_role or PARTY_ROLE_OPTIONS[0]).strip()
    data["party2_role"] = (party2_role or PARTY_ROLE_OPTIONS[2]).strip()
    save_config(data)


def normalize_split(mode, party1_percent, party2_percent):
    if mode == SPLIT_MODE_PARTY1:
        return mode, 100.0, 0.0
    if mode == SPLIT_MODE_PARTY2:
        return mode, 0.0, 100.0
    if mode == SPLIT_MODE_CUSTOM:
        p1 = max(0.0, min(100.0, parse_float(party1_percent, 50.0)))
        p2 = max(0.0, min(100.0, parse_float(party2_percent, 100.0 - p1)))
        total = p1 + p2
        if total <= 0:
            return SPLIT_MODE_EQUAL, 50.0, 50.0
        return mode, (p1 / total) * 100.0, (p2 / total) * 100.0
    return SPLIT_MODE_EQUAL, 50.0, 50.0


def validate_custom_split_total(mode, party1_percent, party2_percent):
    if mode != SPLIT_MODE_CUSTOM:
        return True, ""
    p1 = parse_float(party1_percent, None)
    p2 = parse_float(party2_percent, None)
    if p1 is None or p2 is None:
        return False, "من فضلك أدخل نسب صحيحة للطرفين."
    if p1 < 0 or p2 < 0:
        return False, "النسب لا يمكن أن تكون أقل من صفر."
    if abs((p1 + p2) - 100.0) > 0.001:
        return False, "مجموع نسب الطرفين لازم يساوي 100%."
    return True, ""


def load_split_settings():
    data = load_config()
    mode = str(data.get("split_mode") or SPLIT_MODE_EQUAL).strip()
    return normalize_split(mode, data.get("party1_percent", 50), data.get("party2_percent", 50))


def save_split_settings(mode, party1_percent, party2_percent):
    mode, p1, p2 = normalize_split(mode, party1_percent, party2_percent)
    data = load_config()
    data["split_mode"] = mode
    data["party1_percent"] = p1
    data["party2_percent"] = p2
    save_config(data)


def load_accounting_settings():
    data = load_config()
    mode = str(data.get("accounting_mode") or ACCOUNTING_EQUAL).strip()
    if mode not in ACCOUNTING_MODE_OPTIONS:
        mode = ACCOUNTING_EQUAL
    wedding_policy = str(data.get("wedding_policy") or WEDDING_EQUAL).strip()
    wedding_policy = wedding_policy.replace(LEGACY_WEDDING_TYPE, WEDDING_TYPE)
    if wedding_policy not in WEDDING_POLICY_OPTIONS:
        wedding_policy = WEDDING_EQUAL
    return mode, wedding_policy


def save_accounting_settings(mode, wedding_policy):
    data = load_config()
    data["accounting_mode"] = mode if mode in ACCOUNTING_MODE_OPTIONS else ACCOUNTING_EQUAL
    data["wedding_policy"] = wedding_policy if wedding_policy in WEDDING_POLICY_OPTIONS else WEDDING_EQUAL
    save_config(data)




def ensure_dirs():
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    INVOICES_DIR.mkdir(parents=True, exist_ok=True)
    LOAN_RECEIPTS_DIR.mkdir(parents=True, exist_ok=True)
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)


def parse_float(value, default=0.0):
    try:
        return float(str(value).replace(",", ".").strip())
    except Exception:
        return default


def sanitize_filename(text: str) -> str:
    text = (text or "مرفق").strip()
    bad = '<>:"/\\|?*\n\r\t'
    for ch in bad:
        text = text.replace(ch, " ")
    text = " ".join(text.split())
    text = text[:80].strip(" .")
    return text or "مرفق"


def open_file(path: str):
    if not path or not os.path.exists(path):
        messagebox.showwarning("تنبيه", "الملف غير موجود أو تم نقله.")
        return
    webbrowser.open(Path(path).resolve().as_uri())


def is_image(path: str):
    return Path(path).suffix.lower() in {".png", ".jpg", ".jpeg", ".bmp", ".gif", ".webp"}


def is_pdf(path: str):
    return Path(path).suffix.lower() == ".pdf"


def copy_invoice_file(src: str, item_name: str):
    if not src:
        return "", ""
    if not is_supported_attachment(src) or not is_attachment_size_allowed(src):
        return "", ""
    ensure_dirs()
    ext = Path(src).suffix.lower()
    base_name = sanitize_filename(item_name)
    candidate = INVOICES_DIR / f"{base_name}{ext}"
    counter = 2
    while candidate.exists():
        candidate = INVOICES_DIR / f"{base_name}_{counter}{ext}"
        counter += 1
    shutil.copy2(src, candidate)
    return str(candidate), candidate.name


def copy_loan_receipt_file(src: str, loan_name: str):
    if not src:
        return "", ""
    if not is_supported_attachment(src) or not is_attachment_size_allowed(src):
        return "", ""
    ensure_dirs()
    ext = Path(src).suffix.lower()
    base_name = sanitize_filename(f"سداد {loan_name}")
    candidate = LOAN_RECEIPTS_DIR / f"{base_name}{ext}"
    counter = 2
    while candidate.exists():
        candidate = LOAN_RECEIPTS_DIR / f"{base_name}_{counter}{ext}"
        counter += 1
    shutil.copy2(src, candidate)
    return str(candidate), candidate.name


def format_invoice_names(invoices):
    if not invoices:
        return "لا يوجد ملف مرفق"
    names = [inv.get("original_name") or Path(inv.get("path", "")).name for inv in invoices]
    if len(names) == 1:
        return names[0]
    preview = "، ".join(names[:3])
    if len(names) > 3:
        preview += f" +{len(names) - 3}"
    return f"{len(names)} مرفقات: {preview}"


def maybe_import_pdf_libs():
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.pdfgen import canvas
        import arabic_reshaper
        from bidi.algorithm import get_display
        return {
            "A4": A4,
            "pdfmetrics": pdfmetrics,
            "TTFont": TTFont,
            "canvas": canvas,
            "arabic_reshaper": arabic_reshaper,
            "get_display": get_display,
        }
    except Exception:
        return None


def shape_ar(text, pdf_libs):
    if not text:
        return ""
    try:
        return pdf_libs["get_display"](pdf_libs["arabic_reshaper"].reshape(str(text)))
    except Exception:
        return str(text)


def find_arabic_font_path():
    candidates = [
        Path("C:/Windows/Fonts/arial.ttf"),
        Path("C:/Windows/Fonts/tahoma.ttf"),
        Path("C:/Windows/Fonts/tradbdo.ttf"),
        Path("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"),
    ]
    for p in candidates:
        if p.exists():
            return str(p)
    return None


def connect_db():
    ensure_dirs()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys=ON")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            main_type TEXT NOT NULL,
            finish_type TEXT,
            payer TEXT,
            shared_split INTEGER DEFAULT 1,
            item_name TEXT NOT NULL,
            room TEXT,
            quantity REAL DEFAULT 1,
            unit_price REAL DEFAULT 0,
            total REAL DEFAULT 0,
            notes TEXT,
            invoice_path TEXT,
            invoice_original_name TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS receipts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            amount REAL NOT NULL,
            received_at TEXT NOT NULL,
            method TEXT NOT NULL,
            notes TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS record_invoices (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            record_id INTEGER NOT NULL,
            invoice_path TEXT NOT NULL,
            invoice_original_name TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(record_id) REFERENCES records(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS advances (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            amount REAL NOT NULL,
            direction TEXT NOT NULL,
            reason TEXT NOT NULL,
            advance_at TEXT NOT NULL,
            notes TEXT,
            status TEXT NOT NULL DEFAULT 'غير مسددة',
            settled_at TEXT,
            settlement_method TEXT,
            settlement_notes TEXT,
            created_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS loans (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            amount REAL NOT NULL,
            currency TEXT NOT NULL,
            loan_at TEXT NOT NULL,
            notes TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS loan_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            loan_id INTEGER NOT NULL,
            amount REAL NOT NULL,
            currency TEXT NOT NULL,
            paid_at TEXT NOT NULL,
            method TEXT NOT NULL,
            receipt_path TEXT,
            receipt_original_name TEXT,
            notes TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(loan_id) REFERENCES loans(id) ON DELETE CASCADE
        )
        """
    )

    loan_columns = {row[1] for row in conn.execute("PRAGMA table_info(loans)").fetchall()}
    if "title" not in loan_columns:
        conn.execute("ALTER TABLE loans ADD COLUMN title TEXT")
        conn.execute("UPDATE loans SET title=COALESCE(lender_name, 'سلفة') WHERE title IS NULL OR title=''")
    if "loan_at" not in loan_columns:
        conn.execute("ALTER TABLE loans ADD COLUMN loan_at TEXT")
        conn.execute("UPDATE loans SET loan_at=COALESCE(borrowed_at, created_at, datetime('now')) WHERE loan_at IS NULL OR loan_at=''")
    if "updated_at" not in loan_columns:
        conn.execute("ALTER TABLE loans ADD COLUMN updated_at TEXT")
        conn.execute("UPDATE loans SET updated_at=COALESCE(created_at, loan_at, datetime('now')) WHERE updated_at IS NULL OR updated_at=''")

    payment_columns = {row[1] for row in conn.execute("PRAGMA table_info(loan_payments)").fetchall()}
    if "receipt_path" not in payment_columns:
        conn.execute("ALTER TABLE loan_payments ADD COLUMN receipt_path TEXT")
    if "receipt_original_name" not in payment_columns:
        conn.execute("ALTER TABLE loan_payments ADD COLUMN receipt_original_name TEXT")

    loan_columns = {row[1] for row in conn.execute("PRAGMA table_info(loans)").fetchall()}
    if {"lender_name", "method", "borrowed_at"} & loan_columns:
        legacy_name = f"loans_legacy_{datetime.now().strftime('%Y%m%d%H%M%S')}"
        conn.commit()
        conn.execute("PRAGMA foreign_keys=OFF")
        conn.execute(f"ALTER TABLE loans RENAME TO {legacy_name}")
        conn.execute(
            """
            CREATE TABLE loans (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                amount REAL NOT NULL,
                currency TEXT NOT NULL,
                loan_at TEXT NOT NULL,
                notes TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            f"""
            INSERT INTO loans (id, title, amount, currency, loan_at, notes, created_at, updated_at)
            SELECT
                id,
                COALESCE(NULLIF(title, ''), NULLIF(lender_name, ''), 'سلفة'),
                amount,
                currency,
                COALESCE(NULLIF(loan_at, ''), NULLIF(borrowed_at, ''), created_at, datetime('now')),
                notes,
                COALESCE(created_at, datetime('now')),
                COALESCE(updated_at, created_at, loan_at, borrowed_at, datetime('now'))
            FROM {legacy_name}
            """
        )
        conn.execute(f"DROP TABLE {legacy_name}")
        conn.execute("PRAGMA foreign_keys=ON")

    columns = {row[1] for row in conn.execute("PRAGMA table_info(records)").fetchall()}
    if "payer" not in columns:
        conn.execute("ALTER TABLE records ADD COLUMN payer TEXT")
        conn.execute("UPDATE records SET payer=? WHERE payer IS NULL", (DEFAULT_USER_NAME,))
    conn.execute("UPDATE records SET payer=? WHERE payer=?", (LEGACY_OTHER_PARTY_NAME, DEFAULT_OTHER_PARTY_NAME))
    if "shared_split" not in columns:
        conn.execute("ALTER TABLE records ADD COLUMN shared_split INTEGER DEFAULT 1")
        conn.execute("UPDATE records SET shared_split=1 WHERE shared_split IS NULL")
    conn.execute("UPDATE records SET main_type=? WHERE main_type=?", (WEDDING_TYPE, LEGACY_WEDDING_TYPE))
    conn.execute(
        """
        INSERT INTO record_invoices (record_id, invoice_path, invoice_original_name, created_at)
        SELECT r.id, r.invoice_path, r.invoice_original_name, COALESCE(r.created_at, r.updated_at, datetime('now'))
        FROM records r
        WHERE r.invoice_path IS NOT NULL
          AND r.invoice_path <> ''
          AND NOT EXISTS (
              SELECT 1 FROM record_invoices ri
              WHERE ri.record_id = r.id AND ri.invoice_path = r.invoice_path
          )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_records_updated_at ON records(updated_at DESC, id DESC)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_records_type_room ON records(main_type, room)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_record_invoices_record_id ON record_invoices(record_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_receipts_received_at ON receipts(received_at DESC, id DESC)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_advances_advance_at ON advances(advance_at DESC, id DESC)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_advances_direction ON advances(direction)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_loans_loan_at ON loans(loan_at DESC, id DESC)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_loan_payments_loan_id ON loan_payments(loan_id, paid_at DESC, id DESC)")
    conn.commit()
    return conn


class ApartmentCostsApp:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title(APP_TITLE)
        self._set_app_icon()
        self.root.geometry("1580x960")
        self.root.minsize(980, 620)
        try:
            self.root.state("zoomed")
        except Exception:
            pass
        self.root.after(120, self._ensure_main_window_visible)

        self.selected_invoice_sources = []
        self.editing_record_id = None
        self.invoice_removed_in_edit = False
        self.current_preview_image = None
        self.current_selected_record_id = None
        self.search_rows = []
        self.search_window = None
        self.rooms_window = None
        self.receipt_form_window = None
        self.receipts_view_window = None
        self.advance_form_window = None
        self.advances_view_window = None
        self.loan_form_window = None
        self.loans_view_window = None
        self.loan_payment_window = None
        self.support_window = None
        self.editing_receipt_id = None

        self.style = ttk.Style()
        try:
            self.style.theme_use("clam")
        except Exception:
            pass
        self._configure_style()
        self.user_name, self.other_party_name = self._load_or_prompt_party_names()
        if not self._root_exists():
            return
        self.party1_role, self.party2_role = load_party_roles()
        self.base_currency = load_base_currency()
        self.split_mode, self.party1_percent, self.party2_percent = load_split_settings()
        self.accounting_mode, self.wedding_policy = load_accounting_settings()
        self.payer_options = [self.user_name, self.other_party_name]
        self._ensure_initial_room_options()
        if not self._root_exists():
            return
        self.project_name = load_project_name()
        self.payment_targets = load_payment_targets()
        self.settings_window = None
        self.root.title(self.project_name)
        self.conn = connect_db()
        self._migrate_payer_names()
        self._build_variables()
        self._build_ui()
        self._install_text_editing_support()
        self.refresh_room_option_widgets()
        self._bind_events()
        self.update_form_by_type()
        self.refresh_summary()

    def _root_exists(self):
        try:
            return bool(self.root.winfo_exists())
        except tk.TclError:
            return False

    def _set_app_icon(self):
        self._app_icon_photo = None
        try:
            if APP_ICON_ICO.exists():
                self.root.iconbitmap(default=str(APP_ICON_ICO))
        except Exception:
            pass
        try:
            if APP_ICON_PNG.exists():
                self._app_icon_photo = tk.PhotoImage(file=str(APP_ICON_PNG))
                self.root.iconphoto(True, self._app_icon_photo)
        except Exception:
            pass

    def _ensure_main_window_visible(self):
        try:
            self.root.update_idletasks()
            self.root.state("zoomed")
        except Exception:
            try:
                sw = self.root.winfo_screenwidth()
                sh = self.root.winfo_screenheight()
                self.root.geometry(f"{sw}x{max(860, sh - 80)}+0+0")
            except Exception:
                pass

    def _fit_toplevel(self, window, min_w, min_h, width_ratio=0.62, height_ratio=0.72):
        try:
            window.update_idletasks()
            sw = window.winfo_screenwidth()
            sh = window.winfo_screenheight()
            max_w = min(sw - 80, max(min_w, int(sw * width_ratio)))
            max_h = min(sh - 90, max(min_h, int(sh * height_ratio)))
            requested_w = window.winfo_reqwidth() + 34
            requested_h = window.winfo_reqheight() + 44
            width = min(max(min_w, requested_w), max_w)
            height = min(max(min_h, requested_h), max_h)
            x = max(0, (sw - width) // 2)
            y = max(0, (sh - height) // 2)
            window.geometry(f"{width}x{height}+{x}+{y}")
            window.minsize(min(min_w, width), min(min_h, height))
        except Exception:
            try:
                window.geometry(f"{min_w}x{min_h}")
                window.minsize(min_w, min_h)
            except Exception:
                pass

    def _reveal_window(self, window):
        window.deiconify()
        try:
            window.update_idletasks()
        except Exception:
            pass
        window.lift()
        window.focus_force()

    def _configure_style(self):
        self.root.configure(bg=BG_APP)
        self.style.configure("TFrame", background=BG_APP)
        self.style.configure("Sidebar.TFrame", background=SURFACE)
        self.style.configure("Card.TFrame", background=SURFACE, relief="flat", borderwidth=0)
        self.style.configure("Section.TFrame", background=SURFACE)
        self.style.configure("TLabel", background=BG_APP, font=(FONT_FAMILY, 10), foreground=TEXT)
        self.style.configure("Header.TLabel", background=BG_APP, font=(FONT_FAMILY, 17, "bold"), foreground=PRIMARY)
        self.style.configure("HeaderSub.TLabel", background=BG_APP, font=(FONT_FAMILY, 9), foreground=TEXT_2)
        self.style.configure("SubHeader.TLabel", background=SURFACE, font=(FONT_FAMILY, 13, "bold"), foreground=PRIMARY)
        self.style.configure("SectionTitle.TLabel", background=SURFACE, font=(FONT_FAMILY, 10, "bold"), foreground=PRIMARY)
        self.style.configure("Field.TLabel", background=SURFACE, font=(FONT_FAMILY, 11, "bold"), foreground=NAVY_2)
        self.style.configure("TLabelframe", background=SURFACE, borderwidth=1, bordercolor=BORDER, relief="flat")
        self.style.configure("TLabelframe.Label", background=SURFACE, font=(FONT_FAMILY, 11, "bold"), foreground=PRIMARY)
        self.style.configure("TEntry", font=(FONT_FAMILY, 11), fieldbackground=INPUT_BG, foreground=TEXT, bordercolor=INPUT_BORDER, lightcolor=INPUT_BORDER, darkcolor=INPUT_BORDER, insertcolor=TEXT, relief="flat", borderwidth=1, padding=(12, 9))
        self.style.configure("Focus.TEntry", font=(FONT_FAMILY, 11), fieldbackground=INPUT_BG, foreground=TEXT, bordercolor=INPUT_FOCUS, lightcolor=INPUT_FOCUS, darkcolor=INPUT_FOCUS, insertcolor=TEXT, relief="flat", borderwidth=1, padding=(12, 9))
        self.style.map("TEntry", fieldbackground=[("readonly", "#EFE7D8")], foreground=[("readonly", TEXT)])
        self.style.configure("TCombobox", font=(FONT_FAMILY, 11), fieldbackground=INPUT_BG, foreground=TEXT, bordercolor=INPUT_BORDER, lightcolor=INPUT_BORDER, darkcolor=INPUT_BORDER, arrowsize=16, relief="flat", borderwidth=1, padding=(12, 9, 34, 9))
        self.style.configure("Focus.TCombobox", font=(FONT_FAMILY, 11), fieldbackground=INPUT_BG, foreground=TEXT, bordercolor=INPUT_FOCUS, lightcolor=INPUT_FOCUS, darkcolor=INPUT_FOCUS, arrowsize=16, relief="flat", borderwidth=1, padding=(12, 9, 34, 9))
        self.style.map("TCombobox", fieldbackground=[("readonly", INPUT_BG)], selectbackground=[("readonly", INPUT_BG)], selectforeground=[("readonly", TEXT)])
        self.style.configure("TButton", font=(FONT_FAMILY, 9, "bold"), padding=(12, 8), background=SURFACE, foreground=TEXT, borderwidth=1, bordercolor=BORDER)
        self.style.map("TButton", background=[("active", "#F5EFE3")])
        self.style.configure("Accent.TButton", font=(FONT_FAMILY, 10, "bold"), padding=(14, 9), background=PRIMARY, foreground="#FFF7D6", borderwidth=0)
        self.style.map("Accent.TButton", background=[("active", PRIMARY_HOVER)])
        self.style.configure("Primary.TButton", font=(FONT_FAMILY, 10, "bold"), padding=(14, 9), background=PRIMARY, foreground="#FFF7D6", borderwidth=0)
        self.style.map("Primary.TButton", background=[("active", PRIMARY_HOVER)])
        self.style.configure("Secondary.TButton", font=(FONT_FAMILY, 9, "bold"), padding=(12, 8), background=INPUT_BG, foreground=TEXT, borderwidth=1, bordercolor=INPUT_BORDER)
        self.style.map("Secondary.TButton", background=[("active", "#F1E8D7")])
        self.style.configure("Warning.TButton", font=(FONT_FAMILY, 9, "bold"), padding=(12, 8), background=DANGER, foreground="#FFFFFF", borderwidth=0)
        self.style.map("Warning.TButton", background=[("active", "#7F2D22")])
        self.style.configure("Info.TButton", font=(FONT_FAMILY, 9, "bold"), padding=(12, 8), background=GOLD, foreground=PRIMARY, borderwidth=0)
        self.style.map("Info.TButton", background=[("active", GOLD_HOVER)])
        self.style.configure("Treeview", font=(FONT_FAMILY, 10), rowheight=36, borderwidth=0, relief="flat", background=SURFACE, fieldbackground=SURFACE, foreground=TEXT)
        self.style.configure("Treeview.Heading", font=(FONT_FAMILY, 10, "bold"), background=PRIMARY, foreground="#FFFFFF", relief="flat")
        self.style.map("Treeview", background=[("selected", "#E7D9B9")], foreground=[("selected", TEXT)])
        self.style.configure(
            "Vertical.TScrollbar",
            gripcount=0,
            background="#CDBB94",
            darkcolor="#CDBB94",
            lightcolor="#CDBB94",
            troughcolor=SURFACE,
            bordercolor=SURFACE,
            arrowcolor=PRIMARY,
            relief="flat",
            width=12,
            arrowsize=10,
        )
        self.style.map("Vertical.TScrollbar", background=[("active", GOLD), ("pressed", GOLD)])
        self.style.configure("Secondary.TCheckbutton", background=SURFACE, foreground=TEXT, font=(FONT_FAMILY, 10))

    def _configure_tree_zebra(self, tree):
        tree.tag_configure("oddrow", background=SURFACE)
        tree.tag_configure("evenrow", background="#EEF4EA")

    def _add_corner_flourish(self, parent, size=34):
        for corner, anchors in (
            ("tl", ("left", "top")),
            ("br", ("right", "bottom")),
        ):
            canvas = tk.Canvas(parent, width=size, height=size, bg=SURFACE, highlightthickness=0, bd=0)
            if corner == "tl":
                canvas.place(x=5, y=5)
                canvas.create_arc(2, 2, size * 1.45, size * 1.45, start=90, extent=90, outline=GOLD, width=1)
                canvas.create_arc(9, 9, size * 1.08, size * 1.08, start=90, extent=90, outline="#D6C38E", width=1)
                canvas.create_line(0, 9, size * 0.45, 9, fill=GOLD)
                canvas.create_line(9, 0, 9, size * 0.45, fill=GOLD)
            else:
                canvas.place(relx=1, rely=1, x=-size - 5, y=-size - 5)
                canvas.create_arc(-size * 0.45, -size * 0.45, size - 2, size - 2, start=270, extent=90, outline=GOLD, width=1)
                canvas.create_arc(-size * 0.08, -size * 0.08, size - 9, size - 9, start=270, extent=90, outline="#D6C38E", width=1)
                canvas.create_line(size * 0.55, size - 9, size, size - 9, fill=GOLD)
                canvas.create_line(size - 9, size * 0.55, size - 9, size, fill=GOLD)

    def _make_luxury_card(self, parent, padx=18, pady=18):
        card = tk.Frame(parent, bg="#D8D0C2", bd=0)
        surface = tk.Frame(card, bg=SURFACE, bd=0, highlightbackground=BORDER, highlightthickness=1)
        surface.place(x=0, y=0, relwidth=1, relheight=1, width=-5, height=-6)
        inner = tk.Frame(surface, bg=SURFACE, padx=padx, pady=pady)
        inner.pack(fill="both", expand=True)
        self._add_corner_flourish(inner, size=28)
        return card, inner

    def _popup_card(self, window, title, padx=18, pady=18):
        outer = tk.Frame(window, bg=BG_APP, padx=18, pady=16)
        outer.pack(fill="both", expand=True)
        card, inner = self._make_luxury_card(outer, padx=padx, pady=pady)
        card.pack(fill="both", expand=True)
        if title:
            tk.Label(inner, text=title, bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 15, "bold"), anchor="e", justify="right").pack(fill="x", pady=(0, 12))
        return inner

    def _action_button(self, parent, text, command, kind="primary", width=126, height=38):
        palette = {
            "primary": (PRIMARY, PRIMARY_HOVER, "#FFF7D6"),
            "secondary": (INPUT_BG, "#F1E8D7", TEXT),
            "danger": (DANGER, "#7F2D22", "#FFFFFF"),
            "gold": (GOLD, GOLD_HOVER, PRIMARY),
        }
        bg, hover, fg = palette.get(kind, palette["primary"])
        return RoundedCanvasButton(parent, text, command, bg=bg, hover_bg=hover, fg=fg, width=width, height=max(46, height), radius=8, font_size=10, shadow=True)

    def _attach_vertical_scrollbar(self, widget, parent, manager="pack", **layout):
        scrollbar = ttk.Scrollbar(parent, orient="vertical", command=widget.yview)
        state = {"visible": False}

        def show():
            if state["visible"]:
                return
            if manager == "grid":
                scrollbar.grid(**layout)
            else:
                scrollbar.pack(**layout)
            state["visible"] = True

        def hide():
            if not state["visible"]:
                return
            if manager == "grid":
                scrollbar.grid_remove()
            else:
                scrollbar.pack_forget()
            state["visible"] = False

        def update(first, last):
            scrollbar.set(first, last)
            try:
                needs_scroll = float(first) > 0.0 or float(last) < 1.0
            except Exception:
                needs_scroll = True
            if needs_scroll:
                show()
            else:
                hide()

        widget.configure(yscrollcommand=update)
        parent.after_idle(lambda: update(*widget.yview()))
        return scrollbar

    def create_metric_card(self, parent, title, value, icon):
        card = tk.Frame(parent, bg=SURFACE, bd=0, highlightbackground="#E5E7EB", highlightthickness=1)
        body = tk.Frame(card, bg=SURFACE, padx=25, pady=20)
        body.pack(fill="both", expand=True)
        tk.Label(
            body,
            text=f"{icon}  {title}",
            bg=SURFACE,
            fg="#6B7280",
            font=(FONT_FAMILY, 10),
            anchor="e",
            justify="right",
        ).pack(fill="x")
        value_kwargs = {"textvariable": value} if isinstance(value, tk.Variable) else {"text": value}
        value_label = tk.Label(
            body,
            bg=SURFACE,
            fg=TEXT,
            font=(FONT_FAMILY, 24, "bold"),
            anchor="e",
            justify="right",
            **value_kwargs,
        )
        value_label.pack(fill="x", pady=(4, 0))
        card.value_label = value_label
        return card

    def _draw_metric_icon(self, canvas, key):
        bg = PRIMARY if key in {"balance", "person_cost"} else "#F0EEE8"
        fg = "#FFFFFF" if key in {"balance", "person_cost"} else PRIMARY
        canvas.create_oval(2, 2, 38, 38, fill=bg, outline="")
        if key == "all":
            canvas.create_rectangle(9, 13, 31, 27, outline=fg, width=2)
            canvas.create_oval(13, 17, 17, 21, outline=fg, width=2)
            canvas.create_line(23, 16, 29, 16, fill=fg, width=2)
            canvas.create_line(23, 22, 29, 22, fill=fg, width=2)
        elif key == "furn":
            canvas.create_rectangle(9, 18, 31, 27, fill=fg, outline=fg)
            canvas.create_rectangle(12, 13, 28, 19, outline=fg, width=2)
            canvas.create_line(12, 27, 12, 31, fill=fg, width=2)
            canvas.create_line(28, 27, 28, 31, fill=fg, width=2)
        elif key == "wedding":
            canvas.create_line(13, 12, 28, 27, fill=fg, width=2)
            canvas.create_line(28, 12, 13, 27, fill=fg, width=2)
            canvas.create_oval(10, 9, 16, 15, outline=fg, width=2)
            canvas.create_oval(25, 9, 31, 15, outline=fg, width=2)
        elif key == "finish":
            canvas.create_rectangle(10, 18, 30, 27, outline=fg, width=2)
            canvas.create_rectangle(13, 13, 27, 18, fill=fg, outline=fg)
        elif key == "count":
            for y in (12, 19, 26):
                canvas.create_line(12, y, 30, y, fill=fg, width=2)
                canvas.create_oval(7, y - 2, 10, y + 1, fill=fg, outline=fg)
        elif key == "receipts":
            canvas.create_rectangle(9, 18, 26, 27, outline=fg, width=2)
            canvas.create_oval(12, 27, 17, 32, outline=fg, width=2)
            canvas.create_oval(25, 27, 30, 32, outline=fg, width=2)
            canvas.create_polygon(26, 18, 32, 23, 26, 23, fill=fg, outline=fg)
        elif key == "engineer":
            canvas.create_arc(10, 13, 30, 31, start=0, extent=180, outline=fg, width=2)
            canvas.create_line(9, 22, 31, 22, fill=fg, width=2)
            canvas.create_line(14, 13, 14, 22, fill=fg, width=2)
            canvas.create_line(26, 13, 26, 22, fill=fg, width=2)
        elif key == "groom_cost":
            canvas.create_rectangle(10, 20, 14, 29, fill=fg, outline=fg)
            canvas.create_rectangle(18, 15, 22, 29, fill=fg, outline=fg)
            canvas.create_rectangle(26, 11, 30, 29, fill=fg, outline=fg)
            canvas.create_line(9, 30, 31, 30, fill=fg, width=2)
        elif key == "person_cost":
            canvas.create_oval(15, 10, 25, 20, fill=fg, outline=fg)
            canvas.create_arc(10, 18, 30, 36, start=0, extent=180, outline=fg, width=2)
        elif key == "balance":
            canvas.create_oval(13, 10, 23, 20, fill=fg, outline=fg)
            canvas.create_arc(8, 18, 28, 36, start=0, extent=180, outline=fg, width=2)
            canvas.create_text(28, 22, text="$", fill=fg, font=(FONT_FAMILY, 10, "bold"))
        canvas.scale("all", 20, 20, 0.84, 0.84)

    def create_metric_ribbon_item(self, parent, key, title, value):
        item = tk.Frame(parent, bg="#FFFFFF", padx=12, pady=8)
        icon_image = self._get_metric_icon_image(key)
        if icon_image:
            icon_wrap = tk.Label(item, image=icon_image, bg="#FFFFFF", bd=0)
            icon_wrap.image = icon_image
        else:
            icon_wrap = tk.Canvas(item, width=36, height=36, bg="#FFFFFF", highlightthickness=0, bd=0)
            self._draw_metric_icon(icon_wrap, key)
        icon_wrap.pack(side="left", padx=(3, 9), pady=(0, 0))
        text_box = tk.Frame(item, bg="#FFFFFF")
        text_box.pack(side="right", fill="both", expand=True)
        tk.Label(
            text_box,
            text=title,
            bg="#FFFFFF",
            fg="#111827",
            font=(FONT_FAMILY, 10),
            anchor="e",
            justify="right",
        ).pack(fill="x")
        value_label = tk.Label(
            text_box,
            textvariable=value,
            bg="#FFFFFF",
            fg="#111827",
            font=(FONT_FAMILY, 18, "bold"),
            anchor="e",
            justify="right",
        )
        value_label.pack(fill="x", pady=(1, 0))
        item.value_label = value_label
        return item

    def _get_metric_icon_image(self, key):
        if Image is None or ImageTk is None:
            return None
        if not hasattr(self, "_metric_icon_cache"):
            self._metric_icon_cache = {}
        if key in self._metric_icon_cache:
            return self._metric_icon_cache[key]
        icon_map = {
            "all": "metric_total.png",
            "furn": "metric_furniture.png",
            "wedding": "metric_wedding.png",
            "finish": "metric_finishing.png",
            "count": "metric_count.png",
            "receipts": "metric_receipts.png",
            "engineer": "metric_engineer.png",
            "groom_cost": "metric_cost.png",
            "person_cost": "metric_share.png",
            "balance": "metric_balance.png",
        }
        path = ICONS_DIR / icon_map.get(key, "")
        if not path.exists():
            return None
        try:
            img = Image.open(path).convert("RGBA").resize((40, 40), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            self._metric_icon_cache[key] = photo
            return photo
        except Exception:
            return None

    def _make_flat_button(
        self,
        parent,
        text,
        command,
        bg=PRIMARY,
        fg="#FFFFFF",
        padx=15,
        pady=5,
        relief="flat",
        bd=0,
        font_size=9,
        hover_bg=None,
        highlightbackground=None,
        highlightthickness=0,
    ):
        btn = tk.Button(
            parent,
            text=text,
            command=command,
            bg=bg,
            fg=fg,
            activebackground=hover_bg or bg,
            activeforeground=fg,
            relief=relief,
            bd=bd,
            highlightbackground=highlightbackground or bg,
            highlightthickness=highlightthickness,
            padx=padx,
            pady=pady,
            cursor="hand2",
            anchor="e",
            justify="right",
            font=(FONT_FAMILY, font_size, "bold"),
        )
        if hover_bg:
            btn.bind("<Enter>", lambda _e: btn.configure(bg=hover_bg))
            btn.bind("<Leave>", lambda _e: btn.configure(bg=bg))
        return btn

    def _make_ghost_button(self, parent, text, command):
        return self._make_flat_button(parent, text, command, bg=SURFACE, fg=TEXT, padx=15, pady=5)

    def _add_sidebar_button(self, parent, text, command):
        btn = SidebarButton(parent, text, command, self)
        btn.pack(fill="x", pady=2, padx=12)
        self.sidebar_buttons.append(btn)
        if len(self.sidebar_buttons) == 1:
            btn.set_active(True)
        return btn

    def _set_sidebar_active(self, active_button):
        for button in getattr(self, "sidebar_buttons", []):
            button.set_active(button is active_button)

    def _build_currency_strip(self, parent):
        strip = tk.Frame(parent, bg="#EADDBF", bd=0, padx=0, pady=0, height=42)
        strip.pack_propagate(False)
        self.currency_value_labels = {}
        title_block = tk.Frame(strip, bg="#F8F3E8", padx=20, pady=6)
        title_block.pack(side="right", fill="y")
        tk.Label(title_block, text="السلف", bg="#F8F3E8", fg="#111827", font=(FONT_FAMILY, 10, "bold"), anchor="center", justify="center").pack(fill="both", expand=True)
        tk.Frame(strip, bg="#FFFFFF", width=1).pack(side="right", fill="y", pady=10)
        for idx, currency in enumerate(LOAN_CURRENCIES):
            block = tk.Frame(strip, bg="#EADDBF", padx=18, pady=4)
            block.pack(side="right", fill="both", expand=True)
            tk.Label(block, text=currency, bg="#EADDBF", fg="#111827", font=(FONT_FAMILY, 8, "bold"), anchor="center", justify="center").pack(fill="x")
            value = tk.Label(block, text="0.00", bg="#EADDBF", fg="#111827", font=(FONT_FAMILY, 9, "bold"), anchor="center", justify="center")
            value.pack(fill="x", pady=(1, 0))
            self.currency_value_labels[currency] = value
            if idx < len(LOAN_CURRENCIES) - 1:
                tk.Frame(strip, bg="#FFFFFF", width=1).pack(side="right", fill="y", pady=10)
        return strip

    def _dialog_step_badge(self, parent, number):
        badge = tk.Canvas(parent, width=38, height=38, bg=SURFACE, highlightthickness=0, bd=0)
        badge.create_oval(3, 3, 35, 35, fill=GOLD, outline="#C7A65D", width=1)
        badge.create_text(19, 19, text=str(number), fill=PRIMARY, font=(FONT_FAMILY, 15, "bold"))
        return badge

    def _dialog_field_label(self, parent, text):
        tk.Label(
            parent,
            text=text,
            bg=SURFACE,
            fg="#FFFFFF",
            font=(FONT_FAMILY, 12, "bold"),
            anchor="e",
            justify="right",
        ).pack(fill="x", pady=(0, 5))

    def _dialog_text_input(self, parent, variable, state="normal", font_size=12):
        return RoundedTextInput(
            parent,
            variable=variable,
            state=state,
            font_size=font_size,
            surface_bg=SURFACE,
            input_bg="#F7EFE3",
            border_color="#CBB58A",
            focus_color=PRIMARY,
            radius=10,
            native=False,
            height=46,
        )

    def _dialog_combo_input(self, parent, variable, values, font_size=11):
        return RoundedOptionInput(
            parent,
            variable=variable,
            values=values,
            font_size=font_size,
            surface_bg=SURFACE,
            input_bg="#FFFDF7",
            border_color="#CBB58A",
            focus_color=PRIMARY,
            radius=10,
            height=44,
        )

    def _dialog_save_button(self, parent, text, command):
        holder = tk.Frame(parent, bg=parent.cget("bg"), height=56)
        holder.pack(fill="x", pady=(14, 0))
        holder.pack_propagate(False)
        btn = RoundedCanvasButton(
            holder,
            text,
            command,
            bg=PRIMARY,
            hover_bg=PRIMARY_HOVER,
            fg=GOLD,
            width=380,
            height=46,
            radius=10,
            font_size=12,
        )
        btn.pack(fill="both", expand=True)
        return btn

    def _load_or_prompt_party_names(self):
        user_name, other_party_name = load_party_names()
        if user_name and other_party_name and not TEST_FIRST_RUN_EVERY_LAUNCH:
            return user_name, other_party_name

        dialog = tk.Toplevel(self.root)
        dialog.title("إعداد أول تشغيل")
        dialog.configure(bg=BG_APP)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, True)
        dialog.protocol("WM_DELETE_WINDOW", lambda: None)

        project_var = tk.StringVar(value=load_project_name())
        user_var = tk.StringVar(value=user_name or DEFAULT_USER_NAME)
        other_var = tk.StringVar(value=other_party_name or "")
        party1_role_var = tk.StringVar(value=load_party_roles()[0])
        party2_role_var = tk.StringVar(value=load_party_roles()[1])
        base_currency_var = tk.StringVar(value=load_base_currency())
        accounting_mode_var = tk.StringVar(value=load_accounting_settings()[0])
        wedding_policy_var = tk.StringVar(value=load_accounting_settings()[1])
        split_mode, p1, p2 = load_split_settings()
        split_mode_var = tk.StringVar(value=split_mode)
        party1_percent_var = tk.StringVar(value=f"{p1:.0f}")
        party2_percent_var = tk.StringVar(value=f"{p2:.0f}")
        error_var = tk.StringVar(value="")
        result = {}

        shell = tk.Frame(dialog, bg=BG_APP, padx=18, pady=14)
        shell.pack(fill="both", expand=True)
        outer_shell = RoundedCardFrame(shell, bg=SURFACE, border="#D9CDB4", shadow="#E6D9C4", radius=18, padx=22, pady=18)
        outer_shell.pack(fill="both", expand=True)
        outer = outer_shell.body

        def refresh_dialog_layout():
            def refresh():
                outer_shell.refresh_size()
                self._fit_toplevel(dialog, 500, 720, 0.40, 0.96)
            dialog.after_idle(refresh)

        self._dialog_step_badge(outer, 1).pack(anchor="ne", pady=(0, 0))
        tk.Label(outer, text="إعداد أول تشغيل", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 16), anchor="center", justify="center").pack(fill="x", pady=(0, 5))
        tk.Label(outer, text="أدخل أسماء الأطراف", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 21, "bold"), anchor="center", justify="center").pack(fill="x", pady=(0, 5))
        tk.Label(
            outer,
            text="سيتم استخدام هذه الأسماء في الحسابات، اختيار الطرف المسؤول، والتقارير.",
            bg=SURFACE,
            fg=TEXT,
            font=(FONT_FAMILY, 12),
            anchor="center",
            justify="center",
            wraplength=380,
        ).pack(fill="x", pady=(0, 8))

        fields = (
            ("اسم المشروع", "", project_var),
        )
        for label, note, var in fields:
            field = tk.Frame(outer, bg=SURFACE)
            field.pack(fill="x", pady=4)
            self._dialog_field_label(field, label)
            if note:
                ttk.Label(field, text=note, anchor="e", style="HeaderSub.TLabel").pack(fill="x", pady=(2, 0))
            self._dialog_text_input(field, var).pack(fill="x", pady=(0, 0))

        for title, name_var, role_var in (
            ("الطرف الأول", user_var, party1_role_var),
            ("الطرف الثاني", other_var, party2_role_var),
        ):
            party_row = tk.Frame(outer, bg=SURFACE)
            party_row.pack(fill="x", pady=4)
            party_row.columnconfigure(0, weight=1)
            party_row.columnconfigure(1, weight=1)
            role_frame = tk.Frame(party_row, bg=SURFACE)
            role_frame.grid(row=0, column=0, sticky="ew", padx=(0, 4))
            self._dialog_field_label(role_frame, f"صفة {title}")
            role_combo = self._dialog_combo_input(role_frame, role_var, PARTY_ROLE_OPTIONS)
            role_combo.pack(fill="x")
            name_frame = tk.Frame(party_row, bg=SURFACE)
            name_frame.grid(row=0, column=1, sticky="ew", padx=(4, 0))
            self._dialog_field_label(name_frame, f"اسم {title}")
            self._dialog_text_input(name_frame, name_var).pack(fill="x")

        split_title = tk.Label(outer, text="إعدادات الحسابات", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 15, "bold"), anchor="e", justify="right")
        split_title.pack(fill="x", pady=(7, 4))
        split_box = tk.Frame(outer, bg=SURFACE, bd=0, highlightbackground="#E3D6BF", highlightthickness=1, padx=10, pady=10)
        split_box.pack(fill="x", pady=(4, 0))
        split_box.columnconfigure(0, weight=1)
        split_box.columnconfigure(1, weight=1)
        tk.Label(split_box, text="العملة الأساسية", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 11, "bold"), anchor="e").grid(row=0, column=1, sticky="ew", padx=4)
        self._dialog_combo_input(split_box, base_currency_var, LOAN_CURRENCIES).grid(row=1, column=1, sticky="ew", padx=4, pady=(4, 8))
        tk.Label(split_box, text="نظام المحاسبة", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 11, "bold"), anchor="e").grid(row=0, column=0, sticky="ew", padx=4)
        self._dialog_combo_input(split_box, accounting_mode_var, ACCOUNTING_MODE_OPTIONS).grid(row=1, column=0, sticky="ew", padx=4, pady=(4, 8))
        wedding_policy_label = ttk.Label(split_box, text="حفل الزفاف في هذا النظام", anchor="e", style="Field.TLabel")
        wedding_policy_combo = self._dialog_combo_input(split_box, wedding_policy_var, WEDDING_POLICY_OPTIONS)
        ttk.Label(
            split_box,
            text="نظام الطرف الواحد: كل البنود على الطرف الأول. نظام 50/50: الحساب بالمناصفة. نظام البنود: كل بند تحدد عليه الطرف المسؤول.",
            anchor="e",
            justify="right",
            style="HeaderSub.TLabel",
            wraplength=470,
        ).grid(row=4, column=0, columnspan=2, sticky="ew", padx=4, pady=(8, 0))

        def update_wedding_policy_visibility(*_):
            if accounting_mode_var.get() in (ACCOUNTING_SINGLE, ACCOUNTING_ITEMIZED):
                wedding_policy_label.grid(row=2, column=0, columnspan=2, sticky="ew", padx=4)
                wedding_policy_combo.grid(row=3, column=0, columnspan=2, sticky="ew", padx=4, pady=(4, 0))
            else:
                wedding_policy_label.grid_remove()
                wedding_policy_combo.grid_remove()
            refresh_dialog_layout()

        accounting_mode_var.trace_add("write", update_wedding_policy_visibility)
        update_wedding_policy_visibility()

        def sync_split_fields(*_):
            mode, np1, np2 = normalize_split(split_mode_var.get(), party1_percent_var.get(), party2_percent_var.get())
            if mode != SPLIT_MODE_CUSTOM:
                party1_percent_var.set(f"{np1:.0f}")
                party2_percent_var.set(f"{np2:.0f}")

        split_mode_var.trace_add("write", sync_split_fields)
        ttk.Label(outer, textvariable=error_var, foreground=DANGER, background=SURFACE, anchor="center").pack(fill="x", pady=(8, 0))

        def save_names():
            user = user_var.get().strip()
            other = other_var.get().strip()
            if not user or not other:
                error_var.set("من فضلك أدخل الاسمين.")
                return
            if user == other:
                error_var.set("لازم يكون اسم الطرف الثاني مختلف.")
                return
            split_mode_var.set(SPLIT_MODE_EQUAL)
            party1_percent_var.set("50")
            party2_percent_var.set("50")
            save_project_name(project_var.get().strip() or APP_TITLE)
            save_party_names(user, other)
            save_party_roles(party1_role_var.get(), party2_role_var.get())
            save_base_currency(base_currency_var.get())
            save_split_settings(split_mode_var.get(), party1_percent_var.get(), party2_percent_var.get())
            save_accounting_settings(accounting_mode_var.get(), wedding_policy_var.get())
            result["names"] = (user, other)
            dialog.destroy()

        self._dialog_save_button(outer, "حفظ والانتقال للخطوة التالية", save_names)
        dialog.bind("<Return>", lambda event: save_names())
        self._fit_toplevel(dialog, 500, 720, 0.40, 0.96)
        self.root.wait_window(dialog)
        if "names" not in result:
            fallback_user = user_var.get().strip() or DEFAULT_USER_NAME
            fallback_other = other_var.get().strip() or DEFAULT_OTHER_PARTY_NAME
            if fallback_user == fallback_other:
                fallback_other = DEFAULT_OTHER_PARTY_NAME
            result["names"] = (fallback_user, fallback_other)
        return result["names"]

    def _ensure_initial_room_options(self):
        data = load_config()
        if data.get("room_setup_done") and not TEST_FIRST_RUN_EVERY_LAUNCH:
            save_room_options(data.get("room_options") or DEFAULT_ROOM_OPTIONS)
            return

        dialog = tk.Toplevel(self.root)
        dialog.title("إعداد أماكن الشقة")
        dialog.configure(bg=BG_APP)
        dialog.transient(self.root)
        dialog.grab_set()
        dialog.resizable(False, True)
        dialog.protocol("WM_DELETE_WINDOW", lambda: None)

        shell = tk.Frame(dialog, bg=BG_APP, padx=26, pady=24)
        shell.pack(fill="both", expand=True)
        outer_shell = RoundedCardFrame(shell, bg=SURFACE, border="#D9CDB4", shadow="#E6D9C4", radius=18, padx=26, pady=24)
        outer_shell.pack(fill="both", expand=True)
        outer = outer_shell.body

        def refresh_dialog_layout():
            def refresh():
                outer_shell.refresh_size()
                self._fit_toplevel(dialog, 460, 760, 0.36, 0.92)
            dialog.after_idle(refresh)

        self._dialog_step_badge(outer, 2).pack(anchor="ne", pady=(0, 2))
        tk.Label(outer, text="إعداد أماكن الشقة", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 18), anchor="center", justify="center").pack(fill="x", pady=(0, 8))
        tk.Label(outer, text="أدخل مكونات الشقة", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 22, "bold"), anchor="center", justify="center").pack(fill="x", pady=(0, 8))
        tk.Label(
            outer,
            text="أضف أسماء الأماكن أو الغرف التي ستسجل عليها البنود لاحقًا. خانة مجمع موجودة دائمًا للاستخدام العام.",
            bg=SURFACE,
            fg=TEXT,
            font=(FONT_FAMILY, 12),
            anchor="center",
            justify="center",
            wraplength=430,
        ).pack(fill="x", pady=(0, 12))

        rooms_inner = tk.Frame(outer, bg=SURFACE)
        rooms_inner.pack(fill="both", expand=True, pady=(12, 0))

        tk.Label(rooms_inner, text="مكونات الشقة", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 13, "bold"), anchor="e").pack(fill="x", pady=(0, 8))
        fixed_row = tk.Frame(rooms_inner, bg=SURFACE)
        fixed_row.pack(fill="x", pady=(0, 8))
        fixed_pill = tk.Canvas(fixed_row, height=48, bg=SURFACE, highlightthickness=0, bd=0)
        fixed_pill.pack(fill="x")

        def draw_fixed_pill(event=None):
            fixed_pill.delete("pill")
            w = max(10, fixed_pill.winfo_width())
            fixed_pill.create_polygon(
                10, 1, w - 10, 1, w - 1, 1, w - 1, 11, w - 1, 37, w - 1, 47,
                w - 10, 47, 10, 47, 1, 47, 1, 37, 1, 11, 1, 1,
                fill="#E7ECE7",
                outline="#D1D5DB",
                smooth=True,
                splinesteps=18,
                tags="pill",
            )
            fixed_pill.create_text(w - 22, 24, text=DEFAULT_ROOM_NAME, fill=PRIMARY, font=(FONT_FAMILY, 13), anchor="e", tags="pill")
            fixed_pill.create_text(22, 24, text="موجود دائمًا", fill=TEXT_2, font=(FONT_FAMILY, 12), anchor="w", tags="pill")

        fixed_pill.bind("<Configure>", draw_fixed_pill)

        room_vars = []
        rows_canvas = tk.Canvas(rooms_inner, bg=SURFACE, highlightthickness=0, bd=0, height=330)
        rows_canvas.pack(fill="both", expand=True, pady=(4, 0))
        rows_frame = tk.Frame(rows_canvas, bg=SURFACE)
        rows_window = rows_canvas.create_window(0, 0, anchor="nw", window=rows_frame)

        def sync_rows_scrollregion(_event=None):
            rows_canvas.configure(scrollregion=rows_canvas.bbox("all"))

        def sync_rows_width(event=None):
            rows_canvas.itemconfigure(rows_window, width=rows_canvas.winfo_width())

        def scroll_room_rows(event):
            rows_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        rows_frame.bind("<Configure>", sync_rows_scrollregion, add="+")
        rows_canvas.bind("<Configure>", sync_rows_width, add="+")
        rows_canvas.bind("<MouseWheel>", scroll_room_rows, add="+")
        rows_frame.bind("<MouseWheel>", scroll_room_rows, add="+")

        def rebuild_room_rows(focus_index=None):
            for child in rows_frame.winfo_children():
                child.destroy()
            for idx, var in enumerate(room_vars):
                row = tk.Frame(rows_frame, bg=SURFACE)
                row.pack(fill="x", pady=7)
                RoundedCanvasButton(row, "+", lambda pos=idx + 1: add_room_field(pos), bg=PRIMARY, hover_bg=PRIMARY_HOVER, fg=GOLD, width=46, height=38, radius=9, font_size=15).pack(side="right", padx=(0, 8))
                RoundedCanvasButton(row, "-", lambda pos=idx: remove_room_field(pos), bg=PRIMARY, hover_bg=PRIMARY_HOVER, fg=GOLD, width=46, height=38, radius=9, font_size=15).pack(side="right", padx=(0, 8))
                entry = self._dialog_text_input(row, var)
                entry.pack(side="right", fill="x", expand=True)
                row.bind("<MouseWheel>", scroll_room_rows, add="+")
                entry.bind("<MouseWheel>", scroll_room_rows, add="+")
                if focus_index == idx:
                    entry.focus_set()
            refresh_dialog_layout()

        def add_room_field(pos=None):
            insert_at = len(room_vars) if pos is None else pos
            room_vars.insert(insert_at, tk.StringVar(value=""))
            rebuild_room_rows(insert_at)

        def remove_room_field(pos):
            if 0 <= pos < len(room_vars):
                room_vars.pop(pos)
            if not room_vars:
                room_vars.append(tk.StringVar(value=""))
            rebuild_room_rows(min(pos, len(room_vars) - 1))

        add_room_field()

        error_var = tk.StringVar(value="")
        ttk.Label(outer, textvariable=error_var, foreground=DANGER, background=SURFACE, anchor="center").pack(fill="x", pady=(8, 0))
        tk.Frame(outer, bg=SURFACE).pack(fill="both", expand=True)

        def save_rooms():
            rooms = normalize_room_options(var.get() for var in room_vars)
            save_room_options(rooms)
            data = load_config()
            data["room_setup_done"] = True
            save_config(data)
            dialog.destroy()

        self._dialog_save_button(outer, "حفظ الأماكن وفتح البرنامج", save_rooms)
        dialog.bind("<Return>", lambda event: save_rooms())
        self._fit_toplevel(dialog, 460, 760, 0.36, 0.92)
        self.root.wait_window(dialog)

    def _migrate_payer_names(self):
        self.conn.execute(
            "UPDATE records SET payer=? WHERE payer IS NULL OR payer='' OR payer=?",
            (self.user_name, DEFAULT_USER_NAME),
        )
        self.conn.execute(
            "UPDATE records SET payer=? WHERE payer IN (?, ?)",
            (self.other_party_name, DEFAULT_OTHER_PARTY_NAME, LEGACY_OTHER_PARTY_NAME),
        )
        self.conn.commit()

    def _build_variables(self):
        self.room_options = load_room_options()
        self.main_type_var = tk.StringVar(value="")
        self.finish_type_var = tk.StringVar(value="")
        self.payer_var = tk.StringVar(value="")
        self.shared_only_var = tk.BooleanVar(value=False)
        self.item_name_var = tk.StringVar()
        self.room_var = tk.StringVar(value=self.room_options[0] if self.room_options else "")
        self.quantity_var = tk.StringVar(value="1")
        self.unit_price_var = tk.StringVar(value="0")
        self.total_var = tk.StringVar(value="0.00")
        self.invoice_name_var = tk.StringVar(value="لا يوجد ملف مرفق")
        self.status_var = tk.StringVar(value="جاهز")

        self.search_var = tk.StringVar()
        self.filter_type_var = tk.StringVar(value="الكل")
        self.filter_room_var = tk.StringVar(value="الكل")
        self.preview_title_var = tk.StringVar(value="معاينة الفاتورة")
        self.record_details_var = tk.StringVar(value="اختر سجلًا من نتائج البحث لعرض التفاصيل.")

        self.metric_vars = {
            "count": tk.StringVar(value="0"),
            "furn": tk.StringVar(value="0.00"),
            "finish": tk.StringVar(value="0.00"),
            "wedding": tk.StringVar(value="0.00"),
            "all": tk.StringVar(value="0.00"),
            "receipts": tk.StringVar(value="0.00"),
            "remaining": tk.StringVar(value="0.00"),
            "person_cost": tk.StringVar(value="0.00"),
            "balance": tk.StringVar(value="0.00"),
            "groom_cost": tk.StringVar(value="0.00"),
            "engineer": tk.StringVar(value="0.00"),
            "loans": tk.StringVar(value="لا توجد سلف"),
        }

        self.receipt_amount_var = tk.StringVar(value="0")
        self.receipt_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.receipt_time_var = tk.StringVar(value=datetime.now().strftime("%H:%M"))
        self.receipt_method_var = tk.StringVar(value=PAYMENT_METHODS[0])

        self.advance_amount_var = tk.StringVar(value="0")
        self.advance_direction_var = tk.StringVar(value=self.payment_targets[0])
        self.advance_reason_var = tk.StringVar()
        self.advance_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.advance_time_var = tk.StringVar(value=datetime.now().strftime("%H:%M"))
        self.editing_advance_id = None

        self.loan_title_var = tk.StringVar()
        self.loan_amount_var = tk.StringVar(value="0")
        self.loan_currency_var = tk.StringVar(value=self.base_currency)
        self.loan_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.loan_time_var = tk.StringVar(value=datetime.now().strftime("%H:%M"))
        self.editing_loan_id = None

        self.loan_payment_amount_var = tk.StringVar(value="0")
        self.loan_payment_currency_var = tk.StringVar(value=self.base_currency)
        self.loan_payment_method_var = tk.StringVar(value=LOAN_PAYMENT_METHODS[0])
        self.loan_payment_date_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        self.loan_payment_time_var = tk.StringVar(value=datetime.now().strftime("%H:%M"))
        self.loan_payment_receipt_var = tk.StringVar(value="لا يوجد إيصال مرفق")
        self.loan_payment_context_var = tk.StringVar(value="")
        self.selected_loan_payment_receipt_source = None
        self.current_payment_loan_id = None
        self.current_payment_loan_title = ""
        self.support_name_var = tk.StringVar()
        self.support_email_var = tk.StringVar()
        self.support_type_var = tk.StringVar(value=SUPPORT_MESSAGE_TYPES[0])
        self.support_subject_var = tk.StringVar()
        self.support_sending_var = tk.BooleanVar(value=False)

    def _build_ui(self):
        self.main_canvas = tk.Canvas(self.root, bg=BG_APP, highlightthickness=0, bd=0)
        self.v_scroll = ttk.Scrollbar(self.root, orient="vertical", command=self.main_canvas.yview)
        self.h_scroll = ttk.Scrollbar(self.root, orient="horizontal", command=self.main_canvas.xview)
        self.main_canvas.configure(yscrollcommand=self.v_scroll.set, xscrollcommand=self.h_scroll.set)
        self._v_scroll_visible = False
        self.main_canvas.pack(side="left", fill="both", expand=True)

        outer = ttk.Frame(self.main_canvas, padding=0)
        self._canvas_window = self.main_canvas.create_window((0, 0), window=outer, anchor="nw")

        def _update_scroll_region(event=None):
            try:
                bbox = self.main_canvas.bbox("all")
                self.main_canvas.configure(scrollregion=bbox)
                needs_scroll = bool(bbox and bbox[3] > self.main_canvas.winfo_height() + 2)
                if needs_scroll and not self._v_scroll_visible:
                    self.v_scroll.pack(side="right", fill="y")
                    self._v_scroll_visible = True
                elif not needs_scroll and self._v_scroll_visible:
                    self.v_scroll.pack_forget()
                    self._v_scroll_visible = False
            except Exception:
                pass

        def _resize_canvas_window(event):
            try:
                self.main_canvas.itemconfigure(self._canvas_window, width=max(event.width, 980))
            except Exception:
                pass
            _update_scroll_region()

        outer.bind("<Configure>", _update_scroll_region)
        self.main_canvas.bind("<Configure>", _resize_canvas_window)

        def _on_mousewheel(event):
            try:
                self.main_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass

        self.main_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        content = tk.Frame(outer, bg=BG_APP)
        content.pack(fill="both", expand=True)

        sidebar_shell = tk.Frame(content, bg=SIDEBAR_BG, width=260, highlightbackground="#D4C3A3", highlightthickness=1)
        sidebar_shell.pack(side="right", fill="y")
        sidebar_shell.pack_propagate(False)
        sidebar = tk.Frame(sidebar_shell, bg=SIDEBAR_BG, padx=12, pady=12)
        sidebar.pack(side="right", fill="both", expand=True)
        sidebar.pack_propagate(False)
        self.sidebar_buttons = []

        main_area = tk.Frame(content, bg=BG_APP)
        main_area.pack(side="left", fill="both", expand=True)

        top_bar = tk.Frame(main_area, bg="#F7F0E2", height=64, highlightbackground="#D4C3A3", highlightthickness=1)
        top_bar.pack(side="top", fill="x", pady=(0, 14))
        top_bar.pack_propagate(False)
        self.project_title_var = tk.StringVar(value=self.project_name)
        tk.Label(top_bar, text="إدارة المصاريف", bg="#F7F0E2", fg=TEXT, font=(FONT_FAMILY, 17, "bold"), anchor="e", justify="right").pack(side="right", fill="y", padx=(24, 14))
        tk.Label(top_bar, textvariable=self.project_title_var, bg="#F7F0E2", fg=TEXT, font=(FONT_FAMILY, 10), anchor="e", justify="right").pack(side="right", fill="y", padx=(0, 8))

        metric_shell = RoundedCardFrame(main_area, bg="#FFFFFF", border="#E5E7EB", shadow="#E8E8E8", radius=10, padx=0, pady=0)
        metric_shell.pack(side="top", fill="x", padx=30, pady=(6, 0))
        metric_ribbon = metric_shell.body
        metric_ribbon.configure(bg="#FFFFFF")
        self.summary_cards = {}
        self.summary_card_boxes = {}

        row_groups = (
            [
                ("count", "عدد البنود"),
                ("finish", "التشطيب"),
                ("wedding", "حفل الزواج"),
                ("furn", "الفرش"),
                ("all", "الإجمالي"),
            ],
            [
                ("receipts", "التوريدات"),
                ("engineer", self.payment_targets[1] if len(self.payment_targets) > 1 else "جهة تسديد"),
                ("groom_cost", f"إجمالي تكلفة {self.user_name}"),
                ("person_cost", f"نصيب {self.user_name}"),
                ("balance", "الموازنة"),
            ],
        )
        for row_index, row_cards in enumerate(row_groups):
            if row_index:
                tk.Frame(metric_ribbon, bg="#E5E7EB", height=1).pack(fill="x", padx=12)
            row_frame = tk.Frame(metric_ribbon, bg="#FFFFFF")
            row_frame.pack(fill="x")
            for idx, (key, title) in enumerate(row_cards):
                box = self.create_metric_ribbon_item(row_frame, key, title, self.metric_vars[key])
                box.pack(side="right", expand=True, fill="both")
                self.summary_card_boxes[key] = (box, row_index, idx)
                self.summary_cards[key] = box.value_label
                if idx < len(row_cards) - 1:
                    tk.Frame(row_frame, bg="#E5E7EB", width=1).pack(side="right", fill="y", pady=10)

        currency_strip = self._build_currency_strip(metric_ribbon)
        currency_strip.pack(side="top", fill="x")

        form_container = tk.Frame(main_area, bg=BG_APP)
        form_container.pack(fill="both", expand=True, padx=30)
        form_container.columnconfigure(0, weight=2)
        form_container.columnconfigure(1, weight=3)
        form_container.rowconfigure(0, weight=1)

        self.financial_shell = RoundedCardFrame(form_container, bg="#FFFFFF", border="#E5E7EB", shadow="#D6D6D6", radius=13, padx=20, pady=20)
        self.financial_shell.grid(row=0, column=0, sticky="nsew", padx=(0, 8), pady=(0, 10))
        financial_card = self.financial_shell.body
        self.details_shell = RoundedCardFrame(form_container, bg="#FFFFFF", border="#E5E7EB", shadow="#D6D6D6", radius=13, padx=20, pady=18)
        self.details_shell.grid(row=0, column=1, sticky="nsew", padx=(8, 0), pady=(0, 10))
        details_card = self.details_shell.body

        tk.Label(sidebar, text="💍", bg=SIDEBAR_BG, fg=PRIMARY, font=(FONT_FAMILY, 29), anchor="center", justify="center").pack(fill="x", pady=(0, 0), padx=12)
        tk.Label(sidebar, text=APP_TITLE, bg=SIDEBAR_BG, fg=TEXT, font=(FONT_FAMILY, 15, "bold"), anchor="center", justify="center").pack(fill="x", pady=(0, 10), padx=12)

        records_card = tk.Frame(sidebar, bg=SIDEBAR_BG)
        records_card.pack(fill="x")
        tk.Label(records_card, text="السجلات والطباعة", bg=SIDEBAR_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold"), anchor="e", justify="right").pack(fill="x", pady=(0, 6), padx=12)
        self._add_sidebar_button(records_card, "بحث وإدارة السجلات", self.open_search_window)
        self._add_sidebar_button(records_card, "إدارة الأماكن", self.open_rooms_window)
        self._add_sidebar_button(records_card, "تقارير PDF", self.open_reports_dialog)
        self._add_sidebar_button(records_card, "تصدير Excel", self.export_to_excel)
        self._add_sidebar_button(records_card, "طباعة القائمة", self.print_furniture_list)

        finance_card = tk.Frame(sidebar, bg=SIDEBAR_BG)
        finance_card.pack(fill="x", pady=(6,0))
        tk.Label(finance_card, text="العمليات المالية", bg=SIDEBAR_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold"), anchor="e", justify="right").pack(fill="x", pady=(0, 6), padx=12)
        self._add_sidebar_button(finance_card, "توريد جديد", self.open_receipt_form_window)
        self._add_sidebar_button(finance_card, "عرض التوريدات", self.open_receipts_view_window)
        self._add_sidebar_button(finance_card, "تسديد جديد", self.open_advance_form_window)
        self._add_sidebar_button(finance_card, "عرض التسديدات", self.open_advances_view_window)
        self._add_sidebar_button(finance_card, "سلفة جديدة", self.open_loan_form_window)
        self._add_sidebar_button(finance_card, "إدارة السلف", self.open_loans_view_window)

        helper_panel = tk.Frame(sidebar, bg=SIDEBAR_BG)
        helper_panel.pack(fill="x", pady=(6, 0))
        tk.Label(helper_panel, text="إعدادات ونسخ", bg=SIDEBAR_BG, fg=TEXT, font=(FONT_FAMILY, 10, "bold"), anchor="e", justify="right").pack(fill="x", pady=(0, 6), padx=12)
        self._add_sidebar_button(helper_panel, "إعدادات البرنامج", self.open_settings_window)
        self._add_sidebar_button(helper_panel, "تصدير نسخة", self.export_backup)
        self._add_sidebar_button(helper_panel, "استيراد نسخة", self.restore_backup)

        tk.Frame(sidebar, bg=SIDEBAR_BG).pack(fill="both", expand=True)
        bottom_panel = tk.Frame(sidebar, bg=SIDEBAR_BG, padx=10, pady=8)
        bottom_panel.pack(side="bottom", fill="x", pady=(4, 0))
        self._action_button(bottom_panel, "إبلاغ أو اقتراح", self.open_support_window, kind="gold", width=210, height=40).pack(fill="x", pady=(0, 8))

        copyright_panel = tk.Frame(bottom_panel, bg=SIDEBAR_BG)
        copyright_panel.pack(fill="x")
        tk.Frame(copyright_panel, bg="#D4C3A3", height=1).pack(fill="x", pady=(0, 6))
        tk.Label(copyright_panel, text="</>", bg=SIDEBAR_BG, fg=PRIMARY, font=(FONT_FAMILY, 10, "bold"), anchor="center", justify="center").pack(fill="x")
        tk.Label(copyright_panel, text=COPYRIGHT_NOTICE, bg=SIDEBAR_BG, fg=TEXT, font=(FONT_FAMILY, 7, "bold"), anchor="center", justify="center", wraplength=214).pack(fill="x", pady=(1, 0))
        tk.Label(copyright_panel, text=OWNER_EMAIL, bg=SIDEBAR_BG, fg=TEXT_2, font=(FONT_FAMILY, 7), anchor="center", justify="center", wraplength=214).pack(fill="x", pady=(0, 0))

        tk.Label(
            details_card,
            text="إضافة بند جديد - تفاصيل البند",
            bg="#FFFFFF",
            fg="#111827",
            font=(FONT_FAMILY, 13, "bold"),
            anchor="e",
            justify="right",
        ).pack(fill="x", pady=(0, 8))
        tk.Label(
            details_card,
            textvariable=self.status_var,
            bg="#F4EFE7",
            fg="#111827",
            font=(FONT_FAMILY, 9),
            anchor="e",
            justify="right",
        ).pack(fill="x", pady=(0, 12))

        form = tk.Frame(details_card, bg="#FFFFFF")
        form.pack(fill="x")
        form.columnconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)

        def detail_label(parent, text):
            tk.Label(
                parent,
                text=text,
                bg="#FFFFFF",
                fg="#111827",
                font=(FONT_FAMILY, 11, "bold"),
                anchor="e",
                justify="right",
            ).pack(fill="x", pady=(0, 6))

        def add_detail_combo(parent, label, variable, values, row, col, colspan=1):
            frame = tk.Frame(parent, bg="#FFFFFF")
            frame.grid(row=row, column=col, columnspan=colspan, sticky="nsew", padx=5, pady=(0, 16))
            frame.label = tk.Label(frame, text=label, bg="#FFFFFF", fg="#111827", font=(FONT_FAMILY, 11, "bold"), anchor="e", justify="right")
            frame.label.pack(fill="x", pady=(0, 6))
            combo = RoundedComboInput(
                frame,
                variable=variable,
                values=values,
                font_size=11,
                surface_bg="#FFFFFF",
                input_bg="#FFFFFF",
                border_color="#D1D5DB",
                focus_color=PRIMARY,
                radius=10,
            )
            combo.pack(fill="x")
            frame.widget = combo
            return frame

        def add_detail_entry(parent, label, variable, row, col, colspan=1):
            frame = tk.Frame(parent, bg="#FFFFFF")
            frame.grid(row=row, column=col, columnspan=colspan, sticky="nsew", padx=5, pady=(0, 16))
            detail_label(frame, label)
            entry = RoundedTextInput(
                frame,
                variable=variable,
                font_size=11,
                surface_bg="#FFFFFF",
                input_bg="#FFFFFF",
                border_color="#D1D5DB",
                focus_color=PRIMARY,
                radius=10,
                native=False,
            )
            entry.pack(fill="x")
            frame.widget = entry
            return frame

        self.main_type_frame = add_detail_combo(form, "نوع البند", self.main_type_var, TYPE_OPTIONS, 0, 1)
        self.finish_type_frame = add_detail_combo(form, "تصنيف التشطيب", self.finish_type_var, FINISHING_OPTIONS, 0, 0)
        self.payer_frame = add_detail_combo(form, "دافع البند", self.payer_var, self.payer_options, 1, 1)
        self.shared_only_frame = tk.Frame(form, bg="#FFFFFF")
        self.shared_only_frame.grid(row=1, column=0, sticky="nsew", padx=5, pady=(0, 16))
        detail_label(self.shared_only_frame, "خيار إضافي")
        self.shared_only_check = ttk.Checkbutton(
            self.shared_only_frame,
            text="هذا البند على الطرف الأول بالكامل ولا يدخل في نسب المشاركة",
            variable=self.shared_only_var,
            style="Secondary.TCheckbutton"
        )
        self.shared_only_check.pack(fill="x", pady=(4,0), anchor="e")
        self.item_name_frame = add_detail_entry(form, "اسم البند", self.item_name_var, 2, 0, colspan=2)
        room_line = tk.Frame(form, bg="#FFFFFF")
        room_line.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=5, pady=(0, 16))
        room_line.columnconfigure(0, weight=1)
        room_line.columnconfigure(1, weight=0)
        tk.Label(room_line, text="المكان", bg="#FFFFFF", fg="#111827", font=(FONT_FAMILY, 11, "bold"), anchor="e", justify="right").grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 6))
        self.room_combo = RoundedComboInput(
            room_line,
            variable=self.room_var,
            values=self.room_options,
            font_size=11,
            surface_bg="#FFFFFF",
            input_bg="#FFFFFF",
            border_color="#D1D5DB",
            focus_color=PRIMARY,
            radius=10,
        )
        self.room_combo.grid(row=1, column=0, sticky="ew")
        RoundedCanvasButton(room_line, "إدارة الأماكن", self.open_rooms_window, bg=PRIMARY, hover_bg=PRIMARY_HOVER, fg="#FFFFFF", width=104, height=36, radius=8, font_size=10).grid(row=1, column=1, padx=(8, 0), sticky="e")
        self.room_frame = room_line
        notes_frame = tk.Frame(form, bg="#FFFFFF")
        notes_frame.grid(row=4, column=0, columnspan=2, sticky="nsew", padx=5, pady=(0, 16))
        detail_label(notes_frame, "ملاحظات")
        self.notes_text = RoundedTextInput(
            notes_frame,
            font_size=10,
            multiline=True,
            rows=2,
            surface_bg="#FFFFFF",
            input_bg="#FFFFFF",
            border_color="#D1D5DB",
            focus_color=PRIMARY,
            radius=10,
        )
        self.notes_text.pack(fill="x")

        RoundedCanvasButton(
            details_card,
            "إضافة وحفظ",
            self.save_record,
            bg=GOLD,
            hover_bg=GOLD_HOVER,
            fg="#FFFFFF",
            width=640,
            height=38,
            radius=8,
            font_size=11,
        ).pack(fill="x", pady=(4, 0))

        tk.Label(
            financial_card,
            text="إضافة بند جديد - التكلفة والحسابات",
            bg="#FFFFFF",
            fg="#111827",
            font=(FONT_FAMILY, 13, "bold"),
            anchor="e",
            justify="right",
        ).pack(fill="x", pady=(0, 16))
        fin_form = tk.Frame(financial_card, bg="#FFFFFF")
        fin_form.pack(fill="x", pady=(0, 0))
        fin_form.columnconfigure(0, weight=1)
        fin_form.columnconfigure(1, weight=1)

        def add_clean_money_entry(parent, label, variable, row, col, state="normal", colspan=1):
            frame = tk.Frame(parent, bg="#FFFFFF")
            frame.grid(row=row, column=col, columnspan=colspan, sticky="nsew", padx=6, pady=(0, 18))
            tk.Label(
                frame,
                text=label,
                bg="#FFFFFF",
                fg="#111827",
                font=(FONT_FAMILY, 11, "bold"),
                anchor="e",
                justify="right",
            ).pack(fill="x", pady=(0, 6))
            entry = RoundedTextInput(
                frame,
                variable=variable,
                state=state,
                font_size=11,
                surface_bg="#FFFFFF",
                input_bg="#FFFFFF",
                border_color="#D1D5DB",
                focus_color=PRIMARY,
                radius=11,
                native=False,
            )
            entry.pack(fill="x")
            frame.widget = entry
            return frame

        total_frame = add_clean_money_entry(fin_form, "الإجمالي", self.total_var, 0, 1, state="readonly")
        self.quantity_frame = add_clean_money_entry(fin_form, "الكمية", self.quantity_var, 1, 1)
        self.unit_price_frame = add_clean_money_entry(fin_form, "سعر الوحدة", self.unit_price_var, 1, 0)

        invoice_label = tk.Label(
            financial_card,
            text="الفاتورة / المرفق",
            bg="#FFFFFF",
            fg="#111827",
            font=(FONT_FAMILY, 11, "bold"),
            anchor="e",
            justify="right",
        )
        invoice_label.pack(fill="x", pady=(2, 6))

        drop_bg = "#F8F4EC"
        drop_canvas = tk.Canvas(financial_card, bg="#FFFFFF", highlightthickness=0, bd=0, height=235)
        drop_canvas.pack(fill="x", pady=(0, 0))
        drop_inner = tk.Frame(drop_canvas, bg=drop_bg)
        drop_window = drop_canvas.create_window(0, 0, anchor="nw", window=drop_inner)
        drop_rect = drop_canvas.create_rectangle(1, 1, 10, 10, outline="#D1D5DB", dash=(3, 2), width=1, fill=drop_bg)

        def resize_drop_zone(event=None):
            width = max(1, drop_canvas.winfo_width())
            height = max(1, drop_canvas.winfo_height())
            drop_canvas.coords(drop_rect, 1, 1, width - 2, height - 2)
            drop_canvas.coords(drop_window, 2, 2)
            drop_canvas.itemconfigure(drop_window, width=width - 4, height=height - 4)
            drop_canvas.tag_raise(drop_rect)

        drop_canvas.bind("<Configure>", resize_drop_zone, add="+")
        drop_inner.pack_propagate(False)

        tk.Label(drop_inner, textvariable=self.invoice_name_var, bg=drop_bg, fg="#111827", font=(FONT_FAMILY, 9, "bold"), anchor="e", justify="right").pack(fill="x", padx=14, pady=(8, 6))
        attachment_actions = tk.Frame(drop_inner, bg=drop_bg)
        attachment_actions.pack(fill="x", padx=14, pady=(0, 18))
        RoundedCanvasButton(attachment_actions, "إرفاق فاتورة", self.pick_invoice_file, bg=PRIMARY, hover_bg=PRIMARY_HOVER, fg="#FFFFFF", width=116, height=36, radius=8, font_size=10).pack(side="right", padx=(5, 0))
        RoundedCanvasButton(attachment_actions, "فتح الملف", self.open_selected_invoice_source, bg=PRIMARY, hover_bg=PRIMARY_HOVER, fg="#FFFFFF", width=108, height=36, radius=8, font_size=10).pack(side="right", padx=5)
        RoundedCanvasButton(attachment_actions, "🗑", self.remove_selected_invoice, bg=DANGER, hover_bg="#B91C1C", fg="#FFFFFF", width=42, height=36, radius=8, font_size=11).pack(side="right", padx=5)

        icon_canvas = tk.Canvas(drop_inner, width=58, height=64, bg=drop_bg, highlightthickness=0, bd=0)
        icon_canvas.pack(pady=(6, 4))
        icon_canvas.create_polygon(14, 5, 40, 5, 52, 17, 52, 58, 14, 58, outline=PRIMARY, fill="", width=3)
        icon_canvas.create_line(40, 5, 40, 18, 52, 18, fill=PRIMARY, width=3)
        tk.Label(drop_inner, text="اسحب وأفلت الملفات هنا", bg=drop_bg, fg="#111827", font=(FONT_FAMILY, 10), anchor="center", justify="center").pack(fill="x")
        self._register_invoice_drop_target(drop_canvas, drop_inner, attachment_actions, icon_canvas)

    def _configure_rtl_text_widget(self, widget):
        widget.configure(padx=8, pady=6, insertbackground=TEXT, selectbackground="#2563EB", selectforeground="#FFFFFF")
        widget.tag_configure("rtl", justify="right", rmargin=6, lmargin1=6, lmargin2=6)
        self._set_text_value(widget, "")

        def keep_rtl(event=None):
            try:
                widget.tag_add("rtl", "1.0", "end")
            except Exception:
                pass

        widget.bind("<KeyRelease>", keep_rtl, add="+")
        widget.bind("<<Paste>>", lambda e: widget.after_idle(keep_rtl), add="+")
        widget.bind("<FocusIn>", keep_rtl, add="+")
        return widget

    def _clean_text_value(self, text):
        return (text or "").replace(RTL_START, "").replace(RTL_END, "").strip()

    def _get_text_value(self, widget):
        if isinstance(widget, (tk.Entry, ttk.Entry, NativeArabicEdit, RoundedTextInput)):
            return self._clean_text_value(widget.get())
        return self._clean_text_value(widget.get("1.0", "end-1c"))

    def _set_text_value(self, widget, value):
        if isinstance(widget, (tk.Entry, ttk.Entry, NativeArabicEdit, RoundedTextInput)):
            widget.delete(0, "end")
            widget.insert(0, value or "")
            return
        widget.delete("1.0", "end")
        widget.insert("1.0", value or "")
        widget.tag_add("rtl", "1.0", "end")
        widget.mark_set("insert", "end-1c")

    def _clear_text_value(self, widget):
        self._set_text_value(widget, "")

    def _register_focus_state(self, widget, base_style=None, focus_style=None):
        def focus_in(_event=None):
            try:
                if base_style and focus_style:
                    widget.configure(style=focus_style)
                else:
                    widget.configure(highlightbackground=INPUT_FOCUS, highlightcolor=INPUT_FOCUS)
            except Exception:
                pass

        def focus_out(_event=None):
            try:
                if base_style and focus_style:
                    widget.configure(style=base_style)
                else:
                    widget.configure(highlightbackground=INPUT_BORDER, highlightcolor=INPUT_BORDER)
            except Exception:
                pass

        widget.bind("<FocusIn>", focus_in, add="+")
        widget.bind("<FocusOut>", focus_out, add="+")
        return widget

    def _make_text_entry(self, parent, variable=None, state="normal", font_size=11):
        if os.name == "nt":
            return RoundedTextInput(parent, variable=variable, state=state, font_size=font_size, surface_bg=SURFACE, input_bg=INPUT_BG, border_color=INPUT_BORDER, focus_color=INPUT_FOCUS, radius=10, native=False, height=50)
        entry = tk.Entry(
            parent,
            textvariable=variable,
            justify="right",
            state=state,
            font=(FONT_FAMILY, font_size),
            bg=INPUT_BG,
            fg=TEXT,
            insertbackground=TEXT,
            readonlybackground="#EFE7D8",
            disabledbackground="#EFE7D8",
            disabledforeground=TEXT_2,
            relief="flat",
            bd=0,
            highlightthickness=1,
            highlightbackground=INPUT_BORDER,
            highlightcolor=INPUT_FOCUS,
            selectbackground="#2563EB",
            selectforeground="#FFFFFF",
            exportselection=False,
        )

        return self._register_focus_state(entry)

    def _make_combo_input(self, parent, variable=None, values=(), font_size=11):
        return RoundedOptionInput(parent, variable=variable, values=values, font_size=font_size, surface_bg=SURFACE, input_bg=INPUT_BG, border_color=INPUT_BORDER, focus_color=INPUT_FOCUS, radius=10, height=46)

    def _make_notes_widget(self, parent, rows=3):
        if os.name == "nt":
            return RoundedTextInput(parent, font_size=10, multiline=True, rows=rows, surface_bg=SURFACE, input_bg=INPUT_BG, border_color=INPUT_BORDER, focus_color=INPUT_FOCUS, radius=10, native=False)
        widget = tk.Text(parent, height=rows, wrap="word", font=(FONT_FAMILY, 10), bg=INPUT_BG, fg=TEXT, relief="flat", bd=0, highlightthickness=1, highlightbackground=INPUT_BORDER, highlightcolor=INPUT_FOCUS)
        return self._configure_rtl_text_widget(widget)

    def _add_labeled_entry(self, parent, label, variable, row, col, state="normal", colspan=1):
        frame = ttk.Frame(parent, style="Card.TFrame")
        frame.grid(row=row, column=col, columnspan=colspan, sticky="nsew", padx=7, pady=7)
        ttk.Label(frame, text=label, anchor="e", style="Field.TLabel").pack(fill="x", pady=(0, 4))
        entry = self._make_text_entry(frame, variable, state=state)
        entry.pack(fill="x", pady=(4, 14))
        frame.widget = entry
        return frame

    def _add_labeled_combobox(self, parent, label, variable, values, row, col, colspan=1):
        frame = ttk.Frame(parent, style="Card.TFrame")
        frame.grid(row=row, column=col, columnspan=colspan, sticky="nsew", padx=7, pady=7)
        frame.label = ttk.Label(frame, text=label, anchor="e", style="Field.TLabel")
        frame.label.pack(fill="x", pady=(0, 4))
        combo = self._make_combo_input(frame, variable, values)
        combo.pack(fill="x", pady=(4, 14))
        frame.widget = combo
        return frame

    def _install_text_editing_support(self):
        menu = tk.Menu(self.root, tearoff=0)

        def is_editable(widget):
            try:
                state = str(widget.cget("state")).lower()
                return state not in {"disabled", "readonly"}
            except Exception:
                return True

        def do_action(widget, virtual_event):
            if virtual_event in {"<<Paste>>", "<<Cut>>"} and not is_editable(widget):
                return "break"
            try:
                widget.event_generate(virtual_event)
            except Exception:
                pass
            return "break"

        def select_all(widget):
            try:
                if isinstance(widget, tk.Text):
                    widget.tag_add("sel", "1.0", "end-1c")
                    widget.mark_set("insert", "end-1c")
                else:
                    widget.selection_range(0, "end")
                    widget.icursor("end")
            except Exception:
                pass
            return "break"

        def physical_ctrl_shortcut(event):
            # Windows keeps the Latin virtual keycode even when the active keyboard layout is Arabic.
            if event.keycode in (65, 67, 86, 88):
                actions = {
                    65: lambda: select_all(event.widget),
                    67: lambda: do_action(event.widget, "<<Copy>>"),
                    86: lambda: do_action(event.widget, "<<Paste>>"),
                    88: lambda: do_action(event.widget, "<<Cut>>"),
                }
                return actions[event.keycode]()
            return None

        def popup(event):
            widget = event.widget
            menu.delete(0, "end")
            menu.add_command(label="قص", command=lambda: do_action(widget, "<<Cut>>"), state=("normal" if is_editable(widget) else "disabled"))
            menu.add_command(label="نسخ", command=lambda: do_action(widget, "<<Copy>>"))
            menu.add_command(label="لصق", command=lambda: do_action(widget, "<<Paste>>"), state=("normal" if is_editable(widget) else "disabled"))
            menu.add_separator()
            menu.add_command(label="تحديد الكل", command=lambda: select_all(widget))
            try:
                menu.tk_popup(event.x_root, event.y_root)
            finally:
                menu.grab_release()
            return "break"

        for cls in ("Entry", "TEntry", "Text", "TCombobox"):
            self.root.bind_class(cls, "<Control-a>", lambda e: select_all(e.widget))
            self.root.bind_class(cls, "<Control-A>", lambda e: select_all(e.widget))
            self.root.bind_class(cls, "<Control-c>", lambda e: do_action(e.widget, "<<Copy>>"))
            self.root.bind_class(cls, "<Control-C>", lambda e: do_action(e.widget, "<<Copy>>"))
            self.root.bind_class(cls, "<Control-v>", lambda e: do_action(e.widget, "<<Paste>>"))
            self.root.bind_class(cls, "<Control-V>", lambda e: do_action(e.widget, "<<Paste>>"))
            self.root.bind_class(cls, "<Control-x>", lambda e: do_action(e.widget, "<<Cut>>"))
            self.root.bind_class(cls, "<Control-X>", lambda e: do_action(e.widget, "<<Cut>>"))
            self.root.bind_class(cls, "<Shift-Insert>", lambda e: do_action(e.widget, "<<Paste>>"))
            self.root.bind_class(cls, "<Control-Insert>", lambda e: do_action(e.widget, "<<Copy>>"))
            self.root.bind_class(cls, "<Control-KeyPress>", physical_ctrl_shortcut, add="+")
            self.root.bind_class(cls, "<Button-3>", popup)

    def _bind_events(self):
        self.main_type_var.trace_add("write", lambda *args: self.update_form_by_type())
        self.finish_type_var.trace_add("write", lambda *args: self.update_form_by_type())
        self.quantity_var.trace_add("write", lambda *args: self.update_total())
        self.unit_price_var.trace_add("write", lambda *args: self.update_total())
        self.search_var.trace_add(
            "write",
            lambda *args: self.search_records() if self.search_window and self.search_window.winfo_exists() and self.search_window.state() != "withdrawn" else None,
        )

    def update_form_by_type(self):
        main_type = self.main_type_var.get()
        if not main_type:
            self.finish_type_var.set("")
            self.finish_type_frame.grid_remove()
            self.payer_frame.grid_remove()
            self.shared_only_frame.grid_remove()
            self.shared_only_var.set(False)
        elif main_type == "تشطيب":
            self.finish_type_frame.grid()
            self.payer_frame.grid_remove()
            self.shared_only_frame.grid_remove()
            if not self.finish_type_var.get():
                self.finish_type_var.set("مشتريات")
            self.payer_var.set(self.user_name)
            self.shared_only_var.set(False)
            if self.finish_type_var.get() != "مشتريات":
                self.quantity_var.set("1")
        elif main_type == "فرش":
            self.finish_type_var.set("")
            self.finish_type_frame.grid_remove()
            self.payer_frame.grid()
            if self.accounting_mode == ACCOUNTING_EQUAL:
                self.payer_frame.label.configure(text="دافع البند")
                self.shared_only_frame.grid()
                if self.editing_record_id is None:
                    self.payer_var.set("")
            elif self.accounting_mode == ACCOUNTING_ITEMIZED:
                self.payer_frame.label.configure(text="البند على")
                self.shared_only_frame.grid_remove()
                self.shared_only_var.set(False)
                if self.editing_record_id is None:
                    self.payer_var.set("")
            else:
                self.payer_frame.grid_remove()
                self.shared_only_frame.grid_remove()
                self.payer_var.set(self.user_name)
                self.shared_only_var.set(False)
            if self.accounting_mode != ACCOUNTING_ITEMIZED and self.payer_var.get() not in self.payer_options:
                self.payer_var.set("")
        elif main_type == WEDDING_TYPE and self.accounting_mode == ACCOUNTING_ITEMIZED:
            self.finish_type_var.set("")
            self.finish_type_frame.grid_remove()
            self.payer_frame.grid()
            self.payer_frame.label.configure(text="البند على")
            self.shared_only_frame.grid_remove()
            self.shared_only_var.set(False)
            if self.editing_record_id is None:
                self.payer_var.set("")
        elif main_type == WEDDING_TYPE and (
            (self.accounting_mode == ACCOUNTING_SINGLE and self.wedding_policy == WEDDING_EQUAL)
            or self.accounting_mode == ACCOUNTING_EQUAL
        ):
            self.finish_type_var.set("")
            self.finish_type_frame.grid_remove()
            self.payer_frame.grid()
            self.payer_frame.label.configure(text="دافع البند")
            self.shared_only_frame.grid_remove()
            self.shared_only_var.set(False)
            if self.editing_record_id is None:
                self.payer_var.set("")
        else:
            self.finish_type_var.set("")
            self.finish_type_frame.grid_remove()
            self.payer_frame.grid_remove()
            self.shared_only_frame.grid_remove()
            self.payer_var.set(self.user_name)
            self.shared_only_var.set(False)

        if main_type in (WEDDING_TYPE, OTHER_EXPENSES_TYPE):
            self.room_frame.grid_remove()
            self.room_var.set("")
            self.quantity_frame.grid(row=3, column=1, sticky="nsew", padx=5, pady=5)
            self.unit_price_frame.grid(row=3, column=0, sticky="nsew", padx=5, pady=5)
        else:
            self.room_frame.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=5, pady=5)
            if not self.room_var.get():
                self.room_var.set(self.room_options[0] if self.room_options else "")
            self.quantity_frame.grid(row=4, column=1, sticky="nsew", padx=5, pady=5)
            self.unit_price_frame.grid(row=4, column=0, sticky="nsew", padx=5, pady=5)
        if not self.quantity_var.get().strip():
            self.quantity_var.set("1")
        self.update_total()
        self._refresh_dynamic_form_layout()

    def _refresh_dynamic_form_layout(self):
        def refresh():
            for shell_name in ("details_shell", "financial_shell"):
                shell = getattr(self, shell_name, None)
                if shell is not None:
                    try:
                        shell.refresh_size()
                    except Exception:
                        pass
            try:
                self.root.update_idletasks()
                bbox = self.main_canvas.bbox("all")
                self.main_canvas.configure(scrollregion=bbox)
            except Exception:
                pass

        try:
            self.root.after_idle(refresh)
        except Exception:
            refresh()

    def update_total(self):
        qty = parse_float(self.quantity_var.get(), 0)
        unit = parse_float(self.unit_price_var.get(), 0)
        self.total_var.set(f"{qty * unit:.2f}")

    def pick_invoice_file(self):
        paths = filedialog.askopenfilenames(
            title="اختر الفاتورة أو المرفق",
            filetypes=[
                ("الصور وPDF", "*.png *.jpg *.jpeg *.bmp *.gif *.webp *.pdf"),
                ("الصور", "*.png *.jpg *.jpeg *.bmp *.gif *.webp"),
                ("PDF", "*.pdf"),
            ],
        )
        if paths:
            self._attach_invoice_files(paths)

    def _attach_invoice_files(self, paths):
        clean_paths = []
        rejected = []
        for path in paths:
            if not path or not Path(path).exists():
                continue
            if not is_supported_attachment(path):
                rejected.append(Path(path).name)
                continue
            if not is_attachment_size_allowed(path):
                rejected.append(f"{Path(path).name} (أكبر من 100MB)")
                continue
            clean_paths.append(str(Path(path)))
        if not clean_paths:
            if rejected:
                messagebox.showwarning("مرفق غير مدعوم", "المرفقات المسموحة: صور أو PDF فقط.")
            return
        existing = {inv["path"] for inv in self.selected_invoice_sources}
        for path in clean_paths:
            if path not in existing:
                self.selected_invoice_sources.append({"path": path, "original_name": Path(path).name})
                existing.add(path)
        self.invoice_removed_in_edit = False
        self.invoice_name_var.set(format_invoice_names(self.selected_invoice_sources))
        if rejected:
            self.status_var.set("تم اختيار المرفقات المدعومة فقط")
            messagebox.showwarning("مرفقات تم تجاهلها", "تم تجاهل أي ملفات ليست صورًا أو PDF أو أكبر من 100MB.")
        else:
            self.status_var.set("تم اختيار مرفق جديد")

    def _register_invoice_drop_target(self, *widgets):
        if DND_FILES is None:
            return

        def handle_drop(event):
            try:
                paths = self.root.tk.splitlist(event.data)
            except Exception:
                paths = [event.data]
            self._attach_invoice_files(paths)
            return "break"

        for widget in widgets:
            try:
                widget.drop_target_register(DND_FILES)
                widget.dnd_bind("<<Drop>>", handle_drop)
            except Exception:
                pass

    def open_selected_invoice_source(self):
        if self.selected_invoice_sources:
            for inv in self.selected_invoice_sources:
                open_file(inv["path"])
        else:
            messagebox.showinfo("معلومة", "لا يوجد ملف مرفق حاليًا في نموذج الإدخال.")

    def remove_selected_invoice(self):
        self.selected_invoice_sources = []
        self.invoice_name_var.set("لا يوجد ملف مرفق")
        if self.editing_record_id is not None:
            self.invoice_removed_in_edit = True
        self.status_var.set("تمت إزالة المرفق من النموذج")

    def refresh_room_option_widgets(self):
        values = self.room_options or [""]
        if hasattr(self, "room_combo"):
            self.room_combo["values"] = values
            if self.room_var.get() not in values:
                self.room_var.set(values[0])
        if hasattr(self, "filter_room_combo"):
            self.filter_room_combo["values"] = ["الكل"] + values
            if self.filter_room_var.get() not in (["الكل"] + values):
                self.filter_room_var.set("الكل")

    def _ensure_rooms_window(self):
        if self.rooms_window and self.rooms_window.winfo_exists():
            return
        w = tk.Toplevel(self.root)
        self.rooms_window = w
        w.title("إدارة الأماكن")
        w.configure(bg=BG_APP)
        w.withdraw()
        w.protocol("WM_DELETE_WINDOW", w.withdraw)
        c = self._popup_card(w, "إدارة الأماكن")
        top = tk.Frame(c, bg=SURFACE)
        top.pack(fill="x")
        self.new_room_var = tk.StringVar()
        self._action_button(top, "إضافة جديدة", self.add_room_option, width=112).pack(side="right", padx=(0, 8))
        self._make_text_entry(top, self.new_room_var).pack(side="right", fill="x", expand=True)
        box = tk.Frame(c, bg=SURFACE)
        box.pack(fill="both", expand=True, pady=(12, 0))
        self.rooms_listbox = tk.Listbox(box, font=(FONT_FAMILY, 11), justify="right", bg=INPUT_BG, fg=TEXT, relief="flat", bd=0, highlightthickness=1, highlightbackground=INPUT_BORDER, selectbackground="#D8E7D8", selectforeground=TEXT)
        self.rooms_listbox.pack(side="right", fill="both", expand=True)
        self._attach_vertical_scrollbar(self.rooms_listbox, box, side="left", fill="y")
        actions = tk.Frame(c, bg=SURFACE)
        actions.pack(fill="x", pady=(12, 0))
        self._action_button(actions, "حذف المحدد", self.delete_room_option, kind="danger", width=120).pack(side="right", padx=4)
        self._action_button(actions, "استعادة مجمع", self.reset_room_options, kind="secondary", width=134).pack(side="right", padx=4)
        tk.Label(c, text="يمكنك إضافة أو حذف مكونات الشقة لاستخدامها في البنود لاحقًا. خانة مجمع موجودة دائمًا ولا يمكن حذفها.", bg="#DDE8D8", fg=TEXT, font=(FONT_FAMILY, 9), anchor="e", justify="right", padx=10, pady=7).pack(fill="x", pady=(8,0))
        self._fit_toplevel(w, 620, 520, 0.42, 0.62)

    def refresh_rooms_listbox(self):
        if not hasattr(self, "rooms_listbox"):
            return
        self.rooms_listbox.delete(0, "end")
        for room in self.room_options:
            self.rooms_listbox.insert("end", room)

    def open_rooms_window(self):
        self._ensure_rooms_window()
        self.refresh_rooms_listbox()
        self._reveal_window(self.rooms_window)

    def add_room_option(self):
        room = getattr(self, "new_room_var", tk.StringVar()).get().strip()
        if not room:
            return
        if room in self.room_options:
            messagebox.showinfo("معلومة", "هذا المكان موجود بالفعل.")
            return
        self.room_options = normalize_room_options(self.room_options + [room])
        save_room_options(self.room_options)
        self.refresh_room_option_widgets()
        self.refresh_rooms_listbox()
        self.new_room_var.set("")
        self.status_var.set("تمت إضافة المكان بنجاح")

    def delete_room_option(self):
        if not hasattr(self, "rooms_listbox"):
            return
        sel = self.rooms_listbox.curselection()
        if not sel:
            messagebox.showinfo("معلومة", "اختر مكانًا أولًا.")
            return
        room = self.rooms_listbox.get(sel[0])
        if room == DEFAULT_ROOM_NAME:
            messagebox.showinfo("معلومة", "خانة مجمع موجودة دائمًا ولا يمكن حذفها.")
            return
        used = self.conn.execute("SELECT COUNT(*) c FROM records WHERE room=?", (room,)).fetchone()["c"]
        if used:
            messagebox.showwarning("غير متاح", f"لا يمكن حذف '{room}' لأنه مستخدم في {used} بند/بنود.")
            return
        if not messagebox.askyesno("تأكيد", f"هل تريد حذف المكان: {room}؟"):
            return
        self.room_options = [r for r in self.room_options if r != room]
        self.room_options = normalize_room_options(self.room_options)
        save_room_options(self.room_options)
        self.refresh_room_option_widgets()
        self.refresh_rooms_listbox()
        self.status_var.set("تم حذف المكان")

    def reset_room_options(self):
        if not messagebox.askyesno("تأكيد", "هل تريد استعادة الأماكن إلى خانة مجمع فقط؟"):
            return
        self.room_options = list(DEFAULT_ROOM_OPTIONS)
        save_room_options(self.room_options)
        self.refresh_room_option_widgets()
        self.refresh_rooms_listbox()
        self.status_var.set("تمت استعادة الأماكن الافتراضية")

    def refresh_payment_target_widgets(self):
        if hasattr(self, "advance_direction_frame"):
            self.advance_direction_frame.widget["values"] = self.payment_targets
        if self.advance_direction_var.get() not in self.payment_targets:
            self.advance_direction_var.set(self.payment_targets[0])
        if hasattr(self, "targets_listbox"):
            self.targets_listbox.delete(0, "end")
            for target in self.payment_targets:
                self.targets_listbox.insert("end", target)

    def _ensure_support_window(self):
        if self.support_window and self.support_window.winfo_exists():
            return
        w = tk.Toplevel(self.root)
        self.support_window = w
        w.title("إبلاغ أو اقتراح")
        w.configure(bg=BG_APP)
        w.withdraw()
        w.protocol("WM_DELETE_WINDOW", w.withdraw)

        c = self._popup_card(w, "إبلاغ أو اقتراح", padx=20, pady=18)
        tk.Label(
            c,
            text="استخدم هذا النموذج لإرسال مشكلة تقنية، طلب ميزة، اقتراح تحسين، أو ملاحظة عامة.",
            bg=SURFACE,
            fg=TEXT_2,
            font=(FONT_FAMILY, 10),
            anchor="e",
            justify="right",
            wraplength=620,
        ).pack(fill="x", pady=(0, 10))

        form = ttk.Frame(c, style="Card.TFrame")
        form.pack(fill="both", expand=True)
        form.columnconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)
        self._add_labeled_entry(form, "الاسم", self.support_name_var, 0, 1)
        self._add_labeled_entry(form, "البريد الإلكتروني", self.support_email_var, 0, 0)
        self._add_labeled_combobox(form, "نوع الرسالة", self.support_type_var, SUPPORT_MESSAGE_TYPES, 1, 1)
        self._add_labeled_entry(form, "عنوان مختصر", self.support_subject_var, 1, 0)

        details_frame = ttk.Frame(form, style="Card.TFrame")
        details_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=7, pady=7)
        ttk.Label(details_frame, text="التفاصيل", anchor="e", style="Field.TLabel").pack(fill="x", pady=(0, 4))
        self.support_details_text = self._make_notes_widget(details_frame, rows=7)
        self.support_details_text.pack(fill="both", expand=True, pady=(4, 10))

        self.support_status_label = tk.Label(c, text="", bg=SURFACE, fg=TEXT_2, font=(FONT_FAMILY, 9), anchor="e", justify="right", wraplength=620)
        self.support_status_label.pack(fill="x", pady=(2, 0))
        actions = tk.Frame(c, bg=SURFACE)
        actions.pack(fill="x", pady=(12, 0))
        self.support_submit_button = self._action_button(actions, "إرسال", self.submit_support_message, width=112)
        self.support_submit_button.pack(side="right", padx=4)
        self._action_button(actions, "إغلاق", w.withdraw, kind="secondary", width=96).pack(side="right", padx=4)
        self._fit_toplevel(w, 860, 760, 0.64, 0.90)

    def open_support_window(self):
        self._ensure_support_window()
        self.support_status_label.configure(text="")
        self._fit_toplevel(self.support_window, 860, 760, 0.64, 0.90)
        self._reveal_window(self.support_window)

    def _set_support_form_enabled(self, enabled):
        state = "normal" if enabled else "disabled"
        for widget in (
            getattr(self, "support_submit_button", None),
            getattr(self, "support_details_text", None),
        ):
            if not widget:
                continue
            try:
                widget.configure(state=state)
            except Exception:
                pass

    def _support_payload(self):
        if isinstance(self.support_details_text, tk.Text):
            details = self.support_details_text.get("1.0", "end-1c")
        else:
            details = self.support_details_text.get()
        return {
            "name": self.support_name_var.get().strip(),
            "email": self.support_email_var.get().strip(),
            "type": self.support_type_var.get().strip(),
            "subject": self.support_subject_var.get().strip(),
            "details": details.strip(),
            "appName": APP_TITLE,
            "appVersion": APP_VERSION,
            "projectName": self.project_name,
            "sentAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

    def submit_support_message(self):
        if self.support_sending_var.get():
            return
        url = load_support_web_app_url()
        if not url:
            messagebox.showwarning("الخدمة غير مفعلة", "خدمة استقبال البلاغات غير مفعلة في هذه النسخة.")
            return
        if not is_valid_support_web_app_url(url):
            messagebox.showwarning("إعداد غير آمن", "رابط استقبال البلاغات غير صالح أو غير آمن.")
            return
        payload = self._support_payload()
        if not payload["name"]:
            messagebox.showerror("بيانات ناقصة", "من فضلك أدخل الاسم.")
            return
        if "@" not in payload["email"] or "." not in payload["email"].split("@")[-1]:
            messagebox.showerror("بيانات غير صحيحة", "من فضلك أدخل بريد إلكتروني صحيح.")
            return
        if payload["type"] not in SUPPORT_MESSAGE_TYPES:
            messagebox.showerror("بيانات ناقصة", "من فضلك اختر نوع الرسالة.")
            return
        if not payload["subject"]:
            messagebox.showerror("بيانات ناقصة", "من فضلك أدخل عنوانًا مختصرًا.")
            return
        if not payload["details"]:
            messagebox.showerror("بيانات ناقصة", "من فضلك اكتب تفاصيل الرسالة.")
            return
        length_ok, field_name, limit = validate_support_text_lengths(payload)
        if not length_ok:
            field_labels = {"name": "الاسم", "email": "البريد الإلكتروني", "subject": "العنوان", "details": "التفاصيل"}
            messagebox.showerror(
                "بيانات طويلة",
                f"حقل {field_labels.get(field_name, field_name)} أطول من الحد المسموح ({limit} حرف).",
            )
            return

        self.support_status_label.configure(text="جاري إرسال الرسالة...")
        self.support_sending_var.set(True)
        self._set_support_form_enabled(False)

        def worker():
            try:
                body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
                req = urlrequest.Request(
                    url,
                    data=body,
                    headers={"Content-Type": "application/json; charset=utf-8"},
                    method="POST",
                )
                with urlrequest.urlopen(req, timeout=25) as response:
                    raw = response.read().decode("utf-8")
                result = json.loads(raw or "{}")
                if not result.get("ok"):
                    raise RuntimeError(result.get("error") or "تعذر إرسال الرسالة.")
                ticket_id = str(result.get("ticketId") or "").strip()
                self.root.after(0, lambda: self._support_send_finished(ticket_id, None))
            except (urlerror.URLError, TimeoutError) as exc:
                message = f"تعذر الاتصال بخدمة استقبال البلاغات:\n{exc}"
                self.root.after(0, lambda msg=message: self._support_send_finished("", msg))
            except Exception as exc:
                message = str(exc)
                self.root.after(0, lambda msg=message: self._support_send_finished("", msg))

        threading.Thread(target=worker, daemon=True).start()

    def _support_send_finished(self, ticket_id, error_message):
        self.support_sending_var.set(False)
        self._set_support_form_enabled(True)
        if error_message:
            self.support_status_label.configure(text="فشل الإرسال. راجع الاتصال أو رابط الاستقبال.")
            messagebox.showerror("فشل الإرسال", error_message)
            return
        if ticket_id:
            self.support_status_label.configure(text=f"تم إرسال الرسالة بنجاح. رقم المتابعة: {ticket_id}")
            messagebox.showinfo("تم الإرسال", f"تم استلام رسالتك بنجاح.\nرقم المتابعة: {ticket_id}")
        else:
            self.support_status_label.configure(text="تم إرسال الرسالة بنجاح.")
            messagebox.showinfo("تم الإرسال", "تم استلام رسالتك بنجاح.")
        self.support_subject_var.set("")
        try:
            self.support_details_text.delete("1.0", "end")
        except Exception:
            self.support_details_text.delete(0, "end")

    def _ensure_settings_window(self):
        if self.settings_window and self.settings_window.winfo_exists():
            return
        w = tk.Toplevel(self.root)
        self.settings_window = w
        w.title("إعدادات البرنامج")
        w.configure(bg=BG_APP)
        w.withdraw()
        w.protocol("WM_DELETE_WINDOW", w.withdraw)

        c = self._popup_card(w, "إعدادات البرنامج")

        form = ttk.Frame(c, style="Card.TFrame")
        form.pack(fill="x")
        form.columnconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)
        self.settings_project_var = tk.StringVar(value=self.project_name)
        self.settings_user_var = tk.StringVar(value=self.user_name)
        self.settings_other_var = tk.StringVar(value=self.other_party_name)
        self.settings_party1_role_var = tk.StringVar(value=self.party1_role)
        self.settings_party2_role_var = tk.StringVar(value=self.party2_role)
        self.settings_currency_var = tk.StringVar(value=self.base_currency)
        self.settings_accounting_mode_var = tk.StringVar(value=self.accounting_mode)
        self.settings_wedding_policy_var = tk.StringVar(value=self.wedding_policy)
        self.settings_split_mode_var = tk.StringVar(value=self.split_mode)
        self.settings_party1_percent_var = tk.StringVar(value=f"{self.party1_percent:.0f}")
        self.settings_party2_percent_var = tk.StringVar(value=f"{self.party2_percent:.0f}")
        self._add_labeled_entry(form, "اسم المشروع", self.settings_project_var, 0, 0, colspan=2)
        self._add_labeled_entry(form, "اسم الطرف الأول", self.settings_user_var, 1, 1)
        self._add_labeled_entry(form, "اسم الطرف الثاني", self.settings_other_var, 1, 0)
        self._add_labeled_combobox(form, "صفة الطرف الأول", self.settings_party1_role_var, PARTY_ROLE_OPTIONS, 2, 1)
        self._add_labeled_combobox(form, "صفة الطرف الثاني", self.settings_party2_role_var, PARTY_ROLE_OPTIONS, 2, 0)
        self._add_labeled_combobox(form, "العملة الأساسية", self.settings_currency_var, LOAN_CURRENCIES, 3, 1)
        self._add_labeled_combobox(form, "نظام المحاسبة", self.settings_accounting_mode_var, ACCOUNTING_MODE_OPTIONS, 3, 0)
        self.settings_wedding_policy_frame = self._add_labeled_combobox(form, "حفل الزفاف في هذا النظام", self.settings_wedding_policy_var, WEDDING_POLICY_OPTIONS, 4, 0, colspan=2)
        ttk.Label(
            form,
            text="اختيار نظام المحاسبة يغيّر الحقول والحسابات: الطرف الواحد، المناصفة 50/50، أو تحديد الطرف المسؤول لكل بند.",
            anchor="e",
            justify="right",
            style="HeaderSub.TLabel",
            wraplength=620,
        ).grid(row=5, column=0, columnspan=2, sticky="ew", padx=5, pady=(0, 8))

        def sync_settings_split(*_):
            mode, np1, np2 = normalize_split(self.settings_split_mode_var.get(), self.settings_party1_percent_var.get(), self.settings_party2_percent_var.get())
            if mode != SPLIT_MODE_CUSTOM:
                self.settings_party1_percent_var.set(f"{np1:.0f}")
                self.settings_party2_percent_var.set(f"{np2:.0f}")

        self.settings_split_mode_var.trace_add("write", sync_settings_split)
        self.settings_accounting_mode_var.trace_add("write", lambda *_: self.update_settings_accounting_visibility())

        targets_frame = tk.Frame(c, bg=SURFACE, padx=0, pady=0)
        targets_frame.pack(fill="both", expand=True, pady=(12, 0))
        tk.Label(targets_frame, text="جهات التسديد", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 11, "bold"), anchor="e").pack(fill="x", pady=(0, 6))
        top = tk.Frame(targets_frame, bg=SURFACE)
        top.pack(fill="x")
        self.new_target_var = tk.StringVar()
        self._action_button(top, "إضافة جهة", self.add_payment_target, width=112).pack(side="right", padx=(0, 8))
        self._make_text_entry(top, self.new_target_var).pack(side="right", fill="x", expand=True)
        box = tk.Frame(targets_frame, bg=SURFACE)
        box.pack(fill="both", expand=True, pady=(10, 0))
        self.targets_listbox = tk.Listbox(box, font=(FONT_FAMILY, 11), justify="right", height=6, bg=INPUT_BG, fg=TEXT, relief="flat", bd=0, highlightthickness=1, highlightbackground=INPUT_BORDER, selectbackground="#D8E7D8", selectforeground=TEXT)
        self.targets_listbox.pack(side="right", fill="both", expand=True)
        self._attach_vertical_scrollbar(self.targets_listbox, box, side="left", fill="y")

        actions = tk.Frame(c, bg=SURFACE)
        actions.pack(fill="x", pady=(12, 0))
        self._action_button(actions, "حفظ الإعدادات", self.save_general_settings, width=136).pack(side="right", padx=4)
        self._action_button(actions, "حذف الجهة", self.delete_payment_target, kind="danger", width=112).pack(side="right", padx=4)
        self._action_button(actions, "إدارة الأماكن", self.open_rooms_window, kind="secondary", width=124).pack(side="right", padx=4)
        tk.Label(c, text=COPYRIGHT_NOTICE, bg=SURFACE, fg=TEXT_2, font=(FONT_FAMILY, 9), anchor="center", justify="center").pack(fill="x", pady=(12, 0))
        self._fit_toplevel(w, 900, 820, 0.68, 0.92)

    def open_settings_window(self):
        self._ensure_settings_window()
        self.settings_project_var.set(self.project_name)
        self.settings_user_var.set(self.user_name)
        self.settings_other_var.set(self.other_party_name)
        self.settings_party1_role_var.set(self.party1_role)
        self.settings_party2_role_var.set(self.party2_role)
        self.settings_currency_var.set(self.base_currency)
        self.settings_accounting_mode_var.set(self.accounting_mode)
        self.settings_wedding_policy_var.set(self.wedding_policy)
        self.update_settings_accounting_visibility()
        self.settings_split_mode_var.set(self.split_mode)
        self.settings_party1_percent_var.set(f"{self.party1_percent:.0f}")
        self.settings_party2_percent_var.set(f"{self.party2_percent:.0f}")
        self.refresh_payment_target_widgets()
        self._fit_toplevel(self.settings_window, 900, 820, 0.68, 0.92)
        self._reveal_window(self.settings_window)

    def update_settings_accounting_visibility(self):
        if not hasattr(self, "settings_wedding_policy_frame"):
            return
        if self.settings_accounting_mode_var.get() in (ACCOUNTING_SINGLE, ACCOUNTING_ITEMIZED):
            self.settings_wedding_policy_frame.grid()
        else:
            self.settings_wedding_policy_frame.grid_remove()

    def save_general_settings(self):
        project = self.settings_project_var.get().strip() or APP_TITLE
        user = self.settings_user_var.get().strip()
        other = self.settings_other_var.get().strip()
        if not user or not other:
            messagebox.showerror("خطأ", "من فضلك أدخل أسماء الطرفين.")
            return
        if user == other:
            messagebox.showerror("خطأ", "أسماء الطرفين لازم تكون مختلفة.")
            return
        old_user, old_other = self.user_name, self.other_party_name
        split_mode, party1_percent, party2_percent = normalize_split(SPLIT_MODE_EQUAL, 50, 50)
        self.project_name = project
        self.user_name = user
        self.other_party_name = other
        self.party1_role = self.settings_party1_role_var.get()
        self.party2_role = self.settings_party2_role_var.get()
        self.base_currency = self.settings_currency_var.get()
        self.accounting_mode = self.settings_accounting_mode_var.get()
        self.wedding_policy = self.settings_wedding_policy_var.get()
        self.split_mode = split_mode
        self.party1_percent = party1_percent
        self.party2_percent = party2_percent
        self.payer_options = [self.user_name, self.other_party_name]
        save_project_name(project)
        save_party_names(user, other)
        save_party_roles(self.party1_role, self.party2_role)
        save_base_currency(self.base_currency)
        save_split_settings(self.split_mode, self.party1_percent, self.party2_percent)
        save_accounting_settings(self.accounting_mode, self.wedding_policy)
        self.conn.execute("UPDATE records SET payer=? WHERE payer=?", (self.user_name, old_user))
        self.conn.execute("UPDATE records SET payer=? WHERE payer=?", (self.other_party_name, old_other))
        self.conn.commit()
        self.root.title(self.project_name)
        self.project_title_var.set(self.project_name)
        if hasattr(self, "payer_frame"):
            self.payer_frame.widget["values"] = self.payer_options
        if self.payer_var.get() not in self.payer_options:
            self.payer_var.set(self.user_name)
        self.loan_currency_var.set(self.base_currency)
        self.loan_payment_currency_var.set(self.base_currency)
        self.update_form_by_type()
        self.refresh_summary()
        self.status_var.set("تم حفظ إعدادات البرنامج")
        messagebox.showinfo("تم", "تم حفظ الإعدادات بنجاح.")

    def add_payment_target(self):
        target = self.new_target_var.get().strip()
        if not target:
            return
        if target in self.payment_targets:
            messagebox.showinfo("معلومة", "هذه الجهة موجودة بالفعل.")
            return
        self.payment_targets = normalize_text_options([*self.payment_targets, target], [])
        save_payment_targets(self.payment_targets)
        self.new_target_var.set("")
        self.refresh_payment_target_widgets()

    def delete_payment_target(self):
        if not hasattr(self, "targets_listbox"):
            return
        sel = self.targets_listbox.curselection()
        if not sel:
            messagebox.showinfo("معلومة", "اختر جهة أولًا.")
            return
        target = self.targets_listbox.get(sel[0])
        used = self.conn.execute("SELECT COUNT(*) c FROM advances WHERE direction=?", (target,)).fetchone()["c"]
        if used:
            messagebox.showwarning("غير متاح", f"لا يمكن حذف '{target}' لأنها مستخدمة في {used} تسديد/تسديدات.")
            return
        remaining = [t for t in self.payment_targets if t != target]
        if not remaining:
            messagebox.showwarning("غير متاح", "لازم تفضل جهة تسديد واحدة على الأقل.")
            return
        self.payment_targets = remaining
        save_payment_targets(self.payment_targets)
        self.refresh_payment_target_widgets()

    def export_backup(self):
        ensure_dirs()
        default_name = f"apartment_manager_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"
        output = filedialog.asksaveasfilename(
            title="تصدير نسخة احتياطية",
            defaultextension=".zip",
            filetypes=[("ZIP", "*.zip")],
            initialdir=str(EXPORTS_DIR),
            initialfile=default_name,
        )
        if not output:
            return
        try:
            output_path = Path(output).resolve()
            with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zf:
                for path in iter_backup_source_files(output_path):
                    zf.write(path, Path("app_data") / path.relative_to(DATA_DIR))
            self.status_var.set(f"تم تصدير النسخة الاحتياطية: {Path(output).name}")
            messagebox.showinfo("تم", f"تم تصدير النسخة الاحتياطية بنجاح:\n{output}")
        except Exception as e:
            messagebox.showerror("خطأ", f"تعذر تصدير النسخة الاحتياطية:\n{e}")

    def restore_backup(self):
        path = filedialog.askopenfilename(
            title="استيراد نسخة احتياطية",
            filetypes=[("ZIP", "*.zip"), ("كل الملفات", "*.*")],
        )
        if not path:
            return
        if not messagebox.askyesno(
            "تأكيد الاستعادة",
            "سيتم استبدال بيانات البرنامج الحالية بالنسخة المختارة.\nسيتم حفظ نسخة من البيانات الحالية قبل الاستبدال.\n\nهل تريد المتابعة؟",
        ):
            return
        temp_dir = DATA_DIR.parent / f"_restore_temp_{uuid.uuid4().hex}"
        old_backup = DATA_DIR.parent / f"app_data_before_restore_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        moved_current_data = False
        conn_closed = False
        try:
            temp_dir.mkdir(parents=True, exist_ok=True)
            extract_backup_safely(path, temp_dir)
            extracted_data = temp_dir / "app_data"
            if not extracted_data.exists():
                messagebox.showerror("خطأ", "ملف النسخة الاحتياطية غير صالح: لا يحتوي على مجلد app_data.")
                return
            try:
                self.conn.close()
                conn_closed = True
            except Exception:
                pass
            if DATA_DIR.exists():
                shutil.move(str(DATA_DIR), str(old_backup))
                moved_current_data = True
            shutil.move(str(extracted_data), str(DATA_DIR))
            self.project_name = load_project_name()
            self.user_name, self.other_party_name = load_party_names()
            if not self.user_name or not self.other_party_name:
                self.user_name, self.other_party_name = DEFAULT_USER_NAME, DEFAULT_OTHER_PARTY_NAME
            self.party1_role, self.party2_role = load_party_roles()
            self.base_currency = load_base_currency()
            self.split_mode, self.party1_percent, self.party2_percent = load_split_settings()
            self.accounting_mode, self.wedding_policy = load_accounting_settings()
            self.room_options = load_room_options()
            self.payment_targets = load_payment_targets()
            self.payer_options = [self.user_name, self.other_party_name]
            self.conn = connect_db()
            self._migrate_payer_names()
            self.root.title(self.project_name)
            self.project_title_var.set(self.project_name)
            self.refresh_room_option_widgets()
            self.refresh_payment_target_widgets()
            if hasattr(self, "payer_frame"):
                self.payer_frame.widget["values"] = self.payer_options
            self.clear_form()
            self.refresh_summary()
            self.status_var.set("تم استيراد النسخة الاحتياطية")
            messagebox.showinfo("تم", f"تم استيراد النسخة الاحتياطية بنجاح.\nتم حفظ البيانات السابقة في:\n{old_backup}")
        except Exception as e:
            if moved_current_data and not DATA_DIR.exists() and old_backup.exists():
                try:
                    shutil.move(str(old_backup), str(DATA_DIR))
                except Exception:
                    pass
            if conn_closed:
                try:
                    self.conn = connect_db()
                except Exception:
                    pass
            messagebox.showerror("خطأ", f"تعذر استيراد النسخة الاحتياطية:\n{e}")
        finally:
            if temp_dir.exists():
                shutil.rmtree(temp_dir, ignore_errors=True)

    def party1_ratio(self):
        return max(0.0, min(1.0, float(self.party1_percent or 0) / 100.0))

    def party2_ratio(self):
        return max(0.0, min(1.0, float(self.party2_percent or 0) / 100.0))

    def party1_share_amount(self, total):
        return float(total or 0) * self.party1_ratio()

    def party2_share_amount(self, total):
        return float(total or 0) * self.party2_ratio()

    def record_shared_split(self, row):
        value = row["shared_split"]
        return 1 if value is None else int(value)

    def record_party1_share(self, row_or_type, total=None, payer=None, shared_split=1):
        if isinstance(row_or_type, sqlite3.Row):
            main_type = row_or_type["main_type"]
            total = row_or_type["total"]
            payer = row_or_type["payer"] or self.user_name
            shared_split = self.record_shared_split(row_or_type)
        else:
            main_type = row_or_type
        total = float(total or 0)
        if main_type in ("تشطيب", OTHER_EXPENSES_TYPE):
            return total
        if self.accounting_mode == ACCOUNTING_SINGLE:
            if main_type == WEDDING_TYPE:
                if self.wedding_policy == WEDDING_EQUAL:
                    return total / 2.0
                if self.wedding_policy == WEDDING_PARTY2:
                    return 0.0
            return total
        if self.accounting_mode == ACCOUNTING_ITEMIZED and main_type in ("فرش", WEDDING_TYPE):
            return total if payer == self.user_name else 0.0
        if main_type == "فرش" and shared_split == 0:
            return total
        return total / 2.0

    def record_party2_share(self, row_or_type, total=None, payer=None, shared_split=1):
        if isinstance(row_or_type, sqlite3.Row):
            main_type = row_or_type["main_type"]
            total = row_or_type["total"]
            payer = row_or_type["payer"] or self.user_name
            shared_split = self.record_shared_split(row_or_type)
        else:
            main_type = row_or_type
        total = float(total or 0)
        if main_type in ("تشطيب", OTHER_EXPENSES_TYPE):
            return 0.0
        if self.accounting_mode == ACCOUNTING_SINGLE:
            if main_type == WEDDING_TYPE:
                if self.wedding_policy == WEDDING_EQUAL:
                    return total / 2.0
                if self.wedding_policy == WEDDING_PARTY2:
                    return total
            return 0.0
        if self.accounting_mode == ACCOUNTING_ITEMIZED and main_type in ("فرش", WEDDING_TYPE):
            return total if payer == self.other_party_name else 0.0
        if main_type == "فرش" and shared_split == 0:
            return 0.0
        return total / 2.0

    def record_accounting_label(self, row):
        if self.accounting_mode == ACCOUNTING_ITEMIZED and row["main_type"] in ("فرش", WEDDING_TYPE):
            return f"على {row['payer'] or self.user_name}"
        if self.accounting_mode == ACCOUNTING_SINGLE:
            if row["main_type"] == WEDDING_TYPE:
                return self.wedding_policy.replace("حفل الزفاف ", "")
            return f"على {self.user_name}"
        if row["main_type"] == "فرش" and self.record_shared_split(row) == 0:
            return f"على {self.user_name} بالكامل"
        return "مشترك 50/50"

    def has_party1_shared_balance(self):
        return not (self.accounting_mode == ACCOUNTING_SINGLE and self.wedding_policy == WEDDING_PARTY1)

    def update_summary_visibility(self):
        if not hasattr(self, "summary_card_boxes"):
            return
        packed_ribbon = any(
            self.summary_card_boxes.get(key) and self.summary_card_boxes[key][0].winfo_manager() == "pack"
            for key in self.summary_card_boxes
        )
        if packed_ribbon:
            return
        hide = not self.has_party1_shared_balance()
        for key in ("person_cost", "balance"):
            box_info = self.summary_card_boxes.get(key)
            if not box_info:
                continue
            box, row, col = box_info
            if hide:
                box.grid_remove()
            else:
                box.grid(row=row, column=col, sticky="nsew", padx=4, pady=4)

    def validate_form(self):
        if self.main_type_var.get() not in TYPE_OPTIONS:
            messagebox.showerror("خطأ", "من فضلك اختر نوع البند.")
            return False
        if not self.item_name_var.get().strip():
            messagebox.showerror("خطأ", "من فضلك أدخل اسم البند.")
            return False
        if self.main_type_var.get() not in (WEDDING_TYPE, OTHER_EXPENSES_TYPE) and not self.room_var.get().strip():
            messagebox.showerror("خطأ", "من فضلك اختر المكان.")
            return False
        if self.main_type_var.get() == "تشطيب" and not self.finish_type_var.get().strip():
            messagebox.showerror("خطأ", "من فضلك اختر تصنيف التشطيب.")
            return False
        if self.accounting_mode == ACCOUNTING_ITEMIZED and self.main_type_var.get() in ("فرش", WEDDING_TYPE):
            if self.payer_var.get() not in self.payer_options:
                messagebox.showerror("خطأ", "من فضلك اختر البند على أي طرف.")
                return False
        if (
            self.main_type_var.get() == WEDDING_TYPE
            and (
                (self.accounting_mode == ACCOUNTING_SINGLE and self.wedding_policy == WEDDING_EQUAL)
                or self.accounting_mode == ACCOUNTING_EQUAL
            )
        ):
            if self.payer_var.get() not in self.payer_options:
                messagebox.showerror("خطأ", "من فضلك اختر دافع بند حفل الزفاف.")
                return False
        qty = parse_float(self.quantity_var.get(), None)
        unit = parse_float(self.unit_price_var.get(), None)
        if qty is None or qty < 0:
            messagebox.showerror("خطأ", "الكمية غير صحيحة.")
            return False
        if unit is None or unit < 0:
            messagebox.showerror("خطأ", "سعر الوحدة غير صحيح.")
            return False
        return True

    def delete_invoice_file_if_orphan(self, path, excluding_record_id=None):
        if not path:
            return
        try:
            resolved = Path(path).resolve()
            resolved.relative_to(INVOICES_DIR.resolve())
        except Exception:
            return
        params = [path]
        query = "SELECT COUNT(*) AS c FROM record_invoices WHERE invoice_path=?"
        if excluding_record_id is not None:
            query += " AND record_id <> ?"
            params.append(excluding_record_id)
        count = self.conn.execute(query, params).fetchone()["c"]
        if count == 0 and resolved.exists():
            try:
                resolved.unlink()
            except Exception:
                pass

    def get_record_invoices(self, record_id):
        rows = self.conn.execute(
            "SELECT invoice_path, invoice_original_name FROM record_invoices WHERE record_id=? ORDER BY id",
            (record_id,),
        ).fetchall()
        return [{"path": r["invoice_path"], "original_name": r["invoice_original_name"]} for r in rows]

    def _prepare_selected_invoices(self, item_name):
        invoices = []
        for inv in self.selected_invoice_sources:
            source = inv.get("path")
            if not source:
                continue
            if not os.path.exists(source):
                continue
            try:
                is_stored = os.path.exists(source) and Path(source).resolve().is_relative_to(INVOICES_DIR.resolve())
            except Exception:
                is_stored = False
            if is_stored:
                invoices.append({"path": source, "original_name": inv.get("original_name") or Path(source).name})
            else:
                copied_path, copied_name = copy_invoice_file(source, item_name)
                if copied_path:
                    invoices.append({"path": copied_path, "original_name": copied_name})
        return invoices

    def _replace_record_invoices(self, record_id, invoices):
        self.conn.execute("DELETE FROM record_invoices WHERE record_id=?", (record_id,))
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for inv in invoices:
            self.conn.execute(
                "INSERT INTO record_invoices (record_id, invoice_path, invoice_original_name, created_at) VALUES (?, ?, ?, ?)",
                (record_id, inv["path"], inv.get("original_name"), now),
            )
        first = invoices[0] if invoices else None
        self.conn.execute(
            "UPDATE records SET invoice_path=?, invoice_original_name=? WHERE id=?",
            (first["path"] if first else None, first.get("original_name") if first else None, record_id),
        )

    def save_record(self):
        if not self.validate_form():
            return
        qty = parse_float(self.quantity_var.get(), 0)
        unit = parse_float(self.unit_price_var.get(), 0)
        total = qty * unit
        notes = self._get_text_value(self.notes_text)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        item_name = self.item_name_var.get().strip()

        invoices = self._prepare_selected_invoices(item_name)
        first_invoice = invoices[0] if invoices else None

        if self.accounting_mode == ACCOUNTING_ITEMIZED and self.main_type_var.get() in ("فرش", WEDDING_TYPE):
            payer = self.payer_var.get()
        elif (
            self.main_type_var.get() == WEDDING_TYPE
            and (
                (self.accounting_mode == ACCOUNTING_SINGLE and self.wedding_policy == WEDDING_EQUAL)
                or self.accounting_mode == ACCOUNTING_EQUAL
            )
        ):
            payer = self.payer_var.get()
        elif self.accounting_mode == ACCOUNTING_EQUAL and self.main_type_var.get() == "فرش":
            payer = self.payer_var.get()
        else:
            payer = self.user_name
        shared_split = 0 if (self.accounting_mode == ACCOUNTING_EQUAL and self.main_type_var.get() == "فرش" and self.shared_only_var.get()) else 1

        if self.editing_record_id is None:
            self.conn.execute(
                """
                INSERT INTO records
                (main_type, finish_type, payer, shared_split, item_name, room, quantity, unit_price, total, notes, invoice_path, invoice_original_name, created_at, updated_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    self.main_type_var.get(),
                    self.finish_type_var.get() or None,
                    payer,
                    shared_split,
                    item_name,
                    self.room_var.get().strip(),
                    qty,
                    unit,
                    total,
                    notes,
                    first_invoice["path"] if first_invoice else None,
                    first_invoice.get("original_name") if first_invoice else None,
                    now,
                    now,
                ),
            )
            record_id = self.conn.execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
            self._replace_record_invoices(record_id, invoices)
            self.conn.commit()
            self.status_var.set("تم حفظ البند بنجاح")
            messagebox.showinfo("تم", "تم حفظ البند بنجاح.")
        else:
            old = self.get_record(self.editing_record_id)
            old_invoices = self.get_record_invoices(self.editing_record_id)
            if not old_invoices and old["invoice_path"]:
                old_invoices = [{"path": old["invoice_path"], "original_name": old["invoice_original_name"]}]
            final_invoices = old_invoices
            if self.invoice_removed_in_edit:
                final_invoices = []
            if self.selected_invoice_sources:
                final_invoices = invoices
            first_invoice = final_invoices[0] if final_invoices else None
            self.conn.execute(
                """
                UPDATE records SET
                main_type=?, finish_type=?, payer=?, shared_split=?, item_name=?, room=?, quantity=?, unit_price=?, total=?, notes=?,
                invoice_path=?, invoice_original_name=?, updated_at=?
                WHERE id=?
                """,
                (
                    self.main_type_var.get(),
                    self.finish_type_var.get() or None,
                    payer,
                    shared_split,
                    item_name,
                    self.room_var.get().strip(),
                    qty,
                    unit,
                    total,
                    notes,
                    first_invoice["path"] if first_invoice else None,
                    first_invoice.get("original_name") if first_invoice else None,
                    now,
                    self.editing_record_id,
                ),
            )
            self._replace_record_invoices(self.editing_record_id, final_invoices)
            self.conn.commit()
            final_paths = {inv["path"] for inv in final_invoices}
            for old_invoice in old_invoices:
                old_path = old_invoice.get("path")
                if old_path and old_path not in final_paths:
                    self.delete_invoice_file_if_orphan(old_path, excluding_record_id=self.editing_record_id)
            self.status_var.set("تم تحديث السجل بنجاح")
            messagebox.showinfo("تم", "تم تحديث السجل بنجاح.")

        self.clear_form()
        self.refresh_summary()
        self.search_records()

    def clear_form(self):
        self.editing_record_id = None
        self.invoice_removed_in_edit = False
        self.main_type_var.set("")
        self.finish_type_var.set("")
        self.payer_var.set("")
        self.item_name_var.set("")
        self.room_var.set(self.room_options[0] if self.room_options else "")
        self.quantity_var.set("1")
        self.unit_price_var.set("0")
        self.total_var.set("0.00")
        self.shared_only_var.set(False)
        self._clear_text_value(self.notes_text)
        self.selected_invoice_sources = []
        self.invoice_name_var.set("لا يوجد ملف مرفق")
        self.status_var.set("تم تجهيز النموذج لإدخال جديد")
        self.update_form_by_type()

    def get_record(self, record_id):
        return self.conn.execute("SELECT * FROM records WHERE id=?", (record_id,)).fetchone()

    def _ensure_search_window(self):
        if self.search_window and self.search_window.winfo_exists():
            return
        self.search_window = tk.Toplevel(self.root)
        self.search_window.title("بحث وإدارة السجلات")
        self.search_window.configure(bg=BG_APP)
        self.search_window.withdraw()
        self.search_window.protocol("WM_DELETE_WINDOW", self.search_window.withdraw)

        card = self._popup_card(self.search_window, "", padx=18, pady=18)
        self._build_search_panel(card)
        self._fit_toplevel(self.search_window, 1120, 720, 0.86, 0.86)

    def open_search_window(self):
        self._ensure_search_window()
        self._reveal_window(self.search_window)
        self.search_records()

    def _build_search_panel(self, parent):
        container = tk.Frame(parent, bg=SURFACE)
        container.pack(fill="both", expand=True)
        tk.Label(container, text="بحث وإدارة السجلات", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 15, "bold"), anchor="e", justify="right").pack(fill="x", pady=(0, 12))

        filters = ttk.Frame(container)
        filters.pack(fill="x")
        for c in range(4):
            filters.columnconfigure(c, weight=1)
        self._add_labeled_entry(filters, "بحث باسم البند أو جزء منه", self.search_var, 0, 3)
        self._add_labeled_combobox(filters, "النوع", self.filter_type_var, ["الكل"] + TYPE_OPTIONS, 0, 2)
        self.filter_room_frame = self._add_labeled_combobox(filters, "المكان", self.filter_room_var, ["الكل"] + self.room_options, 0, 1)
        self.filter_room_combo = self.filter_room_frame.widget
        filter_buttons = tk.Frame(filters, bg=SURFACE)
        filter_buttons.grid(row=0, column=0, sticky="ew", padx=8, pady=8)
        self._action_button(filter_buttons, "بحث", self.search_records, width=78).pack(side="right", padx=4)
        self._action_button(filter_buttons, "إعادة تعيين", self.reset_search, kind="secondary", width=104).pack(side="right", padx=4)
        self._action_button(filter_buttons, "حذف السجل", self.delete_selected_record, kind="danger", width=110).pack(side="left", padx=4)

        results_frame = ttk.Frame(container)
        results_frame.pack(fill="both", expand=True, pady=(14, 10))
        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)

        columns = ("id", "type", "subtype", "payer", "item", "room", "total", "date")
        self.results_tree = ttk.Treeview(results_frame, columns=columns, show="headings", selectmode="browse")
        headings = {
            "id": "م",
            "type": "النوع",
            "subtype": "التصنيف",
            "payer": "الدافع / على",
            "item": "البند",
            "room": "المكان",
            "total": "الإجمالي",
            "date": "آخر تحديث",
        }
        widths = {"id": 60, "type": 90, "subtype": 110, "payer": 100, "item": 260, "room": 140, "total": 110, "date": 150}
        for col in columns:
            self.results_tree.heading(col, text=headings[col], anchor="center")
            self.results_tree.column(col, width=widths[col], anchor="center")
        self._configure_tree_zebra(self.results_tree)
        self.results_tree.grid(row=0, column=0, sticky="nsew")
        self._attach_vertical_scrollbar(self.results_tree, results_frame, manager="grid", row=0, column=1, sticky="ns")

        details = ttk.LabelFrame(container, text="تفاصيل السجل المحدد", padding=12)
        details.pack(fill="x", pady=(0, 10))
        ttk.Label(details, textvariable=self.record_details_var, justify="right", anchor="e", wraplength=1050).pack(fill="x", pady=(0, 12))
        preview_area = ttk.Frame(details)
        preview_area.pack(fill="x")
        ttk.Label(preview_area, textvariable=self.preview_title_var, anchor="e").pack(fill="x")
        self.preview_canvas = tk.Label(preview_area, bg="#f5f7fb", relief="solid", bd=1, width=40, height=12)
        self.preview_canvas.pack(fill="x", pady=(8, 8))
        preview_buttons = tk.Frame(details, bg=SURFACE)
        preview_buttons.pack(fill="x")
        self._action_button(preview_buttons, "فتح المرفق", self.open_selected_record_invoice, width=106).pack(side="right", padx=4)
        self._action_button(preview_buttons, "فتح مكان الملف", self.open_selected_record_invoice_folder, kind="secondary", width=130).pack(side="right", padx=4)
        self._action_button(preview_buttons, "تحميل للتعديل", self.load_selected_record_for_edit, kind="gold", width=130).pack(side="left", padx=4)
        self.results_tree.bind("<<TreeviewSelect>>", self.on_tree_select)
        self.results_tree.bind("<Double-1>", lambda e: self.load_selected_record_for_edit())
        self.preview_canvas.bind("<Button-1>", lambda e: self.open_selected_record_invoice())

    def search_records(self):
        if not hasattr(self, "results_tree"):
            return
        q = "SELECT * FROM records WHERE 1=1"
        params = []
        search_text = self.search_var.get().strip()
        if search_text:
            q += " AND item_name LIKE ?"
            params.append(f"%{search_text}%")
        if self.filter_type_var.get() != "الكل":
            q += " AND main_type = ?"
            params.append(self.filter_type_var.get())
        if self.filter_room_var.get() != "الكل":
            q += " AND room = ?"
            params.append(self.filter_room_var.get())
        q += " ORDER BY updated_at DESC, id DESC"
        self.search_rows = self.conn.execute(q, params).fetchall()
        for item in self.results_tree.get_children():
            self.results_tree.delete(item)
        for idx, row in enumerate(self.search_rows):
            self.results_tree.insert(
                "",
                "end",
                iid=str(row["id"]),
                tags=("oddrow" if idx % 2 == 0 else "evenrow",),
                values=(
                    row["id"],
                    row["main_type"],
                    row["finish_type"] or "-",
                    row["payer"] or "-",
                    row["item_name"],
                    row["room"],
                    f"{row['total']:.2f}",
                    row["updated_at"][:16],
                ),
            )
        if not self.search_rows:
            self.record_details_var.set("لا توجد نتائج مطابقة حاليًا.")
            self.clear_preview()
        else:
            self.record_details_var.set("اختر سجلًا من نتائج البحث لعرض التفاصيل.")

    def reset_search(self):
        self.search_var.set("")
        self.filter_type_var.set("الكل")
        self.filter_room_var.set("الكل")
        self.search_records()

    def on_tree_select(self, event=None):
        if not hasattr(self, "results_tree"):
            return
        selected = self.results_tree.selection()
        if not selected:
            return
        record_id = int(selected[0])
        self.current_selected_record_id = record_id
        row = self.get_record(record_id)
        if not row:
            return
        details = [
            f"النوع: {row['main_type']}",
            f"التصنيف: {row['finish_type'] or '-'}",
            f"الدافع / على: {row['payer'] or '-'}",
            f"البند: {row['item_name']}",
            f"المكان: {row['room']}",
            f"الكمية: {row['quantity']}",
            f"سعر الوحدة: {row['unit_price']:.2f}",
            f"الإجمالي: {row['total']:.2f}",
            f"آخر تحديث: {row['updated_at']}",
            f"ملاحظات: {row['notes'] or '-'}",
        ]
        self.record_details_var.set(" | ".join(details))
        self.show_preview(row)

    def show_preview(self, row):
        invoices = self.get_record_invoices(row["id"])
        if not invoices and row["invoice_path"]:
            invoices = [{"path": row["invoice_path"], "original_name": row["invoice_original_name"]}]
        invoice_count = len(invoices)
        current_invoice = invoices[0] if invoices else None
        path = current_invoice["path"] if current_invoice else None
        if not path:
            self.preview_title_var.set("لا توجد فاتورة مرفقة")
            self.clear_preview("لا يوجد مرفق")
            return
        if not os.path.exists(path):
            self.preview_title_var.set("المرفق غير موجود")
            self.clear_preview("تم نقل الملف أو حذفه")
            return
        if is_image(path) and Image is not None:
            try:
                img = Image.open(path)
                if ImageOps is not None:
                    img = ImageOps.contain(img, THUMB_SIZE)
                else:
                    img.thumbnail(THUMB_SIZE)
                bg = Image.new("RGB", THUMB_SIZE, "#f5f7fb")
                x = (THUMB_SIZE[0] - img.size[0]) // 2
                y = (THUMB_SIZE[1] - img.size[1]) // 2
                bg.paste(img, (x, y))
                self.current_preview_image = ImageTk.PhotoImage(bg)
                self.preview_canvas.configure(image=self.current_preview_image, text="")
                self.preview_title_var.set(f"معاينة صورة الفاتورة: {row['invoice_original_name'] or Path(path).name}")
                return
            except Exception:
                pass
        if is_pdf(path) and Image is not None and ImageDraw is not None:
            card = Image.new("RGB", THUMB_SIZE, "#fff6ee")
            draw = ImageDraw.Draw(card)
            draw.rectangle((30, 30, 250, 190), outline="#c84c0c", width=3)
            draw.text((120, 95), "PDF", fill="#c84c0c")
            self.current_preview_image = ImageTk.PhotoImage(card)
            self.preview_canvas.configure(image=self.current_preview_image, text="")
            self.preview_title_var.set(f"المرفق PDF: {row['invoice_original_name'] or Path(path).name}")
            return
        self.preview_title_var.set(f"مرفق: {row['invoice_original_name'] or Path(path).name}")
        self.clear_preview("اضغط لفتح المرفق")

    def clear_preview(self, text=""):
        if hasattr(self, "preview_canvas"):
            self.current_preview_image = None
            self.preview_canvas.configure(image="", text=text, font=(FONT_FAMILY, 11), fg="#44546a")

    def open_selected_record_invoice(self):
        if not self.current_selected_record_id:
            messagebox.showinfo("معلومة", "اختر سجلًا أولًا.")
            return
        row = self.get_record(self.current_selected_record_id)
        invoices = self.get_record_invoices(self.current_selected_record_id) if row else []
        if not invoices and row and row["invoice_path"]:
            invoices = [{"path": row["invoice_path"], "original_name": row["invoice_original_name"]}]
        if not row or not invoices:
            messagebox.showinfo("معلومة", "لا يوجد مرفق لهذا السجل.")
            return
        for inv in invoices:
            open_file(inv["path"])

    def open_selected_record_invoice_folder(self):
        if not self.current_selected_record_id:
            messagebox.showinfo("معلومة", "اختر سجلًا أولًا.")
            return
        row = self.get_record(self.current_selected_record_id)
        invoices = self.get_record_invoices(self.current_selected_record_id) if row else []
        if not invoices and row and row["invoice_path"]:
            invoices = [{"path": row["invoice_path"], "original_name": row["invoice_original_name"]}]
        path = invoices[0]["path"] if invoices else None
        if not path or not os.path.exists(path):
            messagebox.showinfo("معلومة", "ملف المرفق غير موجود.")
            return
        webbrowser.open(Path(path).parent.resolve().as_uri())

    def load_selected_record_for_edit(self):
        if not hasattr(self, "results_tree"):
            return
        selected = self.results_tree.selection()
        if not selected:
            messagebox.showinfo("معلومة", "اختر سجلًا من نتائج البحث أولًا.")
            return
        record_id = int(selected[0])
        row = self.get_record(record_id)
        if not row:
            return
        self.editing_record_id = record_id
        self.invoice_removed_in_edit = False
        self.main_type_var.set(row["main_type"])
        self.finish_type_var.set(row["finish_type"] or "")
        self.payer_var.set(row["payer"] or self.user_name)
        self.shared_only_var.set(bool(row["shared_split"] == 0))
        self.item_name_var.set(row["item_name"])
        self.room_var.set(row["room"] or (self.room_options[0] if self.room_options else ""))
        self.quantity_var.set(str(row["quantity"]))
        self.unit_price_var.set(str(row["unit_price"]))
        self.total_var.set(f"{row['total']:.2f}")
        self._set_text_value(self.notes_text, row["notes"] or "")
        self.selected_invoice_sources = self.get_record_invoices(record_id)
        if not self.selected_invoice_sources and row["invoice_path"]:
            self.selected_invoice_sources = [{"path": row["invoice_path"], "original_name": row["invoice_original_name"]}]
        self.invoice_name_var.set(format_invoice_names(self.selected_invoice_sources))
        self.update_form_by_type()
        self.status_var.set(f"تم تحميل السجل رقم {record_id} للتعديل")
        self.root.lift()
        self.root.focus_force()

    def delete_selected_record(self):
        if not hasattr(self, "results_tree"):
            return
        selected = self.results_tree.selection()
        if not selected:
            messagebox.showinfo("معلومة", "اختر سجلًا من نتائج البحث أولًا.")
            return
        record_id = int(selected[0])
        row = self.get_record(record_id)
        if not row:
            return
        if not messagebox.askyesno("تأكيد الحذف", "هل أنت متأكد أنك تريد حذف هذا السجل؟"):
            return
        summary = (
            f"النوع: {row['main_type']}\n"
            f"التصنيف: {row['finish_type'] or '-'}\n"
            f"البند: {row['item_name']}\n"
            f"المكان: {row['room']}\n"
            f"الإجمالي: {row['total']:.2f}\n\n"
            "خلي بالك، أنت هتمسح السجل ده نهائيًا.\nهل تريد تأكيد الحذف النهائي؟"
        )
        if not messagebox.askyesno("تأكيد نهائي", summary):
            return
        old_invoices = self.get_record_invoices(record_id)
        if not old_invoices and row["invoice_path"]:
            old_invoices = [{"path": row["invoice_path"], "original_name": row["invoice_original_name"]}]
        self.conn.execute("DELETE FROM records WHERE id=?", (record_id,))
        self.conn.execute("DELETE FROM record_invoices WHERE record_id=?", (record_id,))
        self.conn.commit()
        for old_invoice in old_invoices:
            self.delete_invoice_file_if_orphan(old_invoice.get("path"))
        if self.editing_record_id == record_id:
            self.clear_form()
        self.refresh_summary()
        self.search_records()
        self.record_details_var.set("اختر سجلًا من نتائج البحث لعرض التفاصيل.")
        self.clear_preview()
        self.status_var.set(f"تم حذف السجل رقم {record_id}")
        messagebox.showinfo("تم", "تم حذف السجل بنجاح.")

    def refresh_summary(self):
        records = self.conn.execute("SELECT * FROM records").fetchall()
        total_furn = self.conn.execute("SELECT COALESCE(SUM(total),0) AS s FROM records WHERE main_type='فرش'").fetchone()["s"]
        total_finish = self.conn.execute("SELECT COALESCE(SUM(total),0) AS s FROM records WHERE main_type='تشطيب'").fetchone()["s"]
        total_wedding = self.conn.execute("SELECT COALESCE(SUM(total),0) AS s FROM records WHERE main_type=?", (WEDDING_TYPE,)).fetchone()["s"]
        total_other = self.conn.execute("SELECT COALESCE(SUM(total),0) AS s FROM records WHERE main_type=?", (OTHER_EXPENSES_TYPE,)).fetchone()["s"]
        total_all = total_furn + total_finish + total_wedding + total_other
        shared_furn_me = sum(float(r["total"] or 0) for r in records if r["main_type"] == "فرش" and self.record_shared_split(r) == 1 and (r["payer"] or self.user_name) == self.user_name)
        shared_furn_other = sum(float(r["total"] or 0) for r in records if r["main_type"] == "فرش" and self.record_shared_split(r) == 1 and (r["payer"] or self.user_name) == self.other_party_name)
        shared_furn = sum(float(r["total"] or 0) for r in records if r["main_type"] in ("فرش", WEDDING_TYPE) and (self.record_party1_share(r) or self.record_party2_share(r)))
        solo_furn = self.conn.execute("SELECT COALESCE(SUM(total),0) AS s FROM records WHERE main_type='فرش' AND COALESCE(shared_split,1)=0").fetchone()["s"]
        count_all = self.conn.execute("SELECT COUNT(*) AS c FROM records").fetchone()["c"]
        total_receipts = self.conn.execute("SELECT COALESCE(SUM(amount),0) AS s FROM receipts").fetchone()["s"]
        primary_target = self.payment_targets[0]
        secondary_target = self.payment_targets[1] if len(self.payment_targets) > 1 else None
        total_settlements_primary = self.conn.execute("SELECT COALESCE(SUM(amount),0) AS s FROM advances WHERE direction=?", (primary_target,)).fetchone()["s"]
        total_settlements_secondary = 0
        if secondary_target:
            total_settlements_secondary = self.conn.execute("SELECT COALESCE(SUM(amount),0) AS s FROM advances WHERE direction=?", (secondary_target,)).fetchone()["s"]
        loan_rows = self.conn.execute(
            """
            SELECT l.id, l.amount, l.currency, COALESCE(SUM(p.amount), 0) AS paid
            FROM loans l
            LEFT JOIN loan_payments p ON p.loan_id = l.id AND p.currency = l.currency
            GROUP BY l.id
            """
        ).fetchall()
        loan_remaining_by_currency = {currency: 0.0 for currency in LOAN_CURRENCIES}
        for loan in loan_rows:
            remaining = max(float(loan["amount"] or 0) - float(loan["paid"] or 0), 0.0)
            if loan["currency"] in loan_remaining_by_currency:
                loan_remaining_by_currency[loan["currency"]] += remaining

        person_cost = sum(self.record_party1_share(r) for r in records if r["main_type"] in ("فرش", WEDDING_TYPE))
        groom_cost = sum(self.record_party1_share(r) for r in records)
        if self.accounting_mode == ACCOUNTING_ITEMIZED:
            balance_num = 0.0
        else:
            paid_by_party1_for_party2 = sum(max(0.0, float(r["total"] or 0) - self.record_party1_share(r)) for r in records if r["main_type"] in ("فرش", WEDDING_TYPE) and (r["payer"] or self.user_name) == self.user_name)
            paid_by_party2_for_party1 = sum(self.record_party1_share(r) for r in records if r["main_type"] in ("فرش", WEDDING_TYPE) and (r["payer"] or self.user_name) == self.other_party_name)
            balance_num = total_receipts - total_settlements_primary - paid_by_party1_for_party2 + paid_by_party2_for_party1
        engineer_remaining = total_finish - total_settlements_secondary

        if balance_num > 0:
            balance_text = f"عليّ {balance_num:.2f}"
        elif balance_num < 0:
            balance_text = f"ليّ {abs(balance_num):.2f}"
        else:
            balance_text = "متعادل"

        if engineer_remaining > 0:
            engineer_text = f"عليّ {engineer_remaining:.2f}"
        elif engineer_remaining < 0:
            engineer_text = f"ليّ {abs(engineer_remaining):.2f}"
        else:
            engineer_text = "متعادل"

        self.metric_vars["count"].set(str(count_all))
        self.metric_vars["furn"].set(f"{total_furn:.2f}")
        self.metric_vars["finish"].set(f"{total_finish:.2f}")
        self.metric_vars["wedding"].set(f"{total_wedding:.2f}")
        self.metric_vars["all"].set(f"{total_all:.2f}")
        self.metric_vars["receipts"].set(f"{total_receipts:.2f}")
        self.metric_vars["person_cost"].set(f"{person_cost:.2f}")
        self.metric_vars["balance"].set("حسب توزيع البنود" if self.accounting_mode == ACCOUNTING_ITEMIZED else balance_text)
        self.metric_vars["groom_cost"].set(f"{groom_cost:.2f}")
        self.metric_vars["engineer"].set(engineer_text)
        self.metric_vars["loans"].set(" | ".join(f"{currency}: {loan_remaining_by_currency[currency]:.2f}" for currency in LOAN_CURRENCIES))
        if hasattr(self, "currency_value_labels"):
            for currency, label in self.currency_value_labels.items():
                label.configure(text=f"{loan_remaining_by_currency.get(currency, 0.0):.2f}")
        if hasattr(self, "summary_cards"):
            self.update_summary_visibility()
            for label in self.summary_cards.values():
                try:
                    label.configure(fg="#111827")
                except Exception:
                    pass
            if "balance" in self.summary_cards:
                if self.accounting_mode == ACCOUNTING_ITEMIZED:
                    balance_fg = WARNING
                elif balance_num > 0:
                    balance_fg = DANGER
                elif balance_num < 0:
                    balance_fg = SUCCESS
                else:
                    balance_fg = WARNING
                self.summary_cards["balance"].configure(fg=balance_fg)
            if "engineer" in self.summary_cards:
                if engineer_remaining > 0:
                    engineer_fg = DANGER
                elif engineer_remaining < 0:
                    engineer_fg = SUCCESS
                else:
                    engineer_fg = WARNING
                self.summary_cards["engineer"].configure(fg=engineer_fg)
        self.status_var.set(
            f"جاهز | نظام المحاسبة: {self.accounting_mode} | الفرش: {total_furn:.2f} | حفل الزفاف: {total_wedding:.2f}"
        )
    def _ensure_receipt_form_window(self):
        if self.receipt_form_window and self.receipt_form_window.winfo_exists():
            return
        w = tk.Toplevel(self.root)
        self.receipt_form_window = w
        w.title("إضافة توريد جديد")
        w.configure(bg=BG_APP)
        w.withdraw()
        w.protocol("WM_DELETE_WINDOW", w.withdraw)

        c = self._popup_card(w, "")
        self.receipt_form_title = tk.Label(c, text="إضافة توريد جديد", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 15, "bold"), anchor="e", justify="right")
        self.receipt_form_title.pack(fill="x", pady=(0, 12))
        form = ttk.Frame(c, style="Card.TFrame")
        form.pack(fill="x")
        form.columnconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)

        self._add_labeled_entry(form, "المبلغ", self.receipt_amount_var, 0, 1)
        self._add_labeled_entry(form, "التاريخ", self.receipt_date_var, 0, 0)
        self._add_labeled_combobox(form, "طريقة الاستلام", self.receipt_method_var, PAYMENT_METHODS, 1, 1)
        self._add_labeled_entry(form, "الوقت", self.receipt_time_var, 1, 0)

        actions = tk.Frame(c, bg=SURFACE)
        actions.pack(fill="x", pady=(12, 8))
        self.receipt_save_btn = self._action_button(actions, "حفظ التوريد", self.save_receipt, width=120)
        self.receipt_save_btn.pack(side="right", padx=4)
        self._action_button(actions, "تفريغ", self.clear_receipt_form, kind="danger", width=90).pack(side="right", padx=4)

        notes_frame = ttk.Frame(form, style="Card.TFrame")
        notes_frame.grid(row=2, column=0, columnspan=2, sticky="nsew", padx=8, pady=8)
        ttk.Label(notes_frame, text="ملاحظات", anchor="e", style="Field.TLabel").pack(fill="x")
        self.receipt_notes_text = self._make_notes_widget(notes_frame)
        self.receipt_notes_text.pack(fill="x", pady=(6, 0))
        self._fit_toplevel(w, 760, 680, 0.56, 0.84)

    def open_receipt_form_window(self):
        self._ensure_receipt_form_window()
        self._reveal_window(self.receipt_form_window)

    def validate_receipt(self):
        amount = parse_float(self.receipt_amount_var.get(), None)
        if amount is None or amount <= 0:
            messagebox.showerror("خطأ", "من فضلك أدخل مبلغًا صحيحًا.")
            return None
        try:
            dt = datetime.strptime(f"{self.receipt_date_var.get().strip()} {self.receipt_time_var.get().strip()}", "%Y-%m-%d %H:%M")
        except Exception:
            messagebox.showerror("خطأ", "صيغة التاريخ أو الوقت غير صحيحة. استخدم YYYY-MM-DD و HH:MM")
            return None
        if self.receipt_method_var.get() not in PAYMENT_METHODS:
            messagebox.showerror("خطأ", "اختر طريقة استلام صحيحة.")
            return None
        return amount, dt

    def save_receipt(self):
        result = self.validate_receipt()
        if not result:
            return
        amount, dt = result
        notes = self._get_text_value(self.receipt_notes_text) if hasattr(self, "receipt_notes_text") else ""
        received_at = dt.strftime("%Y-%m-%d %H:%M:%S")
        if self.editing_receipt_id is None:
            self.conn.execute(
                "INSERT INTO receipts (amount, received_at, method, notes, created_at) VALUES (?, ?, ?, ?, ?)",
                (amount, received_at, self.receipt_method_var.get(), notes, datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            )
            msg = "تم حفظ التوريد بنجاح."
            self.status_var.set("تم حفظ التوريد بنجاح")
        else:
            self.conn.execute(
                "UPDATE receipts SET amount=?, received_at=?, method=?, notes=? WHERE id=?",
                (amount, received_at, self.receipt_method_var.get(), notes, self.editing_receipt_id),
            )
            msg = "تم تعديل التوريد بنجاح."
            self.status_var.set("تم تعديل التوريد بنجاح")
        self.conn.commit()
        self.clear_receipt_form()
        self.refresh_receipts_tree()
        self.refresh_summary()
        messagebox.showinfo("تم", msg)

    def clear_receipt_form(self):
        self.editing_receipt_id = None
        self.receipt_amount_var.set("0")
        self.receipt_date_var.set(datetime.now().strftime("%Y-%m-%d"))
        self.receipt_time_var.set(datetime.now().strftime("%H:%M"))
        self.receipt_method_var.set(PAYMENT_METHODS[0])
        if hasattr(self, "receipt_notes_text"):
            self._clear_text_value(self.receipt_notes_text)
        if hasattr(self, "receipt_form_title"):
            self.receipt_form_title.configure(text="إضافة توريد جديد")
        if hasattr(self, "receipt_save_btn"):
            self.receipt_save_btn.configure(text="حفظ التوريد")

    def _ensure_receipts_view_window(self):
        if self.receipts_view_window and self.receipts_view_window.winfo_exists():
            return
        w = tk.Toplevel(self.root)
        self.receipts_view_window = w
        w.title("عرض التوريدات")
        w.configure(bg=BG_APP)
        w.withdraw()
        w.protocol("WM_DELETE_WINDOW", w.withdraw)
        c = self._popup_card(w, "التوريدات المسجلة")
        tree_wrap = ttk.Frame(c)
        tree_wrap.pack(fill="both", expand=True)
        cols = ("id", "amount", "method", "dt", "notes")
        self.receipts_tree = ttk.Treeview(tree_wrap, columns=cols, show="headings", selectmode="browse")
        heads = {"id": "م", "amount": "المبلغ", "method": "الطريقة", "dt": "التاريخ والوقت", "notes": "ملاحظات"}
        widths = {"id": 60, "amount": 120, "method": 120, "dt": 170, "notes": 280}
        for col in cols:
            self.receipts_tree.heading(col, text=heads[col], anchor="center")
            self.receipts_tree.column(col, width=widths[col], anchor="center")
        self._configure_tree_zebra(self.receipts_tree)
        self.receipts_tree.pack(side="right", fill="both", expand=True)
        self._attach_vertical_scrollbar(self.receipts_tree, tree_wrap, side="left", fill="y")
        actions = tk.Frame(c, bg=SURFACE)
        actions.pack(fill="x", pady=(14, 10))
        self._action_button(actions, "حذف التوريد", self.delete_selected_receipt, kind="danger", width=120).pack(side="right", padx=4)
        self._action_button(actions, "تعديل التوريد", self.edit_selected_receipt, kind="gold", width=120).pack(side="right", padx=4)
        self._action_button(actions, "توريد جديد", self.open_receipt_form_window, width=108).pack(side="right", padx=4)
        self._fit_toplevel(w, 920, 520, 0.76, 0.72)

    def open_receipts_view_window(self):
        self._ensure_receipts_view_window()
        self.refresh_receipts_tree()
        self._reveal_window(self.receipts_view_window)


    def edit_selected_receipt(self):
        if not hasattr(self, "receipts_tree"):
            return
        selected = self.receipts_tree.selection()
        if not selected:
            messagebox.showinfo("معلومة", "اختر توريدًا أولًا.")
            return
        receipt_id = int(selected[0].replace("r", ""))
        row = self.conn.execute("SELECT * FROM receipts WHERE id=?", (receipt_id,)).fetchone()
        if not row:
            return
        self._ensure_receipt_form_window()
        self.editing_receipt_id = receipt_id
        self.receipt_amount_var.set(str(row["amount"]))
        dt = datetime.strptime(row["received_at"], "%Y-%m-%d %H:%M:%S")
        self.receipt_date_var.set(dt.strftime("%Y-%m-%d"))
        self.receipt_time_var.set(dt.strftime("%H:%M"))
        self.receipt_method_var.set(row["method"] or PAYMENT_METHODS[0])
        self._set_text_value(self.receipt_notes_text, row["notes"] or "")
        self.receipt_form_title.configure(text="تعديل توريد")
        self.receipt_save_btn.configure(text="حفظ التعديل")
        self._reveal_window(self.receipt_form_window)

    def refresh_receipts_tree(self):
        if not hasattr(self, "receipts_tree"):
            return
        for item in self.receipts_tree.get_children():
            self.receipts_tree.delete(item)
        rows = self.conn.execute("SELECT * FROM receipts ORDER BY received_at DESC, id DESC").fetchall()
        for idx, row in enumerate(rows):
            self.receipts_tree.insert("", "end", iid=f"r{row['id']}", tags=("oddrow" if idx % 2 == 0 else "evenrow",), values=(row["id"], f"{row['amount']:.2f}", row["method"], row["received_at"][:16], row["notes"] or "-"))

    def delete_selected_receipt(self):
        if not hasattr(self, "receipts_tree"):
            return
        selected = self.receipts_tree.selection()
        if not selected:
            messagebox.showinfo("معلومة", "اختر توريدًا أولًا.")
            return
        receipt_id = int(selected[0].replace("r", ""))
        row = self.conn.execute("SELECT * FROM receipts WHERE id=?", (receipt_id,)).fetchone()
        if not row:
            return
        if not messagebox.askyesno("تأكيد الحذف", f"هل تريد حذف هذا التوريد؟\n\nالمبلغ: {row['amount']:.2f}\nالطريقة: {row['method']}\nالتاريخ: {row['received_at'][:16]}"):
            return
        self.conn.execute("DELETE FROM receipts WHERE id=?", (receipt_id,))
        self.conn.commit()
        self.refresh_receipts_tree()
        self.refresh_summary()
        self.status_var.set("تم حذف التوريد")

    # -------- Advances --------
    def _ensure_advance_form_window(self):
        if self.advance_form_window and self.advance_form_window.winfo_exists():
            return
        w = tk.Toplevel(self.root)
        self.advance_form_window = w
        w.title("إضافة تسديد جديد")
        w.configure(bg=BG_APP)
        w.withdraw()
        w.protocol("WM_DELETE_WINDOW", w.withdraw)
        c = self._popup_card(w, "")
        self.advance_form_title = tk.Label(c, text="إضافة تسديد جديد", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 15, "bold"), anchor="e", justify="right")
        self.advance_form_title.pack(fill="x", pady=(0, 12))
        form = ttk.Frame(c, style="Card.TFrame")
        form.pack(fill="x")
        form.columnconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)
        self._add_labeled_entry(form, "المبلغ", self.advance_amount_var, 0, 1)
        self.advance_direction_frame = self._add_labeled_combobox(form, "جهة التسديد", self.advance_direction_var, self.payment_targets, 0, 0)
        self._add_labeled_entry(form, "البيان", self.advance_reason_var, 1, 0, colspan=2)
        self._add_labeled_entry(form, "التاريخ", self.advance_date_var, 2, 1)
        self._add_labeled_entry(form, "وقت الدفع", self.advance_time_var, 2, 0)
        actions = tk.Frame(c, bg=SURFACE)
        actions.pack(fill="x", pady=(12, 8))
        self.advance_save_btn = self._action_button(actions, "حفظ التسديد", self.save_advance, width=120)
        self.advance_save_btn.pack(side="right", padx=4)
        self._action_button(actions, "تفريغ", self.clear_advance_form, kind="danger", width=90).pack(side="right", padx=4)

        notes_frame = ttk.Frame(form, style="Card.TFrame")
        notes_frame.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=8, pady=8)
        ttk.Label(notes_frame, text="ملاحظات", anchor="e").pack(fill="x")
        self.advance_notes_text = self._make_notes_widget(notes_frame)
        self.advance_notes_text.pack(fill="x", pady=(6, 0))
        tk.Label(c, text="اختر جهة التسديد من الجهات التي حددتها في إعدادات البرنامج.", bg="#DDE8D8", fg=TEXT, font=(FONT_FAMILY, 9), anchor="e", justify="right", padx=10, pady=7).pack(fill="x", pady=(8,0))
        self._fit_toplevel(w, 760, 720, 0.56, 0.88)

    def open_advance_form_window(self):
        self._ensure_advance_form_window()
        self._reveal_window(self.advance_form_window)

    def validate_advance(self):
        amount = parse_float(self.advance_amount_var.get(), None)
        if amount is None or amount <= 0:
            messagebox.showerror("خطأ", "من فضلك أدخل مبلغ تسديد صحيحًا.")
            return None
        if not self.advance_reason_var.get().strip():
            messagebox.showerror("خطأ", "من فضلك أدخل بيان التسديد.")
            return None
        try:
            dt = datetime.strptime(f"{self.advance_date_var.get().strip()} {self.advance_time_var.get().strip()}", "%Y-%m-%d %H:%M")
        except Exception:
            messagebox.showerror("خطأ", "صيغة التاريخ أو الوقت غير صحيحة. استخدم YYYY-MM-DD و HH:MM")
            return None
        if self.advance_direction_var.get() not in self.payment_targets:
            messagebox.showerror("خطأ", "اختر جهة تسديد صحيحة.")
            return None
        return amount, dt

    def save_advance(self):
        result = self.validate_advance()
        if not result:
            return
        amount, dt = result
        notes = self._get_text_value(self.advance_notes_text) if hasattr(self, "advance_notes_text") else ""
        dt_str = dt.strftime("%Y-%m-%d %H:%M:%S")
        if self.editing_advance_id is None:
            self.conn.execute(
                "INSERT INTO advances (amount, direction, reason, advance_at, notes, status, settled_at, settlement_method, settlement_notes, created_at) VALUES (?, ?, ?, ?, ?, 'مسددة', ?, ?, ?, ?)",
                (
                    amount,
                    self.advance_direction_var.get(),
                    self.advance_reason_var.get().strip(),
                    dt_str,
                    notes,
                    dt_str,
                    "تحويل/دفع",
                    notes,
                    datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                ),
            )
            msg = "تم حفظ التسديد بنجاح."
            self.status_var.set("تم حفظ التسديد بنجاح")
        else:
            self.conn.execute(
                "UPDATE advances SET amount=?, direction=?, reason=?, advance_at=?, notes=?, settled_at=?, settlement_notes=? WHERE id=?",
                (
                    amount,
                    self.advance_direction_var.get(),
                    self.advance_reason_var.get().strip(),
                    dt_str,
                    notes,
                    dt_str,
                    notes,
                    self.editing_advance_id,
                ),
            )
            msg = "تم تعديل التسديد بنجاح."
            self.status_var.set("تم تعديل التسديد بنجاح")
        self.conn.commit()
        self.clear_advance_form()
        self.refresh_advances_tree()
        self.refresh_summary()
        messagebox.showinfo("تم", msg)
    def clear_advance_form(self):
        self.editing_advance_id = None
        self.advance_amount_var.set("0")
        self.advance_direction_var.set(self.payment_targets[0])
        self.advance_reason_var.set("")
        self.advance_date_var.set(datetime.now().strftime("%Y-%m-%d"))
        self.advance_time_var.set(datetime.now().strftime("%H:%M"))
        if hasattr(self, "advance_notes_text"):
            self._clear_text_value(self.advance_notes_text)
        if hasattr(self, "advance_form_title"):
            self.advance_form_title.configure(text="إضافة تسديد جديد")
        if hasattr(self, "advance_save_btn"):
            self.advance_save_btn.configure(text="حفظ التسديد")

    def _ensure_advances_view_window(self):
        if self.advances_view_window and self.advances_view_window.winfo_exists():
            return
        w = tk.Toplevel(self.root)
        self.advances_view_window = w
        w.title("عرض التسديدات")
        w.configure(bg=BG_APP)
        w.withdraw()
        w.protocol("WM_DELETE_WINDOW", w.withdraw)
        c = self._popup_card(w, "التسديدات المسجلة")
        tree_wrap = ttk.Frame(c)
        tree_wrap.pack(fill="both", expand=True)
        cols = ("id", "amount", "direction", "reason", "advance_at", "notes")
        self.advances_tree = ttk.Treeview(tree_wrap, columns=cols, show="headings", selectmode="browse")
        heads = {
            "id": "م",
            "amount": "المبلغ",
            "direction": "الجهة",
            "reason": "البيان",
            "advance_at": "وقت الدفع",
            "notes": "ملاحظات",
        }
        widths = {"id": 60, "amount": 120, "direction": 160, "reason": 220, "advance_at": 170, "notes": 220}
        for col in cols:
            self.advances_tree.heading(col, text=heads[col], anchor="center")
            self.advances_tree.column(col, width=widths[col], anchor="center")
        self._configure_tree_zebra(self.advances_tree)
        self.advances_tree.pack(side="right", fill="both", expand=True)
        self._attach_vertical_scrollbar(self.advances_tree, tree_wrap, side="left", fill="y")
        actions = tk.Frame(c, bg=SURFACE)
        actions.pack(fill="x", pady=(10, 0))
        self._action_button(actions, "حذف التسديد", self.delete_selected_advance, kind="danger", width=122).pack(side="right", padx=4)
        self._action_button(actions, "تعديل التسديد", self.edit_selected_advance, kind="gold", width=122).pack(side="right", padx=4)
        self._action_button(actions, "تسديد جديد", self.open_advance_form_window, width=110).pack(side="right", padx=4)
        self._fit_toplevel(w, 1020, 620, 0.78, 0.80)

    def open_advances_view_window(self):
        self._ensure_advances_view_window()
        self.refresh_advances_tree()
        self._reveal_window(self.advances_view_window)


    def edit_selected_advance(self):
        row = self._get_selected_advance()
        if not row:
            return
        self._ensure_advance_form_window()
        self.editing_advance_id = row["id"]
        self.advance_amount_var.set(str(row["amount"]))
        dt = datetime.strptime(row["advance_at"], "%Y-%m-%d %H:%M:%S")
        self.advance_date_var.set(dt.strftime("%Y-%m-%d"))
        self.advance_time_var.set(dt.strftime("%H:%M"))
        self.advance_direction_var.set(row["direction"] or self.payment_targets[0])
        self.advance_reason_var.set(row["reason"] or "")
        self._set_text_value(self.advance_notes_text, row["notes"] or "")
        self.advance_form_title.configure(text="تعديل تسديد")
        self.advance_save_btn.configure(text="حفظ التعديل")
        self._reveal_window(self.advance_form_window)

    def refresh_advances_tree(self):
        if not hasattr(self, "advances_tree"):
            return
        for item in self.advances_tree.get_children():
            self.advances_tree.delete(item)
        rows = self.conn.execute("SELECT * FROM advances ORDER BY advance_at DESC, id DESC").fetchall()
        for idx, row in enumerate(rows):
            self.advances_tree.insert(
                "",
                "end",
                iid=f"a{row['id']}",
                tags=("oddrow" if idx % 2 == 0 else "evenrow",),
                values=(row["id"], f"{row['amount']:.2f}", row["direction"], row["reason"], row["advance_at"][:16], (row["notes"] or "")[:40]),
            )
    def _get_selected_advance(self):
        if not hasattr(self, "advances_tree"):
            return None
        selected = self.advances_tree.selection()
        if not selected:
            messagebox.showinfo("معلومة", "اختر تسديدًا أولًا.")
            return None
        advance_id = int(selected[0].replace("a", ""))
        return self.conn.execute("SELECT * FROM advances WHERE id=?", (advance_id,)).fetchone()

    def settle_selected_advance(self):
        messagebox.showinfo("معلومة", "تم استبدال السلف بنظام التسديدات المباشرة.")
    def delete_selected_advance(self):
        row = self._get_selected_advance()
        if not row:
            return
        msg = f"هل تريد حذف هذا التسديد؟\n\nالمبلغ: {row['amount']:.2f}\nالجهة: {row['direction']}\nالبيان: {row['reason']}\nالحالة: {row['status']}"
        if not messagebox.askyesno("تأكيد الحذف", msg):
            return
        self.conn.execute("DELETE FROM advances WHERE id=?", (row["id"],))
        self.conn.commit()
        self.refresh_advances_tree()
        self.refresh_summary()
        self.status_var.set("تم حذف التسديد")

    # -------- Loans --------
    def get_loan_totals(self, loan_id):
        row = self.conn.execute(
            """
            SELECT l.amount AS amount, l.currency AS currency, COALESCE(SUM(p.amount), 0) AS paid
            FROM loans l
            LEFT JOIN loan_payments p ON p.loan_id = l.id AND p.currency = l.currency
            WHERE l.id=?
            GROUP BY l.id
            """,
            (loan_id,),
        ).fetchone()
        if not row:
            return 0.0, 0.0, 0.0, ""
        original = float(row["amount"] or 0)
        paid = float(row["paid"] or 0)
        remaining = max(original - paid, 0.0)
        return original, paid, remaining, row["currency"]

    def _ensure_loan_form_window(self):
        if self.loan_form_window and self.loan_form_window.winfo_exists():
            return
        w = tk.Toplevel(self.root)
        self.loan_form_window = w
        w.title("تسجيل سلفة جديدة")
        w.configure(bg=BG_APP)
        w.withdraw()
        w.protocol("WM_DELETE_WINDOW", w.withdraw)

        c = self._popup_card(w, "")
        self.loan_form_title = tk.Label(c, text="تسجيل سلفة جديدة", bg=SURFACE, fg=PRIMARY, font=(FONT_FAMILY, 15, "bold"), anchor="e", justify="right")
        self.loan_form_title.pack(fill="x", pady=(0, 12))
        form = ttk.Frame(c, style="Card.TFrame")
        form.pack(fill="x")
        form.columnconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)

        self._add_labeled_entry(form, "بيان / جهة السلفة", self.loan_title_var, 0, 0, colspan=2)
        self._add_labeled_entry(form, "قيمة السلفة", self.loan_amount_var, 1, 1)
        self._add_labeled_combobox(form, "العملة", self.loan_currency_var, LOAN_CURRENCIES, 1, 0)
        self._add_labeled_entry(form, "التاريخ", self.loan_date_var, 2, 1)
        self._add_labeled_entry(form, "الوقت", self.loan_time_var, 2, 0)

        notes_frame = ttk.Frame(form, style="Card.TFrame")
        notes_frame.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=8, pady=8)
        ttk.Label(notes_frame, text="ملاحظات", anchor="e", style="Field.TLabel").pack(fill="x")
        self.loan_notes_text = self._make_notes_widget(notes_frame)
        self.loan_notes_text.pack(fill="x", pady=(6, 0))

        actions = tk.Frame(c, bg=SURFACE)
        actions.pack(fill="x", pady=(12, 0))
        self.loan_save_btn = self._action_button(actions, "حفظ السلفة", self.save_loan, width=118)
        self.loan_save_btn.pack(side="right", padx=4)
        self._action_button(actions, "تفريغ", self.clear_loan_form, kind="danger", width=90).pack(side="right", padx=4)
        self._fit_toplevel(w, 760, 740, 0.56, 0.90)

    def open_loan_form_window(self):
        self._ensure_loan_form_window()
        self._reveal_window(self.loan_form_window)

    def validate_loan(self):
        title = self.loan_title_var.get().strip()
        if not title:
            messagebox.showerror("خطأ", "من فضلك أدخل بيان أو جهة السلفة.")
            return None
        amount = parse_float(self.loan_amount_var.get(), None)
        if amount is None or amount <= 0:
            messagebox.showerror("خطأ", "من فضلك أدخل قيمة سلفة صحيحة.")
            return None
        if self.loan_currency_var.get() not in LOAN_CURRENCIES:
            messagebox.showerror("خطأ", "اختر عملة صحيحة.")
            return None
        try:
            dt = datetime.strptime(f"{self.loan_date_var.get().strip()} {self.loan_time_var.get().strip()}", "%Y-%m-%d %H:%M")
        except Exception:
            messagebox.showerror("خطأ", "صيغة التاريخ أو الوقت غير صحيحة. استخدم YYYY-MM-DD و HH:MM")
            return None
        if self.editing_loan_id is not None:
            original, paid, _, currency = self.get_loan_totals(self.editing_loan_id)
            if currency == self.loan_currency_var.get() and amount < paid:
                messagebox.showerror("خطأ", f"لا يمكن جعل أصل السلفة أقل من المسدد بالفعل: {paid:.2f} {currency}.")
                return None
        return title, amount, dt

    def save_loan(self):
        result = self.validate_loan()
        if not result:
            return
        title, amount, dt = result
        notes = self._get_text_value(self.loan_notes_text) if hasattr(self, "loan_notes_text") else ""
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        loan_at = dt.strftime("%Y-%m-%d %H:%M:%S")
        if self.editing_loan_id is None:
            self.conn.execute(
                "INSERT INTO loans (title, amount, currency, loan_at, notes, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?)",
                (title, amount, self.loan_currency_var.get(), loan_at, notes, now, now),
            )
            msg = "تم حفظ السلفة بنجاح."
        else:
            self.conn.execute(
                "UPDATE loans SET title=?, amount=?, currency=?, loan_at=?, notes=?, updated_at=? WHERE id=?",
                (title, amount, self.loan_currency_var.get(), loan_at, notes, now, self.editing_loan_id),
            )
            msg = "تم تعديل السلفة بنجاح."
        self.conn.commit()
        self.clear_loan_form()
        self.refresh_loans_tree()
        self.refresh_summary()
        self.status_var.set(msg.replace(".", ""))
        messagebox.showinfo("تم", msg)

    def clear_loan_form(self):
        self.editing_loan_id = None
        self.loan_title_var.set("")
        self.loan_amount_var.set("0")
        self.loan_currency_var.set(self.base_currency)
        self.loan_date_var.set(datetime.now().strftime("%Y-%m-%d"))
        self.loan_time_var.set(datetime.now().strftime("%H:%M"))
        if hasattr(self, "loan_notes_text"):
            self._clear_text_value(self.loan_notes_text)
        if hasattr(self, "loan_form_title"):
            self.loan_form_title.configure(text="تسجيل سلفة جديدة")
        if hasattr(self, "loan_save_btn"):
            self.loan_save_btn.configure(text="حفظ السلفة")

    def _ensure_loans_view_window(self):
        if self.loans_view_window and self.loans_view_window.winfo_exists():
            return
        w = tk.Toplevel(self.root)
        self.loans_view_window = w
        w.title("إدارة السلف")
        w.configure(bg=BG_APP)
        w.withdraw()
        w.protocol("WM_DELETE_WINDOW", w.withdraw)

        c = self._popup_card(w, "إدارة السلف")

        tree_wrap = ttk.Frame(c)
        tree_wrap.pack(fill="both", expand=True)
        tree_wrap.columnconfigure(0, weight=1)
        tree_wrap.rowconfigure(0, weight=1)
        loan_cols = ("id", "title", "amount", "currency", "paid", "remaining", "loan_at", "notes")
        self.loans_tree = ttk.Treeview(tree_wrap, columns=loan_cols, show="headings", selectmode="browse")
        loan_heads = {
            "id": "م",
            "title": "بيان السلفة",
            "amount": "الأصل",
            "currency": "العملة",
            "paid": "المسدد",
            "remaining": "المتبقي",
            "loan_at": "تاريخ السلفة",
            "notes": "ملاحظات",
        }
        loan_widths = {"id": 55, "title": 220, "amount": 110, "currency": 100, "paid": 110, "remaining": 110, "loan_at": 150, "notes": 220}
        for col in loan_cols:
            self.loans_tree.heading(col, text=loan_heads[col], anchor="center")
            self.loans_tree.column(col, width=loan_widths[col], anchor="center")
        self._configure_tree_zebra(self.loans_tree)
        self.loans_tree.grid(row=0, column=0, sticky="nsew")
        self._attach_vertical_scrollbar(self.loans_tree, tree_wrap, manager="grid", row=0, column=1, sticky="ns")

        actions = tk.Frame(c, bg=SURFACE)
        actions.pack(fill="x", pady=(10, 12))
        self._action_button(actions, "سداد من السلفة", self.open_loan_payment_window_for_selected, width=130).pack(side="right", padx=4)
        self._action_button(actions, "سلفة جديدة", self.open_loan_form_window, width=112).pack(side="right", padx=4)
        self._action_button(actions, "تعديل السلفة", self.edit_selected_loan, kind="gold", width=112).pack(side="right", padx=4)
        self._action_button(actions, "حذف السلفة", self.delete_selected_loan, kind="danger", width=112).pack(side="right", padx=4)

        payments_frame = ttk.LabelFrame(c, text="هيستوري تسديد السلفة المحددة", padding=10)
        payments_frame.pack(fill="both", expand=True)
        pay_cols = ("id", "amount", "currency", "method", "paid_at", "receipt", "notes")
        self.loan_payments_tree = ttk.Treeview(payments_frame, columns=pay_cols, show="headings", selectmode="browse", height=7)
        pay_heads = {"id": "م", "amount": "المبلغ", "currency": "العملة", "method": "طريقة السداد", "paid_at": "وقت السداد", "receipt": "الإيصال", "notes": "ملاحظات"}
        pay_widths = {"id": 55, "amount": 110, "currency": 100, "method": 120, "paid_at": 150, "receipt": 220, "notes": 260}
        for col in pay_cols:
            self.loan_payments_tree.heading(col, text=pay_heads[col], anchor="center")
            self.loan_payments_tree.column(col, width=pay_widths[col], anchor="center")
        self._configure_tree_zebra(self.loan_payments_tree)
        self.loan_payments_tree.pack(side="right", fill="both", expand=True)
        self._attach_vertical_scrollbar(self.loan_payments_tree, payments_frame, side="left", fill="y")

        pay_actions = tk.Frame(c, bg=SURFACE)
        pay_actions.pack(fill="x", pady=(10, 0))
        self._action_button(pay_actions, "فتح إيصال السداد", self.open_selected_loan_payment_receipt, kind="secondary", width=142).pack(side="right", padx=4)
        self._action_button(pay_actions, "حذف السداد", self.delete_selected_loan_payment, kind="danger", width=112).pack(side="right", padx=4)

        self.loans_tree.bind("<<TreeviewSelect>>", lambda e: self.refresh_loan_payments_tree())
        self._fit_toplevel(w, 1160, 720, 0.86, 0.84)

    def open_loans_view_window(self):
        self._ensure_loans_view_window()
        self.refresh_loans_tree()
        self._reveal_window(self.loans_view_window)

    def refresh_loans_tree(self):
        if not hasattr(self, "loans_tree"):
            return
        selected_id = None
        selected = self.loans_tree.selection()
        if selected:
            selected_id = selected[0]
        for item in self.loans_tree.get_children():
            self.loans_tree.delete(item)
        rows = self.conn.execute("SELECT * FROM loans ORDER BY loan_at DESC, id DESC").fetchall()
        for idx, row in enumerate(rows):
            original, paid, remaining, currency = self.get_loan_totals(row["id"])
            self.loans_tree.insert(
                "",
                "end",
                iid=f"l{row['id']}",
                tags=("oddrow" if idx % 2 == 0 else "evenrow",),
                values=(row["id"], row["title"], f"{original:.2f}", currency, f"{paid:.2f}", f"{remaining:.2f}", row["loan_at"][:16], (row["notes"] or "")[:45]),
            )
        if selected_id and self.loans_tree.exists(selected_id):
            self.loans_tree.selection_set(selected_id)
        self.refresh_loan_payments_tree()

    def _get_selected_loan(self, show_message=True):
        if not hasattr(self, "loans_tree"):
            return None
        selected = self.loans_tree.selection()
        if not selected:
            if show_message:
                messagebox.showinfo("معلومة", "اختر سلفة أولًا.")
            return None
        loan_id = int(selected[0].replace("l", ""))
        return self.conn.execute("SELECT * FROM loans WHERE id=?", (loan_id,)).fetchone()

    def edit_selected_loan(self):
        row = self._get_selected_loan()
        if not row:
            return
        self._ensure_loan_form_window()
        self.editing_loan_id = row["id"]
        self.loan_title_var.set(row["title"] or "")
        self.loan_amount_var.set(str(row["amount"]))
        self.loan_currency_var.set(row["currency"] or self.base_currency)
        dt = datetime.strptime(row["loan_at"], "%Y-%m-%d %H:%M:%S")
        self.loan_date_var.set(dt.strftime("%Y-%m-%d"))
        self.loan_time_var.set(dt.strftime("%H:%M"))
        self._set_text_value(self.loan_notes_text, row["notes"] or "")
        self.loan_form_title.configure(text="تعديل سلفة")
        self.loan_save_btn.configure(text="حفظ التعديل")
        self._reveal_window(self.loan_form_window)

    def delete_selected_loan(self):
        row = self._get_selected_loan()
        if not row:
            return
        _, paid, remaining, currency = self.get_loan_totals(row["id"])
        msg = f"هل تريد حذف السلفة بالكامل؟\n\nالبيان: {row['title']}\nالمسدد: {paid:.2f} {currency}\nالمتبقي: {remaining:.2f} {currency}\n\nسيتم حذف هيستوري التسديدات والمرفقات المرتبطة بها من البرنامج."
        if not messagebox.askyesno("تأكيد الحذف", msg):
            return
        receipts = self.conn.execute("SELECT receipt_path FROM loan_payments WHERE loan_id=?", (row["id"],)).fetchall()
        self.conn.execute("DELETE FROM loan_payments WHERE loan_id=?", (row["id"],))
        self.conn.execute("DELETE FROM loans WHERE id=?", (row["id"],))
        self.conn.commit()
        for receipt in receipts:
            self.delete_loan_receipt_file_if_orphan(receipt["receipt_path"])
        self.refresh_loans_tree()
        self.refresh_summary()
        self.status_var.set("تم حذف السلفة")

    def refresh_loan_payments_tree(self):
        if not hasattr(self, "loan_payments_tree"):
            return
        for item in self.loan_payments_tree.get_children():
            self.loan_payments_tree.delete(item)
        row = self._get_selected_loan(show_message=False)
        if not row:
            return
        rows = self.conn.execute("SELECT * FROM loan_payments WHERE loan_id=? ORDER BY paid_at DESC, id DESC", (row["id"],)).fetchall()
        for idx, payment in enumerate(rows):
            receipt_name = payment["receipt_original_name"] or (Path(payment["receipt_path"]).name if payment["receipt_path"] else "-")
            self.loan_payments_tree.insert(
                "",
                "end",
                iid=f"lp{payment['id']}",
                tags=("oddrow" if idx % 2 == 0 else "evenrow",),
                values=(payment["id"], f"{payment['amount']:.2f}", payment["currency"], payment["method"], payment["paid_at"][:16], receipt_name, (payment["notes"] or "")[:55]),
            )

    def _ensure_loan_payment_window(self):
        if self.loan_payment_window and self.loan_payment_window.winfo_exists():
            return
        w = tk.Toplevel(self.root)
        self.loan_payment_window = w
        w.title("سداد سلفة")
        w.configure(bg=BG_APP)
        w.withdraw()
        w.protocol("WM_DELETE_WINDOW", w.withdraw)

        c = self._popup_card(w, "سداد سلفة")
        tk.Label(c, textvariable=self.loan_payment_context_var, bg="#DDE8D8", fg=TEXT, font=(FONT_FAMILY, 9), anchor="e", justify="right", padx=10, pady=7).pack(fill="x", pady=(0, 10))

        form = ttk.Frame(c, style="Card.TFrame")
        form.pack(fill="x")
        form.columnconfigure(0, weight=1)
        form.columnconfigure(1, weight=1)
        self._add_labeled_entry(form, "مبلغ السداد", self.loan_payment_amount_var, 0, 1)
        self._add_labeled_combobox(form, "العملة", self.loan_payment_currency_var, LOAN_CURRENCIES, 0, 0)
        self._add_labeled_combobox(form, "طريقة السداد", self.loan_payment_method_var, LOAN_PAYMENT_METHODS, 1, 1)
        self._add_labeled_entry(form, "التاريخ", self.loan_payment_date_var, 1, 0)
        self._add_labeled_entry(form, "الوقت", self.loan_payment_time_var, 2, 1)

        receipt_frame = ttk.Frame(form, style="Card.TFrame")
        receipt_frame.grid(row=2, column=0, sticky="nsew", padx=8, pady=8)
        ttk.Label(receipt_frame, text="إيصال السداد", anchor="e", style="Field.TLabel").pack(fill="x")
        tk.Label(receipt_frame, textvariable=self.loan_payment_receipt_var, bg=SURFACE, fg=TEXT_2, font=(FONT_FAMILY, 9), anchor="e", justify="right").pack(fill="x", pady=(6, 4))
        receipt_actions = tk.Frame(receipt_frame, bg=SURFACE)
        receipt_actions.pack(fill="x")
        self._action_button(receipt_actions, "إرفاق صورة", self.pick_loan_payment_receipt, width=108, height=34).pack(side="right", padx=3)
        self._action_button(receipt_actions, "إزالة", self.remove_loan_payment_receipt, kind="danger", width=76, height=34).pack(side="right", padx=3)

        notes_frame = ttk.Frame(form, style="Card.TFrame")
        notes_frame.grid(row=3, column=0, columnspan=2, sticky="nsew", padx=8, pady=8)
        ttk.Label(notes_frame, text="ملاحظات", anchor="e", style="Field.TLabel").pack(fill="x")
        self.loan_payment_notes_text = self._make_notes_widget(notes_frame)
        self.loan_payment_notes_text.pack(fill="x", pady=(6, 0))

        actions = tk.Frame(c, bg=SURFACE)
        actions.pack(fill="x", pady=(12, 0))
        self._action_button(actions, "حفظ السداد", self.save_loan_payment, width=118).pack(side="right", padx=4)
        self._action_button(actions, "تفريغ", self.clear_loan_payment_form, kind="danger", width=90).pack(side="right", padx=4)
        self._fit_toplevel(w, 780, 760, 0.58, 0.92)

    def open_loan_payment_window_for_selected(self):
        row = self._get_selected_loan()
        if not row:
            return
        self.open_loan_payment_window(row["id"])

    def open_loan_payment_window(self, loan_id):
        row = self.conn.execute("SELECT * FROM loans WHERE id=?", (loan_id,)).fetchone()
        if not row:
            messagebox.showinfo("معلومة", "السلفة غير موجودة.")
            return
        _, paid, remaining, currency = self.get_loan_totals(loan_id)
        if remaining <= 0:
            messagebox.showinfo("معلومة", "هذه السلفة مسددة بالكامل.")
            return
        self._ensure_loan_payment_window()
        self.current_payment_loan_id = loan_id
        self.current_payment_loan_title = row["title"]
        self.loan_payment_context_var.set(f"{row['title']} | المسدد: {paid:.2f} {currency} | المتبقي: {remaining:.2f} {currency}")
        self.loan_payment_currency_var.set(currency)
        self.loan_payment_amount_var.set(f"{remaining:.2f}")
        self._reveal_window(self.loan_payment_window)

    def clear_loan_payment_form(self):
        self.loan_payment_amount_var.set("0")
        self.loan_payment_method_var.set(LOAN_PAYMENT_METHODS[0])
        self.loan_payment_date_var.set(datetime.now().strftime("%Y-%m-%d"))
        self.loan_payment_time_var.set(datetime.now().strftime("%H:%M"))
        self.selected_loan_payment_receipt_source = None
        self.loan_payment_receipt_var.set("لا يوجد إيصال مرفق")
        if hasattr(self, "loan_payment_notes_text"):
            self._clear_text_value(self.loan_payment_notes_text)
        if self.current_payment_loan_id:
            row = self.conn.execute("SELECT * FROM loans WHERE id=?", (self.current_payment_loan_id,)).fetchone()
            if row:
                _, paid, remaining, currency = self.get_loan_totals(self.current_payment_loan_id)
                self.loan_payment_currency_var.set(currency)
                self.loan_payment_context_var.set(f"{row['title']} | المسدد: {paid:.2f} {currency} | المتبقي: {remaining:.2f} {currency}")

    def pick_loan_payment_receipt(self):
        path = filedialog.askopenfilename(
            title="اختر صورة أو ملف إيصال السداد",
            filetypes=[
                ("الصور وPDF", "*.png *.jpg *.jpeg *.bmp *.gif *.webp *.pdf"),
                ("الصور", "*.png *.jpg *.jpeg *.bmp *.gif *.webp"),
                ("PDF", "*.pdf"),
            ],
        )
        if path:
            if not is_supported_attachment(path):
                messagebox.showwarning("مرفق غير مدعوم", "المرفقات المسموحة: صور أو PDF فقط.")
                return
            if not is_attachment_size_allowed(path):
                messagebox.showwarning("مرفق كبير", "حجم المرفق يجب ألا يزيد عن 100MB.")
                return
            self.selected_loan_payment_receipt_source = path
            self.loan_payment_receipt_var.set(Path(path).name)

    def remove_loan_payment_receipt(self):
        self.selected_loan_payment_receipt_source = None
        self.loan_payment_receipt_var.set("لا يوجد إيصال مرفق")

    def save_loan_payment(self):
        if not self.current_payment_loan_id:
            messagebox.showerror("خطأ", "اختر سلفة أولًا.")
            return
        row = self.conn.execute("SELECT * FROM loans WHERE id=?", (self.current_payment_loan_id,)).fetchone()
        if not row:
            messagebox.showerror("خطأ", "السلفة غير موجودة.")
            return
        amount = parse_float(self.loan_payment_amount_var.get(), None)
        if amount is None or amount <= 0:
            messagebox.showerror("خطأ", "من فضلك أدخل مبلغ سداد صحيح.")
            return
        _, _, remaining, currency = self.get_loan_totals(self.current_payment_loan_id)
        if self.loan_payment_currency_var.get() != currency:
            messagebox.showerror("خطأ", "عملة السداد يجب أن تكون نفس عملة السلفة حتى يتم حساب المتبقي بدقة.")
            return
        if amount > remaining + 0.0001:
            messagebox.showerror("خطأ", f"مبلغ السداد أكبر من المتبقي: {remaining:.2f} {currency}.")
            return
        if self.loan_payment_method_var.get() not in LOAN_PAYMENT_METHODS:
            messagebox.showerror("خطأ", "اختر طريقة سداد صحيحة.")
            return
        try:
            dt = datetime.strptime(f"{self.loan_payment_date_var.get().strip()} {self.loan_payment_time_var.get().strip()}", "%Y-%m-%d %H:%M")
        except Exception:
            messagebox.showerror("خطأ", "صيغة التاريخ أو الوقت غير صحيحة. استخدم YYYY-MM-DD و HH:MM")
            return
        receipt_path = None
        receipt_name = None
        if self.selected_loan_payment_receipt_source:
            receipt_path, receipt_name = copy_loan_receipt_file(self.selected_loan_payment_receipt_source, row["title"])
        notes = self._get_text_value(self.loan_payment_notes_text) if hasattr(self, "loan_payment_notes_text") else ""
        self.conn.execute(
            """
            INSERT INTO loan_payments
            (loan_id, amount, currency, paid_at, method, receipt_path, receipt_original_name, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                self.current_payment_loan_id,
                amount,
                currency,
                dt.strftime("%Y-%m-%d %H:%M:%S"),
                self.loan_payment_method_var.get(),
                receipt_path,
                receipt_name,
                notes,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            ),
        )
        self.conn.commit()
        self.clear_loan_payment_form()
        self.refresh_loans_tree()
        self.refresh_summary()
        messagebox.showinfo("تم", "تم حفظ سداد السلفة بنجاح.")

    def _get_selected_loan_payment(self):
        if not hasattr(self, "loan_payments_tree"):
            return None
        selected = self.loan_payments_tree.selection()
        if not selected:
            messagebox.showinfo("معلومة", "اختر سدادًا من الهيستوري أولًا.")
            return None
        payment_id = int(selected[0].replace("lp", ""))
        return self.conn.execute("SELECT * FROM loan_payments WHERE id=?", (payment_id,)).fetchone()

    def open_selected_loan_payment_receipt(self):
        row = self._get_selected_loan_payment()
        if not row:
            return
        if not row["receipt_path"]:
            messagebox.showinfo("معلومة", "لا يوجد إيصال مرفق لهذا السداد.")
            return
        open_file(row["receipt_path"])

    def delete_selected_loan_payment(self):
        row = self._get_selected_loan_payment()
        if not row:
            return
        if not messagebox.askyesno("تأكيد الحذف", f"هل تريد حذف هذا السداد؟\n\nالمبلغ: {row['amount']:.2f} {row['currency']}\nالتاريخ: {row['paid_at'][:16]}"):
            return
        receipt_path = row["receipt_path"]
        self.conn.execute("DELETE FROM loan_payments WHERE id=?", (row["id"],))
        self.conn.commit()
        self.delete_loan_receipt_file_if_orphan(receipt_path)
        self.refresh_loans_tree()
        self.refresh_summary()
        self.status_var.set("تم حذف سداد السلفة")

    def delete_loan_receipt_file_if_orphan(self, path):
        if not path:
            return
        try:
            resolved = Path(path).resolve()
            resolved.relative_to(LOAN_RECEIPTS_DIR.resolve())
        except Exception:
            return
        count = self.conn.execute("SELECT COUNT(*) AS c FROM loan_payments WHERE receipt_path=?", (path,)).fetchone()["c"]
        if count == 0 and resolved.exists():
            try:
                resolved.unlink()
            except Exception:
                pass

    # -------- Export / PDF --------
    def export_to_excel(self):
        rows = self.conn.execute("SELECT * FROM records ORDER BY main_type, updated_at DESC, id DESC").fetchall()
        receipts = self.conn.execute("SELECT * FROM receipts ORDER BY received_at DESC, id DESC").fetchall()
        advances = self.conn.execute("SELECT * FROM advances ORDER BY advance_at DESC, id DESC").fetchall()
        loans = self.conn.execute("SELECT * FROM loans ORDER BY loan_at DESC, id DESC").fetchall()
        loan_payments = self.conn.execute(
            """
            SELECT p.*, l.title AS loan_title
            FROM loan_payments p
            JOIN loans l ON l.id = p.loan_id
            ORDER BY p.paid_at DESC, p.id DESC
            """
        ).fetchall()
        if not rows and not receipts and not advances and not loans and not loan_payments:
            messagebox.showinfo("معلومة", "لا توجد بيانات لتصديرها.")
            return
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "الفرش"
        ws2 = wb.create_sheet("التشطيب")
        ws3 = wb.create_sheet("حفل الزفاف")
        ws4 = wb.create_sheet(OTHER_EXPENSES_TYPE)
        ws5 = wb.create_sheet("التوريدات")
        ws6 = wb.create_sheet("التسديدات")
        ws7 = wb.create_sheet("السلف")
        ws8 = wb.create_sheet("سداد السلف")
        ws9 = wb.create_sheet("ملخص")

        furn_rows = [r for r in rows if r["main_type"] == "فرش"]
        finish_rows = [r for r in rows if r["main_type"] == "تشطيب"]
        wedding_rows = [r for r in rows if r["main_type"] == WEDDING_TYPE]
        other_rows = [r for r in rows if r["main_type"] == OTHER_EXPENSES_TYPE]

        headers_furn = ["م", "البند", "الدافع", "نوع المشاركة", "المكان", "الكمية", "سعر الوحدة", "الإجمالي", f"نصيب {self.user_name}", "الملاحظات", "اسم الفاتورة", "مسار الفاتورة", "تاريخ الإنشاء", "آخر تحديث"]
        headers_finish = ["م", "تصنيف التشطيب", "البند", "المكان", "الكمية", "سعر الوحدة", "الإجمالي", "الملاحظات", "اسم الفاتورة", "مسار الفاتورة", "تاريخ الإنشاء", "آخر تحديث"]
        headers_wedding = ["م", "البند", "المكان", "الكمية", "سعر الوحدة", "الإجمالي", "الملاحظات", "اسم الفاتورة", "مسار الفاتورة", "تاريخ الإنشاء", "آخر تحديث"]
        headers_other = ["م", "البند", "الكمية", "سعر الوحدة", "الإجمالي", "الملاحظات", "اسم الفاتورة", "مسار الفاتورة", "تاريخ الإنشاء", "آخر تحديث"]

        self._fill_sheet(ws1, headers_furn, furn_rows, mode="furn")
        self._fill_sheet(ws2, headers_finish, finish_rows, mode="finish")
        self._fill_sheet(ws3, headers_wedding, wedding_rows, mode="wedding")
        self._fill_sheet(ws4, headers_other, other_rows, mode="other")
        self._fill_receipts_sheet(ws5, receipts)
        self._fill_advances_sheet(ws6, advances)
        self._fill_loans_sheet(ws7, loans)
        self._fill_loan_payments_sheet(ws8, loan_payments)
        self._fill_summary(ws9, furn_rows, finish_rows, wedding_rows, other_rows, receipts, advances, loans)

        now = datetime.now().strftime("%Y%m%d_%H%M%S")
        output = EXPORTS_DIR / f"marriage_expenses_export_{now}.xlsx"
        wb.save(output)
        self.status_var.set(f"تم التصدير: {output.name}")
        if messagebox.askyesno("تم التصدير", f"تم إنشاء ملف Excel بنجاح:\n{output}\n\nهل تريد فتح الملف الآن؟"):
            open_file(str(output))

    def _fill_sheet(self, ws, headers, rows, mode="furn"):
        ws.sheet_view.rightToLeft = True
        ws.append(headers)
        header_fill = PatternFill("solid", fgColor="12324A")
        header_font = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E1F2")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(bottom=thin)
        for row in rows:
            invoices = self.get_record_invoices(row["id"])
            if not invoices and row["invoice_path"]:
                invoices = [{"path": row["invoice_path"], "original_name": row["invoice_original_name"]}]
            invoice_names = " | ".join(inv.get("original_name") or Path(inv["path"]).name for inv in invoices)
            invoice_paths = " | ".join(inv["path"] for inv in invoices)
            if mode == "finish":
                values = [row["id"], row["finish_type"] or "-", row["item_name"], row["room"], row["quantity"], row["unit_price"], row["total"], row["notes"], invoice_names, invoice_paths, row["created_at"], row["updated_at"]]
            elif mode == "wedding":
                values = [row["id"], row["item_name"], row["room"], row["quantity"], row["unit_price"], row["total"], row["notes"], invoice_names, invoice_paths, row["created_at"], row["updated_at"]]
            elif mode == "other":
                values = [row["id"], row["item_name"], row["quantity"], row["unit_price"], row["total"], row["notes"], invoice_names, invoice_paths, row["created_at"], row["updated_at"]]
            else:
                party1_value = self.record_party1_share(row)
                values = [row["id"], row["item_name"], row["payer"] or self.user_name, self.record_accounting_label(row), row["room"], row["quantity"], row["unit_price"], row["total"], party1_value, row["notes"], invoice_names, invoice_paths, row["created_at"], row["updated_at"]]
            ws.append(values)
        self._beautify_ws(ws)
        if ws.max_row >= 2:
            ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            table_name = {"furn": "FurnTable", "finish": "FinishTable", "wedding": "WeddingTable", "other": "OtherExpensesTable"}[mode]
            table = Table(displayName=table_name, ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)

    def _fill_receipts_sheet(self, ws, receipts):
        ws.sheet_view.rightToLeft = True
        ws.append(["م", "المبلغ", "طريقة الاستلام", "التاريخ والوقت", "ملاحظات", "تاريخ الإضافة"])
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="12324A")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in receipts:
            ws.append([row["id"], row["amount"], row["method"], row["received_at"], row["notes"], row["created_at"]])
        self._beautify_ws(ws)
        if ws.max_row >= 2:
            ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            table = Table(displayName="ReceiptsTable", ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)

    def _fill_advances_sheet(self, ws, advances):
        ws.sheet_view.rightToLeft = True
        ws.append(["م", "المبلغ", "الاتجاه", "البيان", "وقت الدفع", "الحالة", "وقت السداد", "طريقة السداد", "ملاحظات السداد", "ملاحظات", "تاريخ الإضافة"])
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="12324A")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in advances:
            ws.append([row["id"], row["amount"], row["direction"], row["reason"], row["advance_at"], row["status"], row["settled_at"], row["settlement_method"], row["settlement_notes"], row["notes"], row["created_at"]])
        self._beautify_ws(ws)
        if ws.max_row >= 2:
            ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            table = Table(displayName="AdvancesTable", ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)

    def _fill_loans_sheet(self, ws, loans):
        ws.sheet_view.rightToLeft = True
        ws.append(["م", "بيان السلفة", "الأصل", "العملة", "المسدد", "المتبقي", "تاريخ السلفة", "ملاحظات", "تاريخ الإضافة", "آخر تحديث"])
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="12324A")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in loans:
            original, paid, remaining, currency = self.get_loan_totals(row["id"])
            ws.append([row["id"], row["title"], original, currency, paid, remaining, row["loan_at"], row["notes"], row["created_at"], row["updated_at"]])
        self._beautify_ws(ws)
        if ws.max_row >= 2:
            ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            table = Table(displayName="LoansTable", ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)

    def _fill_loan_payments_sheet(self, ws, payments):
        ws.sheet_view.rightToLeft = True
        ws.append(["م", "السلفة", "المبلغ", "العملة", "طريقة السداد", "وقت السداد", "اسم الإيصال", "مسار الإيصال", "ملاحظات", "تاريخ الإضافة"])
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="12324A")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in payments:
            ws.append([row["id"], row["loan_title"], row["amount"], row["currency"], row["method"], row["paid_at"], row["receipt_original_name"], row["receipt_path"], row["notes"], row["created_at"]])
        self._beautify_ws(ws)
        if ws.max_row >= 2:
            ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
            table = Table(displayName="LoanPaymentsTable", ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)

    def _beautify_ws(self, ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for c in col_cells:
                c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                txt = "" if c.value is None else str(c.value)
                max_len = max(max_len, len(txt))
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 34)
        ws.freeze_panes = "A2"

    def _fill_summary(self, ws, furn_rows, finish_rows, wedding_rows, other_rows, receipts, advances, loans):
        ws.sheet_view.rightToLeft = True
        ws["A1"] = "البيان"
        ws["B1"] = "القيمة"
        total_furn = sum(r["total"] for r in furn_rows)
        total_finish = sum(r["total"] for r in finish_rows)
        total_wedding = sum(r["total"] for r in wedding_rows)
        total_other = sum(r["total"] for r in other_rows)
        all_rows = list(furn_rows) + list(finish_rows) + list(wedding_rows) + list(other_rows)
        shared_furn_me = sum(r["total"] for r in furn_rows if self.record_shared_split(r) == 1 and (r["payer"] or self.user_name) == self.user_name)
        shared_furn_other = sum(r["total"] for r in furn_rows if self.record_shared_split(r) == 1 and (r["payer"] or self.user_name) == self.other_party_name)
        shared_furn = sum(self.record_party1_share(r) + self.record_party2_share(r) for r in furn_rows)
        solo_furn = sum(r["total"] for r in furn_rows if self.record_shared_split(r) == 0)
        total_all = total_furn + total_finish + total_wedding + total_other
        total_receipts = sum(r["amount"] for r in receipts)
        person_cost = sum(self.record_party1_share(r) for r in list(furn_rows) + list(wedding_rows))
        primary_target = self.payment_targets[0]
        secondary_target = self.payment_targets[1] if len(self.payment_targets) > 1 else None
        primary_payments = sum(a["amount"] for a in advances if a["direction"] == primary_target)
        secondary_payments = sum(a["amount"] for a in advances if secondary_target and a["direction"] == secondary_target)
        loan_remaining_by_currency = {currency: 0.0 for currency in LOAN_CURRENCIES}
        for loan in loans:
            _, _, remaining, currency = self.get_loan_totals(loan["id"])
            if currency in loan_remaining_by_currency:
                loan_remaining_by_currency[currency] += remaining
        if self.accounting_mode == ACCOUNTING_ITEMIZED:
            balance_num = 0.0
        else:
            paid_by_party1_for_party2 = sum(max(0.0, r["total"] - self.record_party1_share(r)) for r in list(furn_rows) + list(wedding_rows) if (r["payer"] or self.user_name) == self.user_name)
            paid_by_party2_for_party1 = sum(self.record_party1_share(r) for r in list(furn_rows) + list(wedding_rows) if (r["payer"] or self.user_name) == self.other_party_name)
            balance_num = total_receipts - primary_payments - paid_by_party1_for_party2 + paid_by_party2_for_party1
        groom_cost = sum(self.record_party1_share(r) for r in all_rows)
        engineer_remaining = total_finish - secondary_payments
        summary_data = [
            ("عدد بنود الفرش", len(furn_rows)),
            ("عدد بنود التشطيب", len(finish_rows)),
            ("عدد بنود حفل الزفاف", len(wedding_rows)),
            ("إجمالي الفرش", total_furn),
            ("الفرش المشترك", shared_furn),
            (f"فرش مشترك دفعه {self.user_name}", shared_furn_me),
            (f"فرش مشترك دفعه {self.other_party_name}", shared_furn_other),
            ("فرش على الطرف الأول بالكامل", solo_furn),
            ("إجمالي التشطيب", total_finish),
            ("إجمالي حفل الزفاف", total_wedding),
            ("الإجمالي الكلي", total_all),
            ("إجمالي التوريدات", total_receipts),
            ("التكاليف للشخص", person_cost),
            ("الموازنة", balance_num),
            ("اللي دفعته", groom_cost),
            (f"تسديدات {primary_target}", primary_payments),
            (f"تسديدات {secondary_target or 'جهة التسديد الثانية'}", secondary_payments),
            (f"المتبقي لـ {secondary_target or 'جهة التسديد الثانية'}", engineer_remaining),
        ]
        summary_data.extend((f"السلف المتبقية - {currency}", loan_remaining_by_currency[currency]) for currency in LOAN_CURRENCIES)
        for i, (label, value) in enumerate(summary_data, start=2):
            ws[f"A{i}"] = label
            ws[f"B{i}"] = value
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="12324A")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18

    def choose_print_mode(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("اختيار نوع القائمة")
        dialog.configure(bg=BG_APP)
        dialog.transient(self.root)
        dialog.grab_set()
        result = {"mode": None}
        outer = self._popup_card(dialog, "اختر نوع القائمة المطلوبة", padx=16, pady=16)
        self._action_button(outer, f"بنسبة {self.user_name}", lambda: (result.update(mode="half"), dialog.destroy()), width=250).pack(fill="x", pady=4)
        self._action_button(outer, "كاملة", lambda: (result.update(mode="full"), dialog.destroy()), width=250).pack(fill="x", pady=4)
        self._action_button(outer, "إلغاء", dialog.destroy, kind="secondary", width=250).pack(fill="x", pady=(10, 0))
        self._fit_toplevel(dialog, 360, 260, 0.28, 0.30)
        self.root.wait_window(dialog)
        return result["mode"]

    def open_reports_dialog(self):
        dialog = tk.Toplevel(self.root)
        dialog.title("تقارير PDF")
        dialog.configure(bg=BG_APP)
        dialog.transient(self.root)
        dialog.grab_set()
        outer = self._popup_card(dialog, "اختر التقرير المطلوب", padx=16, pady=16)
        report_buttons = [
            ("تقرير الفرش", "furniture"),
            ("تقرير التسديدات", "payments"),
            ("تقرير السلف", "loans"),
            ("تقرير التشطيبات", "finishing"),
        ]
        for title, report_key in report_buttons:
            self._action_button(
                outer,
                title,
                lambda key=report_key, w=dialog: (w.destroy(), self.generate_pdf_report(key)),
                width=270,
                height=40,
            ).pack(fill="x", pady=4)
        self._action_button(outer, "إلغاء", dialog.destroy, kind="secondary", width=270, height=40).pack(fill="x", pady=(10, 0))
        self._fit_toplevel(dialog, 380, 360, 0.30, 0.42)
        self.root.wait_window(dialog)

    def generate_pdf_report(self, report_key):
        pdf_libs = maybe_import_pdf_libs()
        if not pdf_libs:
            messagebox.showerror(
                "مكتبات ناقصة",
                "التقارير PDF تحتاج تثبيت مكتبات إضافية مرة واحدة:\n\npy -m pip install reportlab arabic-reshaper python-bidi",
            )
            return
        font_path = find_arabic_font_path()
        if not font_path:
            messagebox.showerror("خطأ", "لم أتمكن من العثور على خط عربي مناسب لإنشاء PDF.")
            return
        try:
            spec = self._build_report_spec(report_key)
            if not spec["rows"]:
                messagebox.showinfo("معلومة", f"لا توجد بيانات في {spec['title']}.")
                return
            now = datetime.now()
            output = EXPORTS_DIR / f"{sanitize_filename(spec['title'])}_{now.strftime('%Y%m%d_%H%M%S')}.pdf"
            self._create_report_pdf(output, spec, pdf_libs, font_path, now)
            self.status_var.set(f"تم حفظ التقرير: {output.name}")
            if messagebox.askyesno("تم", f"تم إنشاء التقرير بنجاح:\n{output}\n\nهل تريد فتح الملف الآن؟"):
                open_file(str(output))
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء إنشاء التقرير:\n{e}")

    def _build_report_spec(self, report_key):
        if report_key == "furniture":
            rows = self.conn.execute(
                """
                SELECT id, item_name, room, quantity, unit_price, total, payer, shared_split, notes, updated_at
                FROM records
                WHERE main_type='فرش'
                ORDER BY updated_at DESC, id DESC
                """
            ).fetchall()
            data = []
            for idx, row in enumerate(rows, start=1):
                data.append([
                    idx,
                    row["item_name"],
                    row["room"] or "-",
                    self._fmt_num(row["quantity"]),
                    self._fmt_money(row["unit_price"]),
                    self._fmt_money(row["total"]),
                    row["payer"] or self.user_name,
                    self.record_accounting_label(row),
                ])
            total = sum(row["total"] for row in rows)
            return {
                "title": "تقرير الفرش",
                "columns": [
                    ("م", 30), ("البند", 178), ("المكان", 82), ("العدد", 50),
                    ("سعر الوحدة", 78), ("الإجمالي", 82), ("الدافع", 74), ("المشاركة", 82),
                ],
                "rows": data,
                "summary": [("إجمالي الفرش", self._fmt_money(total)), ("عدد البنود", str(len(rows)))],
            }
        if report_key == "payments":
            rows = self.conn.execute(
                """
                SELECT id, amount, direction, reason, advance_at, notes
                FROM advances
                ORDER BY advance_at DESC, id DESC
                """
            ).fetchall()
            data = [[idx, self._fmt_money(row["amount"]), row["direction"], row["reason"], row["advance_at"][:16], row["notes"] or "-"] for idx, row in enumerate(rows, start=1)]
            total = sum(row["amount"] for row in rows)
            primary_target = self.payment_targets[0]
            secondary_target = self.payment_targets[1] if len(self.payment_targets) > 1 else None
            primary = sum(row["amount"] for row in rows if row["direction"] == primary_target)
            secondary = sum(row["amount"] for row in rows if secondary_target and row["direction"] == secondary_target)
            return {
                "title": "تقرير التسديدات",
                "columns": [("م", 34), ("المبلغ", 90), ("الجهة", 130), ("البيان", 220), ("وقت الدفع", 110), ("ملاحظات", 204)],
                "rows": data,
                "summary": [
                    ("إجمالي التسديدات", self._fmt_money(total)),
                    (f"تسديدات {primary_target}", self._fmt_money(primary)),
                    (f"تسديدات {secondary_target or 'جهة التسديد الثانية'}", self._fmt_money(secondary)),
                ],
            }
        if report_key == "loans":
            rows = self.conn.execute("SELECT * FROM loans ORDER BY loan_at DESC, id DESC").fetchall()
            data = []
            remaining_by_currency = {currency: 0.0 for currency in LOAN_CURRENCIES}
            original_by_currency = {currency: 0.0 for currency in LOAN_CURRENCIES}
            paid_by_currency = {currency: 0.0 for currency in LOAN_CURRENCIES}
            for idx, row in enumerate(rows, start=1):
                original, paid, remaining, currency = self.get_loan_totals(row["id"])
                if currency in remaining_by_currency:
                    original_by_currency[currency] += original
                    paid_by_currency[currency] += paid
                    remaining_by_currency[currency] += remaining
                data.append([idx, row["title"], self._fmt_money(original), currency, self._fmt_money(paid), self._fmt_money(remaining), row["loan_at"][:16], row["notes"] or "-"])
            summary = []
            for currency in LOAN_CURRENCIES:
                summary.append((f"الأصل - {currency}", self._fmt_money(original_by_currency[currency])))
                summary.append((f"المسدد - {currency}", self._fmt_money(paid_by_currency[currency])))
                summary.append((f"المتبقي - {currency}", self._fmt_money(remaining_by_currency[currency])))
            return {
                "title": "تقرير السلف",
                "columns": [
                    ("م", 30), ("بيان السلفة", 190), ("الأصل", 76), ("العملة", 64),
                    ("المسدد", 76), ("المتبقي", 76), ("تاريخ السلفة", 110), ("ملاحظات", 176),
                ],
                "rows": data,
                "summary": summary,
            }
        if report_key == "finishing":
            rows = self.conn.execute(
                """
                SELECT id, finish_type, item_name, room, quantity, unit_price, total, notes, updated_at
                FROM records
                WHERE main_type='تشطيب'
                ORDER BY updated_at DESC, id DESC
                """
            ).fetchall()
            data = []
            totals_by_type = {}
            for idx, row in enumerate(rows, start=1):
                finish_type = row["finish_type"] or "-"
                totals_by_type[finish_type] = totals_by_type.get(finish_type, 0.0) + row["total"]
                data.append([
                    idx,
                    finish_type,
                    row["item_name"],
                    row["room"] or "-",
                    self._fmt_num(row["quantity"]),
                    self._fmt_money(row["unit_price"]),
                    self._fmt_money(row["total"]),
                    row["notes"] or "-",
                ])
            total = sum(row["total"] for row in rows)
            summary = [("إجمالي التشطيبات", self._fmt_money(total)), ("عدد البنود", str(len(rows)))]
            summary.extend((f"إجمالي {key}", self._fmt_money(value)) for key, value in sorted(totals_by_type.items()))
            return {
                "title": "تقرير التشطيبات",
                "columns": [
                    ("م", 30), ("التصنيف", 92), ("البند", 174), ("المكان", 76),
                    ("العدد", 48), ("سعر الوحدة", 78), ("الإجمالي", 82), ("ملاحظات", 190),
                ],
                "rows": data,
                "summary": summary,
            }
        raise ValueError("نوع التقرير غير معروف")

    def _fmt_money(self, value):
        return f"{float(value or 0):,.2f}"

    def _fmt_num(self, value):
        value = float(value or 0)
        return str(int(value)) if value.is_integer() else f"{value:.2f}"

    def _fit_text_for_pdf(self, text, width, font_size):
        text = str(text if text is not None else "-").replace("\n", " ").strip() or "-"
        max_chars = max(4, int(width / max(font_size * 0.46, 1)))
        if len(text) > max_chars:
            text = text[: max_chars - 3].rstrip() + "..."
        return text

    def _create_report_pdf(self, output, spec, pdf_libs, font_path, created_at):
        A4 = pdf_libs["A4"]
        pdfmetrics = pdf_libs["pdfmetrics"]
        TTFont = pdf_libs["TTFont"]
        canvas = pdf_libs["canvas"]
        font_name = "ArabicFont"
        try:
            pdfmetrics.getFont(font_name)
        except Exception:
            pdfmetrics.registerFont(TTFont(font_name, font_path))

        page_size = (A4[1], A4[0])
        c = canvas.Canvas(str(output), pagesize=page_size)
        width, height = page_size
        margins = {"left": 28, "right": 28, "top": 28, "bottom": 32}
        table_width = width - margins["left"] - margins["right"]
        raw_columns = spec["columns"]
        total_col_width = sum(col_w for _, col_w in raw_columns) or table_width
        scale = table_width / total_col_width
        columns = [(title, col_w * scale) for title, col_w in raw_columns]
        row_h = 22
        header_h = 24
        page_no = 0

        def draw_rtl(text, x_right, y, size=10, color=(0, 0, 0)):
            c.setFont(font_name, size)
            c.setFillColorRGB(*color)
            c.drawRightString(x_right, y, shape_ar(text, pdf_libs))

        def draw_center(text, y, size=16, color=(0, 0, 0)):
            c.setFont(font_name, size)
            c.setFillColorRGB(*color)
            c.drawCentredString(width / 2, y, shape_ar(text, pdf_libs))

        def draw_page_header():
            nonlocal page_no
            page_no += 1
            c.setFillColorRGB(0.07, 0.20, 0.29)
            c.rect(margins["left"], height - 66, table_width, 34, fill=1, stroke=0)
            draw_center(spec["title"], height - 54, 16, (1, 1, 1))
            draw_rtl(f"تاريخ التقرير: {created_at.strftime('%Y-%m-%d %H:%M')}", width - margins["right"], height - 84, 9, (0.35, 0.40, 0.48))
            draw_rtl(f"عدد الصفوف: {len(spec['rows'])}", width - margins["right"] - 210, height - 84, 9, (0.35, 0.40, 0.48))
            draw_center(f"صفحة {page_no}", 18, 9, (0.35, 0.40, 0.48))
            return height - 112

        def draw_table_header(y):
            c.setFillColorRGB(0.07, 0.20, 0.29)
            c.rect(margins["left"], y - header_h + 6, table_width, header_h, fill=1, stroke=0)
            x = width - margins["right"]
            for title, col_w in columns:
                draw_rtl(self._fit_text_for_pdf(title, col_w - 8, 9), x - 5, y - 9, 9, (1, 1, 1))
                x -= col_w
            return y - header_h

        y = draw_page_header()
        y = draw_table_header(y)
        for idx, row in enumerate(spec["rows"], start=1):
            if y < margins["bottom"] + 82:
                c.showPage()
                y = draw_page_header()
                y = draw_table_header(y)
            c.setFillColorRGB(0.96, 0.98, 1.0) if idx % 2 == 0 else c.setFillColorRGB(1, 1, 1)
            c.rect(margins["left"], y - row_h + 6, table_width, row_h, fill=1, stroke=0)
            c.setStrokeColorRGB(0.86, 0.89, 0.93)
            c.line(margins["left"], y - row_h + 6, width - margins["right"], y - row_h + 6)
            x = width - margins["right"]
            for value, (_, col_w) in zip(row, columns):
                draw_rtl(self._fit_text_for_pdf(value, col_w - 8, 8.5), x - 5, y - 8, 8.5)
                x -= col_w
            y -= row_h

        if y < margins["bottom"] + 90:
            c.showPage()
            y = draw_page_header()
        y -= 10
        c.setStrokeColorRGB(0.82, 0.70, 0.22)
        c.line(margins["left"], y, width - margins["right"], y)
        y -= 24
        draw_rtl("ملخص التقرير", width - margins["right"], y, 12, (0.07, 0.20, 0.29))
        y -= 20
        for label, value in spec["summary"]:
            if y < margins["bottom"] + 24:
                c.showPage()
                y = draw_page_header()
            draw_rtl(f"{label}: {value}", width - margins["right"], y, 10)
            y -= 17
        c.save()

    def print_furniture_list(self):
        rows = self.conn.execute("SELECT * FROM records WHERE main_type='فرش' ORDER BY updated_at DESC, id DESC").fetchall()
        if not rows:
            messagebox.showinfo("معلومة", "لا توجد بنود فرش لطباعة القائمة.")
            return
        mode = self.choose_print_mode()
        if not mode:
            return
        pdf_libs = maybe_import_pdf_libs()
        if not pdf_libs:
            messagebox.showerror(
                "مكتبات ناقصة",
                "ميزة حفظ القائمة PDF تحتاج تثبيت مكتبات إضافية مرة واحدة:\n\npy -m pip install reportlab arabic-reshaper python-bidi",
            )
            return
        font_path = find_arabic_font_path()
        if not font_path:
            messagebox.showerror("خطأ", "لم أتمكن من العثور على خط عربي مناسب لإنشاء PDF.")
            return
        default_name = f"قائمة_منقولات_زوجية_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        output = filedialog.asksaveasfilename(
            title="حفظ القائمة PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
            initialdir=str(EXPORTS_DIR),
            initialfile=default_name,
        )
        if not output:
            return
        try:
            self._create_furniture_pdf(output, rows, mode, pdf_libs, font_path)
            self.status_var.set(f"تم حفظ القائمة: {Path(output).name}")
            if messagebox.askyesno("تم", f"تم حفظ القائمة بنجاح:\n{output}\n\nهل تريد فتح الملف الآن؟"):
                open_file(output)
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء إنشاء PDF:\n{e}")

    def _create_furniture_pdf(self, output, rows, mode, pdf_libs, font_path):
        A4 = pdf_libs["A4"]
        pdfmetrics = pdf_libs["pdfmetrics"]
        TTFont = pdf_libs["TTFont"]
        canvas = pdf_libs["canvas"]
        font_name = "ArabicFont"
        try:
            pdfmetrics.getFont(font_name)
        except Exception:
            pdfmetrics.registerFont(TTFont(font_name, font_path))
        c = canvas.Canvas(output, pagesize=A4)
        width, height = A4
        page_no = 0

        def draw_rtl(text, x_right, y, size=12):
            c.setFont(font_name, size)
            shaped = shape_ar(text, pdf_libs)
            c.drawRightString(x_right, y, shaped)

        def draw_center(text, y, size=16):
            c.setFont(font_name, size)
            shaped = shape_ar(text, pdf_libs)
            c.drawCentredString(width / 2, y, shaped)

        def new_page():
            nonlocal page_no
            page_no += 1
            c.setFillColorRGB(0.07, 0.20, 0.29)
            c.rect(30, height - 78, width - 60, 38, fill=1, stroke=0)
            c.setFillColorRGB(1, 1, 1)
            draw_center("قائمة منقولات زوجية", height - 63, 18)
            c.setFillColorRGB(0, 0, 0)
            draw_center(f"قيمة {self.user_name} حسب النسبة" if mode == "half" else "القيمة كاملة", height - 98, 11)
            draw_rtl(f"تاريخ التصدير: {datetime.now().strftime('%Y-%m-%d %H:%M')}", width - 45, height - 118, 10)
            draw_center(f"الصفحة {page_no}", 24, 10)
            c.line(40, height - 128, width - 40, height - 128)
            return height - 156

        y = new_page()
        total = 0.0
        row_height = 24
        for idx, row in enumerate(rows, start=1):
            amount = self.record_party1_share(row) if mode == "half" else row["total"]
            total += amount
            if y < 80:
                c.showPage()
                y = new_page()
            if idx % 2 == 0:
                c.setFillColorRGB(0.95, 0.97, 0.99)
            else:
                c.setFillColorRGB(1, 1, 1)
            c.rect(40, y - 6, width - 80, row_height, fill=1, stroke=0)
            c.setFillColorRGB(0, 0, 0)
            draw_rtl(f"{amount:.2f}", 140, y + 8, 11)
            draw_rtl(row["item_name"], width - 55, y + 8, 11)
            draw_rtl(str(idx), width - 20, y + 8, 10)
            y -= row_height
        if y < 90:
            c.showPage()
            y = new_page()
        c.line(40, y, width - 40, y)
        y -= 24
        draw_rtl(f"الإجمالي: {total:.2f}", width - 45, y, 14)
        c.save()

    def on_close(self):
        if hasattr(self, "conn"):
            self.conn.close()
        if self._root_exists():
            self.root.destroy()


def main():
    try:
        ensure_dirs()
        if os.name == "nt":
            try:
                ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(APP_USER_MODEL_ID)
            except Exception:
                pass
        root = tk.Tk()
        if TkinterDnD is not None:
            try:
                TkinterDnD._require(root)
            except Exception:
                pass
        app = ApartmentCostsApp(root)
        try:
            root_exists = bool(root.winfo_exists())
        except tk.TclError:
            root_exists = False
        if root_exists:
            root.protocol("WM_DELETE_WINDOW", app.on_close)
            root.mainloop()
    except Exception as e:
        try:
            ensure_dirs()
            from traceback import format_exc
            log_path = DATA_DIR / "startup_error.log"
            with open(log_path, "a", encoding="utf-8") as f:
                f.write("\n" + "="*70 + "\n")
                f.write(datetime.now().strftime("%Y-%m-%d %H:%M:%S") + "\n")
                f.write(format_exc())
        except Exception:
            pass
        try:
            root = tk.Tk()
            root.withdraw()
            messagebox.showerror("خطأ في بدء التشغيل", f"تعذر فتح البرنامج.\n\nتم حفظ سبب الخطأ في:\n{DATA_DIR / 'startup_error.log'}")
            root.destroy()
        except Exception:
            pass
        raise


if __name__ == "__main__":
    main()

